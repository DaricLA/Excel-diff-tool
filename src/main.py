import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import threading, time, os, sys, json, zipfile, re, copy, colorsys, random
import pythoncom, win32com.client
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import MultiCellRange
from lxml import etree

NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
PROGRAM_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
VERSION = "v3.94"

DEFAULT_CHECK_OPTIONS = {
    'value': True, 'formula': True, 'rich_text': True, 'font': True,
    'fill': True, 'border': True, 'alignment': True, 'number_format': True,
    'merged_cells': True, 'row_height': True, 'col_width': True,
    'images': True, 'conditional_format': True
}
CHECK_OPTION_LABELS = {
    'value': '值变化', 'formula': '公式变化', 'rich_text': '富文本',
    'font': '字体', 'fill': '填充/背景色', 'border': '边框',
    'alignment': '对齐方式', 'number_format': '数字格式',
    'merged_cells': '合并单元格', 'row_height': '行高', 'col_width': '列宽',
    'images': '图片', 'conditional_format': '条件格式'
}
CHECK_OPTION_GROUPS = [
    ("内容检测", ['value', 'formula', 'rich_text']),
    ("格式检测", ['font', 'fill', 'border', 'alignment', 'number_format']),
    ("结构检测", ['merged_cells', 'row_height', 'col_width', 'images', 'conditional_format'])
]

# ECMA-376 默认索引色调色板（indexedColors 缺省值）
DEFAULT_INDEXED = [
    '000000','FFFFFF','FF0000','00FF00','0000FF','FFFF00','FF00FF','00FFFF',
    '000000','FFFFFF','FF0000','00FF00','0000FF','FFFF00','FF00FF','00FFFF',
    '800000','008000','000080','808000','800080','008080','C0C0C0','808080',
    '808080','FF9999','663300','FFFF99','990000','999900','009900','990099',
    '660099','3366FF','CCCCFF','0066CC','FF00FF','FFFF00','00FFFF','FF00FF',
    'FF00FF','FFFF00','00FFFF','66FFFF','00CCFF','CCFFFF','006600','666600',
    '99CC00','00CC99','CC99FF','993366','FF99CC','CC99FF','FFCC00','336600',
    '993300','993399','333399','339999','333333','666666','999999','CCCCCC',
]
# ECMA-376 内置数字格式（numFmtId < 164）
BUILTIN_NUMFMTS = {
    0:'General',1:'0',2:'0.00',3:'#,##0',4:'#,##0.00',9:'0%',10:'0.00%',
    11:'0.00E+00',12:'?/?',13:'??/??',14:'mm-dd-yy',15:'d-mmm-yy',16:'d-mmm',
    17:'mmm-yy',18:'h:mm AM/PM',19:'h:mm:ss AM/PM',20:'h:mm',21:'h:mm:ss',
    22:'m/d/yy h:mm',37:'#,##0 ;(#,##0)',38:'#,##0 ;[Red](#,##0)',
    39:'#,##0.00;(#,##0.00)',40:'#,##0.00;[Red](#,##0.00)',
    45:'mm:ss',46:'[h]:mm:ss',47:'mmss.0',48:'##0.0E+0',49:'@',
}
# 常见数字格式 -> 可读描述（覆盖 Excel 下拉菜单常见格式）
_NUMFMT_READABLE = {
    'general': '常规', '@': '文本',
    '0': '整数', '0.00': '数字(2位小数)',
    '#,##0': '千分位整数', '#,##0.00': '千分位2位小数',
    '0%': '百分比', '0.00%': '百分比2位小数',
    '¥#,##0': '人民币整数', '¥#,##0.00': '人民币2位小数',
    '$#,##0': '美元整数', '$#,##0.00': '美元2位小数',
    'yyyy-mm-dd': '日期(年-月-日)', 'yyyy/m/d': '日期(年/月/日)',
    'm/d/yyyy': '日期(月/日/年)', 'd/m/yyyy': '日期(日/月/年)',
    'h:mm': '时间(时:分)', 'h:mm:ss': '时间(时:分:秒)',
}

def format_numfmt_readable(fmt):
    """数字格式码 -> 可读描述（常见格式转中文，自定义格式去掉语言前缀后如实显示）；模块级，供两类比对共用"""
    if not fmt: return '常规'
    s = re.sub(r'\[\$-[^\]]*\]', '', str(fmt)).strip()  # 去掉 [$-441] 等语言前缀
    return _NUMFMT_READABLE.get(s.lower(), s)  # 常见格式转中文；自定义格式去前缀后如实显示

def excel_tint(hex6, tint):
    """ECMA-376 主题色 tint 变换：RGB->HLS(浮点)->亮度调整->RGB。
    Excel 实际使用浮点 HLS 数学（shlwapi 整数 0-240 算法对 4472C4-0.25 得
    2F5695，与 Excel 实际值 2F5597 不符）；浮点版对 Office 标准色板主体精确
    命中，仅在 tint 接近 +/-0.8 极端值时差 1 色阶（8bit 量化舍入噪声）。"""
    r, g, b = (int(hex6[i:i+2], 16) for i in (0, 2, 4))
    h, l, sc = colorsys.rgb_to_hls(r / 255.0, g / 255.0, b / 255.0)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, sc)
    return '%02X%02X%02X' % (int(round(r2 * 255)), int(round(g2 * 255)), int(round(b2 * 255)))

def rgb_channel_close(h1, h2, limit=1):
    """两个 6 位 HEX 颜色逐通道差 <= limit（仅用于 HLS tint 浮点回算的 8bit 舍入噪声）"""
    try:
        a = [int(h1[i:i+2], 16) for i in (0, 2, 4)]
        b = [int(h2[i:i+2], 16) for i in (0, 2, 4)]
    except Exception:
        return False
    return all(abs(x - y) <= limit for x, y in zip(a, b))

def rgb_euclidean(h1, h2):
    """两个 6 位 HEX 颜色的 RGB 欧氏距离（调色板量化吸附容差用）"""
    try:
        a = [int(h1[i:i+2], 16) for i in (0, 2, 4)]
        b = [int(h2[i:i+2], 16) for i in (0, 2, 4)]
        return sum((x-y)**2 for x, y in zip(a, b)) ** 0.5
    except Exception:
        return 9999.0

# 可合并相邻区域的单元格级差异类型
MERGEABLE_DIFF_TYPES = {'内容变化','公式变化','字体变化','填充变化','边框变化','对齐变化','数字格式变化','富文本变化','单元格新增','单元格删除','图片新增','图片变动','图片尺寸变化'}

def _grid_segments(cell_set):
    """4连通单元格集合 -> 矩形块列表（每块为 {(col,row),...}）。
    非矩形（L/T形）先按行切连续段，再把相邻行同列范围的段纵向合并为矩形。"""
    remaining=set(cell_set); blocks=[]
    while remaining:
        seed=min(remaining); comp=set(); stack=[seed]
        while stack:
            c,r=stack.pop()
            if (c,r) in comp or (c,r) not in remaining: continue
            comp.add((c,r)); stack.extend([(c+1,r),(c-1,r),(c,r+1),(c,r-1)])
        rows={}
        for c,r in comp: rows.setdefault(r,[]).append(c)
        segs=[]
        for r in sorted(rows):
            cls=sorted(rows[r]); st=pr=cls[0]
            for c in cls[1:]+[None]:
                if c is not None and c==pr+1: pr=c
                else: segs.append([r,st,pr]); st=pr=c
        segs.sort(key=lambda x:(x[1],x[2],x[0])); merged=[]
        for r0,c0,c1 in segs:
            if merged and merged[-1][2]==c0 and merged[-1][3]==c1 and merged[-1][1]+1==r0: merged[-1][1]=r0
            else: merged.append([r0,r0,c0,c1])
        for rs,re_,c0,c1 in merged:
            blocks.append({(c,r) for r in range(rs,re_+1) for c in range(c0,c1+1)})
        remaining-=comp
    return blocks

def dedup_rich_text_diffs(diffs):
    """同地址已有内容变化 diff 时，富文本变化 diff 冗余（规则体检已含富文本检查），返回过滤后列表"""
    vc={(d['sheet'],d['address']) for d in diffs if d['type']=='内容变化'}
    return [d for d in diffs if not (d['type']=='富文本变化' and (d['sheet'],d['address']) in vc)]
def merge_adjacent_diffs(diffs):
    """同 sheet/类型/描述/豁免状态的相邻单元格差异合并为矩形区域（地址形如 A1:C4），
    非单元格级差异（结构/图片/条件格式/插件结果）原样保留。"""
    import collections
    groups=collections.defaultdict(list); out=[]
    for d in diffs:
        m=re.match(r'^([A-Za-z]+)([0-9]+)$', d.get('address','') or '')
        if d.get('type') in MERGEABLE_DIFF_TYPES and d.get('sheet')!='🔍 数据检查' and m:
            groups[(d['sheet'],d['type'],d['desc'],bool(d.get('rule_pass')),d.get('rule_name',''))].append((column_index_from_string(m.group(1)),int(m.group(2)),d))
        else: out.append(d)
    for (sheet,typ,desc,rp,_rn),items in groups.items():
        cells={(c,r):d for c,r,d in items}
        for block in _grid_segments(set(cells)):
            cs=[c for c,r in block]; rs=[r for c,r in block]
            addr=cell_address(min(cs),min(rs))
            if (min(cs),min(rs))!=(max(cs),max(rs)): addr+=':'+cell_address(max(cs),max(rs))
            nd=dict(cells[next(iter(block))]); nd['address']=addr
            ncells=len(block)
            if ncells>1: nd['desc']=f"{nd.get('desc','')} [共计{ncells}单元格]"
            out.append(nd)
    return out

def color_signature(color):
    """openpyxl 颜色对象的存储类型签名（无法解析时兜底）"""
    if color is None: return 'none'
    ct = getattr(color,'type',None)
    if ct == 'rgb': return 'rgb:%s' % (color.rgb or '')
    if ct == 'theme': return 'theme:%s:%s' % (color.theme, color.tint)
    if ct == 'indexed': return 'idx:%s' % color.indexed
    if ct == 'auto': return 'auto'
    return 'unknown'

def center_window(win, parent=None):
    """Toplevel 相对父窗口（父不可用时相对屏幕）居中"""
    win.update_idletasks()
    try: w=win.winfo_width(); h=win.winfo_height()
    except Exception: w=h=400
    if parent is not None:
        try:
            x=parent.winfo_rootx()+parent.winfo_width()//2-w//2
            y=parent.winfo_rooty()+parent.winfo_height()//2-h//2
        except Exception:
            x=(win.winfo_screenwidth()-w)//2; y=(win.winfo_screenheight()-h)//2
    else:
        x=(win.winfo_screenwidth()-w)//2; y=(win.winfo_screenheight()-h)//2
    win.geometry(f"+{max(x,0)}+{max(y,0)}")


TYPE_DISPLAY = {
    'value': '值变化', 'formula': '公式变化', 'rich_text': '富文本',
    'font': '字体', 'fill': '填充/背景色', 'border': '边框',
    'alignment': '对齐方式', 'number_format': '数字格式',
    'merged_cells': '合并单元格', 'row_height': '行高', 'col_width': '列宽',
    'images': '图片', 'conditional_format': '条件格式'
}

class WorkbookStyleCache:
    """直接解析 OOXML（styles.xml / theme1.xml / sheetN.xml），
    提供精确的颜色、数字格式、行高列宽解析，绕过 openpyxl 的样式继承缺陷。"""

    NSM = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    NSA = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
    NSR = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'

    def __init__(self, xlsx_path):
        self.path = xlsx_path
        self.indexed = {i: v for i, v in enumerate(DEFAULT_INDEXED)}
        self.indexed[64] = '000000'; self.indexed[65] = 'FFFFFF'
        self.theme = {}
        self.font_equiv = []  # 字体等价组列表
        self.num_fmts = {}
        self.cell_xfs = []      # [{numFmtId, fontId, fillId, borderId, xfId, applyNF}]
        self.style_xfs = []     # cellStyleXfs（命名样式基类）
        self.fonts = []         # 解析后的字体 dict
        self.fills = []         # 解析后的填充 dict
        self.borders = []       # 解析后的边框 dict
        self.sheet_files = {}   # sheet名 -> xl/worksheets/sheetN.xml
        self.dims_cache = {}
        self.formula_cache = {} # sheet名 -> {(r1,c1,r2,c2): 公式文本}（数组/共享公式区域）
        self.smap_cache = {}   # sheet名 -> {单元格地址: 原始s样式索引}
        self.merge_cache = {}  # sheet名 -> {成员地址: 锚点地址}
        self._zip = None
        try:
            self._zip = zipfile.ZipFile(xlsx_path)
            self._parse_theme()
            self._parse_styles()
            self._parse_sheet_map()
        except Exception:
            self._zip = None

    def close(self):
        if self._zip:
            try: self._zip.close()
            except Exception: pass
        self._zip = None

    # ---------- 数组/共享公式区域扫描 ----------
    # 只匹配带公式体的锚点 <f ...>公式</f>；自闭合的成员格 <f t="array" ref="..."/> 不匹配
    _AF_OPEN = re.compile(r'<f\b(?=[^>]*\bt="(?:array|shared)")[^>]*\bref="([^"]+)"[^>]*(?<!/)>')
    _AF_CLOSE = re.compile(r'</f>')
    _XML_ENTITIES = (('&lt;','<'), ('&gt;','>'), ('&quot;','"'), ('&apos;',"'"), ('&amp;','&'))

    @classmethod
    def _scan_formula_refs(cls, stream):
        """流式扫描 sheet XML，提取数组/共享公式锚点的 ref 区域与公式文本。
        数组公式（<f t="array" ref="I3:Z42">公式</f>）和共享公式（<f t="shared" ...>）
        只有锚点格存公式文本；区域内成员格是自闭合 <f t="array" ref="..."/>（空体），
        若错误匹配成员格，.*?</f> 会吞到下一个锚点公式，拼入 XML 碎片和其他公式（误报'异常符号'）。"""
        out = []
        buf = ''
        while True:
            chunk = stream.read(1 << 20)
            if not chunk: break
            buf += chunk.decode('utf-8', errors='ignore')
            pos = 0
            while True:
                mo = cls._AF_OPEN.search(buf, pos)
                if not mo:
                    pos = len(buf)
                    break
                mc = cls._AF_CLOSE.search(buf, mo.end())
                if not mc:
                    pos = mo.start()   # 开标签已到但闭合标签可能在下一 chunk，保留尾部
                    break
                ref, ftext = mo.group(1), buf[mo.end():mc.start()]
                for ent, ch in cls._XML_ENTITIES: ftext = ftext.replace(ent, ch)
                try:
                    first = ref.split(':')[0]
                    last = ref.split(':')[1] if ':' in ref else first
                    c1 = ''.join(ch for ch in first if ch.isalpha())
                    r1 = int(''.join(ch for ch in first if ch.isdigit()))
                    c2 = ''.join(ch for ch in last if ch.isalpha())
                    r2 = int(''.join(ch for ch in last if ch.isdigit()))
                    out.append(((r1, column_index_from_string(c1), r2, column_index_from_string(c2)), ftext))
                except Exception:
                    pass
                pos = mc.end()
            buf = buf[pos:] if pos >= len(buf) - 65536 else buf[-65536:]
        return out

    def get_formula_refs(self, sheet_name):
        if sheet_name in self.formula_cache: return self.formula_cache[sheet_name]
        refs = {}
        try:
            fn = self.sheet_files.get(sheet_name)
            if fn and self._zip is not None and fn in self._zip.namelist():
                with self._zip.open(fn) as fh:
                    for box, ftext in self._scan_formula_refs(fh):
                        refs[box] = ftext
        except Exception:
            pass
        self.formula_cache[sheet_name] = refs
        return refs

    def find_array_formula(self, sheet_name, row, col):
        """单元格若落在某数组/共享公式区域内，返回锚点公式文本，否则 None"""
        for (r1, c1, r2, c2), ftext in self.get_formula_refs(sheet_name).items():
            if r1 <= row <= r2 and c1 <= col <= c2:
                return ftext
        return None

    # ---------- theme ----------
    def _parse_theme(self):
        try:
            root = etree.fromstring(self._zip.read('xl/theme/theme1.xml'))
            scheme = root.find('.//%sclrScheme' % self.NSA)
            if scheme is None: return
            # clrScheme 子元素顺序为 dk1,lt1,dk2,lt2,accent1..6,hlink,folHlink；
            # 但单元格 <color theme="n"> 的槽位编号是交叉的：0=lt1(白)、1=dk1(黑)、2=lt2、3=dk2、4..=accent1..。
            # 直接按子元素顺序填会让 theme:1(黑色) 错取到 lt1(白色)，产生 Theme:1→Indexed:8 这类同色误报。
            THEME_SLOT = {0:1, 1:0, 2:3, 3:2, 4:4, 5:5, 6:6, 7:7, 8:8, 9:9, 10:10, 11:11}
            for idx, child in enumerate(scheme):
                srgb = child.find('%ssrgbClr' % self.NSA)
                slot = THEME_SLOT.get(idx)
                if srgb is not None and srgb.get('val') and slot is not None:
                    self.theme[slot] = srgb.get('val').upper()[-6:]; continue
                sysc = child.find('%ssysClr' % self.NSA)
                if sysc is not None and slot is not None:
                    val = sysc.get('lastClr') or sysc.get('val')
                    if val: self.theme[slot] = val.upper()[-6:]
        except Exception:
            pass

        # 解析字体等价表
        try:
            root = etree.fromstring(self._zip.read('xl/theme/theme1.xml'))
            for fg_tag in ('majorFont', 'minorFont'):
                fg = root.find('.//%s%s' % (self.NSA, fg_tag))
                if fg is None: continue
                names = set()
                latin = fg.find('%slatin' % self.NSA)
                if latin is not None:
                    tf = latin.get('typeface', '')
                    if tf: names.add(tf)
                for fs in fg.findall('%sfont' % self.NSA):
                    tf = fs.get('typeface', '')
                    if tf: names.add(tf)
                ea = fg.find('%sea' % self.NSA)
                if ea is not None:
                    tf = ea.get('typeface', '')
                    if not tf:
                        hans = fg.find('%sfont[@script="Hans"]' % self.NSA)
                        if hans is not None and hans.get('typeface'):
                            names.add(hans.get('typeface'))
                    else:
                        names.add(tf)
                if len(names) > 1:
                    self.font_equiv.append(names)
        except Exception:
            pass

    def _font_names_equivalent(self, n1, n2):
        if not n1 or not n2: return False
        if n1 == n2: return True
        for group in self.font_equiv:
            if n1 in group and n2 in group:
                return True
        return False
    # ---------- styles.xml ----------
    def _color_elem_to_dict(self, ce):
        if ce is None: return None
        rgb = ce.get('rgb')
        if rgb: return {'type':'rgb','rgb':rgb[-6:].upper(),'raw':rgb}
        t = ce.get('theme')
        if t is not None: return {'type':'theme','theme':int(t),'tint':float(ce.get('tint','0') or 0)}
        idx = ce.get('indexed')
        if idx is not None: return {'type':'indexed','idx':int(idx)}
        if ce.get('auto') is not None: return {'type':'auto'}
        return None

    def _parse_styles(self):
        try:
            root = etree.fromstring(self._zip.read('xl/styles.xml'))
        except Exception:
            return
        try:
            ice = root.find('%scolors/%sindexedColors' % (self.NSM, self.NSM))
            if ice is not None:
                entries = ice.findall('%srgbColor' % self.NSM)
                # <indexedColors> 有两种布局：
                #  Excel 保存：56 条，覆盖索引 8..63（ECMA 旧 EGA 调色板规范）
                #  openpyxl 保存：64 条，从索引 0 开始（其内部 COLOR_INDEX 全量序列化）
                start = 8 if len(entries) == 56 else 0
                for i, rc in enumerate(entries):
                    v = rc.get('rgb')
                    if v and 0 <= start + i < 64: self.indexed[start + i] = v[-6:].upper()
        except Exception: pass
        try:
            for nf in root.findall('%snumFmts/%snumFmt' % (self.NSM, self.NSM)):
                self.num_fmts[int(nf.get('numFmtId'))] = nf.get('formatCode')
        except Exception: pass
        try:
            for fnt in root.findall('%sfonts/%sfont' % (self.NSM, self.NSM)):
                sz_el = fnt.find('%ssz' % self.NSM)
                nm_el = fnt.find('%sname' % self.NSM)
                self.fonts.append({
                    'name': nm_el.get('val') if nm_el is not None else '',
                    'size': float(sz_el.get('val')) if sz_el is not None and sz_el.get('val') else None,
                    'bold': fnt.find('%sb' % self.NSM) is not None,
                    'italic': fnt.find('%si' % self.NSM) is not None,
                    'underline': fnt.find('%su' % self.NSM) is not None,
                    'color': self._color_elem_to_dict(fnt.find('%scolor' % self.NSM)),
                })
        except Exception: pass
        try:
            for fill in root.findall('%sfills/%sfill' % (self.NSM, self.NSM)):
                pf = fill.find('%spatternFill' % self.NSM)
                fd = {'pattern':None,'fg':None,'bg':None}
                if pf is not None:
                    fd['pattern'] = pf.get('patternType')
                    fd['fg'] = self._color_elem_to_dict(pf.find('%sfgColor' % self.NSM))
                    fd['bg'] = self._color_elem_to_dict(pf.find('%sbgColor' % self.NSM))
                self.fills.append(fd)
        except Exception: pass
        try:
            for bd in root.findall('%sborders/%sborder' % (self.NSM, self.NSM)):
                sides = {}
                for side in ('left','right','top','bottom'):
                    se = bd.find('%s%s' % (self.NSM, side))
                    if se is None: sides[side] = {'style':None,'color':None}
                    else:
                        _ce = self._color_elem_to_dict(se.find('%scolor' % self.NSM))
                        # 有线但缺省 <color>：OOXML 语义=自动色(黑)，与 auto / indexed=64 等价
                        sides[side] = {'style':se.get('style'),'color':_ce or {'type':'auto'}}
                self.borders.append(sides)
        except Exception: pass
        for tag, store in (('cellStyleXfs', self.style_xfs), ('cellXfs', self.cell_xfs)):
            try:
                for xf in root.findall('%s%s/%sxf' % (self.NSM, tag, self.NSM)):
                    # OOXML 语义：apply* 属性缺省 = 1（应用该属性），显式 "0" 才不应用。
                    # openpyxl 写出的文件根本不带这些属性，缺省必须按 True 处理
                    if tag == 'cellXfs':
                        _ap = lambda name: xf.get(name) != '0'
                    else:
                        _ap = lambda name: xf.get(name) == '1'
                    store.append({
                        'numFmtId': int(xf.get('numFmtId','0')),
                        'fontId': int(xf.get('fontId','0')),
                        'fillId': int(xf.get('fillId','0')),
                        'borderId': int(xf.get('borderId','0')),
                        'xfId': int(xf.get('xfId','-1')),
                        'applyNF': _ap('applyNumberFormat'),
                        'applyFont': _ap('applyFont'),
                        'applyFill': _ap('applyFill'),
                        'applyBorder': _ap('applyBorder'),
                        'applyAlign': _ap('applyAlignment'),
                    })
            except Exception: pass

    # ---------- sheet 文件映射与行列尺寸 ----------
    def _parse_sheet_map(self):
        try:
            wb = etree.fromstring(self._zip.read('xl/workbook.xml'))
            rels = etree.fromstring(self._zip.read('xl/_rels/workbook.xml.rels'))
            rid_target = {rel.get('Id'): rel.get('Target') for rel in rels}
            for s in wb.findall('%ssheets/%ssheet' % (self.NSM, self.NSM)):
                rid = s.get('%sid' % self.NSR)
                tgt = rid_target.get(rid,'')
                name = s.get('name','')
                if tgt.endswith('.xml') and ('worksheet' in tgt):
                    if tgt.startswith('/'):
                        tgt = tgt.lstrip('/')          # /xl/worksheets/sheet1.xml
                    elif not tgt.startswith('xl/'):
                        tgt = 'xl/' + tgt              # worksheets/sheet1.xml
                    self.sheet_files[name] = tgt
        except Exception:
            pass

    def get_dims(self, sheet_name):
        if sheet_name in self.dims_cache: return self.dims_cache[sheet_name]
        dims = {'rows':{}, 'cols':{}, 'def_h': 15.0, 'def_w': 8.43}
        try:
            fn = self.sheet_files.get(sheet_name)
            if fn and self._zip is not None and fn in self._zip.namelist():
                self._scan_sheet_dims(self._zip.open(fn), dims)
        except Exception:
            pass
        self.dims_cache[sheet_name] = dims
        return dims

    _ROW_TAG = re.compile(r'<row\b([^>]*?)/?>')
    _COL_TAG = re.compile(r'<col\b([^>]*?)/?>')
    _SFP_TAG = re.compile(r'<sheetFormatPr\b[^>]*>')
    _ATTR = re.compile(r'(\w+)="([^"]*)"')

    @classmethod
    def _scan_sheet_dims(cls, stream, dims):
        # 流式正则扫描：只取 sheetFormatPr/cols/row 开始标签属性，不解析单元格树
        buf = ''
        head_done = False
        while True:
            chunk = stream.read(1 << 20)
            if not chunk: break
            buf += chunk.decode('utf-8', errors='ignore')
            if not head_done:
                sd = buf.find('<sheetData')
                if sd >= 0:
                    cls._parse_head_dims(buf[:sd], dims)
                    buf = buf[sd:]
                    head_done = True
                else:
                    # 保留尾部防止标签被截断
                    buf = buf[-8192:]
            if head_done:
                last = 0
                for m in cls._ROW_TAG.finditer(buf):
                    attrs = dict(cls._ATTR.findall(m.group(1)))
                    r = attrs.get('r'); ht = attrs.get('ht')
                    if r and ht:
                        dims['rows'][int(r)] = (float(ht), attrs.get('customHeight') == '1')
                    last = m.end()
                # 保留最后不完整片段
                buf = buf[max(last, len(buf)-4096):] if last else buf[-4096:]
        if not head_done and buf:
            cls._parse_head_dims(buf, dims)

    @classmethod
    def _parse_head_dims(cls, text, dims):
        m = cls._SFP_TAG.search(text)
        if m:
            attrs = dict(cls._ATTR.findall(m.group(0)))
            if attrs.get('defaultRowHeight'):
                try: dims['def_h'] = float(attrs['defaultRowHeight'])
                except Exception: pass
            if attrs.get('defaultColWidth'):
                try: dims['def_w'] = float(attrs['defaultColWidth'])
                except Exception: pass
        for cm in cls._COL_TAG.finditer(text):
            attrs = dict(cls._ATTR.findall(cm.group(1)))
            cw = attrs.get('width'); cmin = attrs.get('min'); cmax = attrs.get('max')
            if cw and cmin:
                try:
                    w = float(cw); custom = attrs.get('customWidth') == '1'
                    for ci in range(int(cmin), int(cmax or cmin) + 1):
                        dims['cols'][ci] = (w, custom)
                except Exception:
                    pass

    # ---------- 颜色精确解析 ----------
    def _color_dict_from_openpyxl(self, color):
        if color is None: return None
        ct = getattr(color, 'type', None)
        if ct == 'rgb':
            rgb = color.rgb
            if not rgb: return None
            return {'type':'rgb','rgb':rgb[-6:].upper(),'raw':str(rgb)}
        if ct == 'theme':
            return {'type':'theme','theme':int(color.theme),
                    'tint':float(color.tint) if color.tint is not None else 0.0}
        if ct == 'indexed': return {'type':'indexed','idx':int(color.indexed)}
        if ct == 'auto': return {'type':'auto'}
        return None

    def _auto_color(self, context='text'):
        # theme 槽位：1=dk1(windowText 窗口文字，默认黑)；0=lt1(window 窗口背景，默认白)
        if context == 'text':
            return self.theme.get(1, '000000')
        return self.theme.get(0, 'FFFFFF')

    def resolve_color(self, color, context='text'):
        """颜色对象/dict -> 精确 RGB（6位HEX）。无法解析返回 None。"""
        cd = color
        if not isinstance(cd, dict):
            cd = self._color_dict_from_openpyxl(color)
        if cd is None: return None
        t = cd.get('type')
        if t == 'rgb':
            rgb = cd.get('rgb')
            if not rgb: return None
            rgb = rgb[-6:].upper()
            if rgb == '000000' and str(cd.get('raw','')).upper() == '00000000':
                return self._auto_color(context)
            return rgb
        if t == 'indexed':
            idx = cd.get('idx')
            if idx == 64: return self._auto_color('text')
            if idx == 65: return self._auto_color('bg')
            return self.indexed.get(idx)
        if t == 'theme':
            base = self.theme.get(cd.get('theme'))
            if not base: return None
            tint = cd.get('tint', 0.0) or 0.0
            if abs(tint) < 1e-9: return base
            try: return excel_tint(base, tint)
            except Exception: return base
        if t == 'auto':
            return self._auto_color(context)
        return None

    _C_TAG = re.compile(r'<c\b([^>]*?)/?>')
    _R_ATTR = re.compile(r'\br="([A-Z]+\d+)"')
    _S_ATTR = re.compile(r'\bs="(\d+)"')

    def _sheet_smap(self, sheet_name):
        """从 sheet XML 直接解析 地址->原始s样式索引（权威），绕过 openpyxl 对 WPS/超大
        xf 表重建 style_id 时的错位（如加粗 <b/> 被错读成非粗体）。"""
        if sheet_name in self.smap_cache: return self.smap_cache[sheet_name]
        m = {}
        try:
            fn = self.sheet_files.get(sheet_name)
            if fn and self._zip is not None and fn in self._zip.namelist():
                for mt in self._C_TAG.finditer(self._zip.read(fn).decode('utf-8', 'ignore')):
                    tag = mt.group(1)
                    mr = self._R_ATTR.search(tag); ms = self._S_ATTR.search(tag)
                    if mr and ms: m[mr.group(1)] = int(ms.group(1))
        except Exception:
            pass
        self.smap_cache[sheet_name] = m
        return m

    def _raw_sid(self, cell):
        try:
            smap = self._sheet_smap(cell.parent.title)
            sid = smap.get(cell.coordinate)
            if sid is None:
                # 合并区域成员格通常无独立s属性：Excel/WPS中其显示样式由锚点格决定，回退取锚点
                anchor = self._merged_anchor(cell.parent.title).get(cell.coordinate)
                if anchor: sid = smap.get(anchor)
            return sid
        except Exception: return None

    def _merged_anchor(self, sheet_name):
        if sheet_name in self.merge_cache: return self.merge_cache[sheet_name]
        m = {}
        try:
            fn = self.sheet_files.get(sheet_name)
            if fn and self._zip is not None and fn in self._zip.namelist():
                root = etree.fromstring(self._zip.read(fn))
                for mc in root.findall('.//%smergeCells/%smergeCell' % (self.NSM, self.NSM)):
                    ref = mc.get('ref') or ''
                    if ':' not in ref: continue
                    a, b = ref.split(':', 1)
                    ac = ''.join(ch for ch in a if ch.isalpha()); ar = int(''.join(ch for ch in a if ch.isdigit()))
                    bc = ''.join(ch for ch in b if ch.isalpha()); br = int(''.join(ch for ch in b if ch.isdigit()))
                    for r in range(ar, br+1):
                        for c in range(column_index_from_string(ac), column_index_from_string(bc)+1):
                            addr = '%s%d' % (get_column_letter(c), r)
                            if addr != a: m[addr] = a
        except Exception:
            pass
        self.merge_cache[sheet_name] = m
        return m

    def _xf_by_cell(self, cell):
        """单元格样式索引优先取 sheet XML 原始 s 属性（与 Excel/WPS 实际一致）；
        openpyxl 的 cell.style_id 仅作兜底。"""
        sid = self._raw_sid(cell)
        if sid is None:
            try: sid = cell.style_id
            except Exception: return None
        if 0 <= sid < len(self.cell_xfs):
            return self.cell_xfs[sid]
        return None

    def resolve_numfmt(self, cell):
        """按 xf -> cellStyleXfs 命名样式继承链解析真实数字格式"""
        xf = self._xf_by_cell(cell)
        if xf is None:
            try: return cell.number_format or 'General'
            except Exception: return 'General'
        fid = xf['numFmtId']
        xfid = xf.get('xfId', -1)
        # 单元格 xf 未指定数字格式（numFmtId=0 且未显式 applyNumberFormat）时，继承命名样式
        if fid == 0 and not xf['applyNF'] and 0 <= xfid < len(self.style_xfs):
            base_fid = self.style_xfs[xfid].get('numFmtId', 0)
            if base_fid: fid = base_fid
        if fid == 0: return 'General'
        return self.num_fmts.get(fid) or BUILTIN_NUMFMTS.get(fid) or 'General'

    # ---------- font / fill / border（命名样式继承场景下补全 openpyxl 残缺对象） ----------
    def _base_xf(self, xf):
        xfid = xf.get('xfId', -1)
        if xfid is not None and 0 <= xfid < len(self.style_xfs):
            return self.style_xfs[xfid]
        return None

    def resolve_font(self, cell, opxl_font):
        xf = self._xf_by_cell(cell)
        sxf = self._base_xf(xf) if xf else None
        # xf 显式不应用字体（applyFont=0）时整个字体对象取命名样式基线
        if xf is not None and not xf['applyFont'] and sxf is not None:
            own = self.fonts[sxf['fontId']] if 0 <= sxf['fontId'] < len(self.fonts) else {}
            base = {}
        else:
            own = self.fonts[xf['fontId']] if xf and 0 <= xf['fontId'] < len(self.fonts) else {}
            base = self.fonts[sxf['fontId']] if sxf and 0 <= sxf['fontId'] < len(self.fonts) else {}
        col = self._color_dict_from_openpyxl(getattr(opxl_font, 'color', None) if opxl_font is not None else None)
        if col is None:
            col = own.get('color') or base.get('color')
        # 布尔属性：字体记录里有明确值就用它（没写 b 标签 = False，是明确的非粗体），
        # 不能从命名样式继承 True——否则"命名样式加粗 + 单元格显式取消加粗"会被错判为加粗
        return {
            'name': own.get('name') or base.get('name') or getattr(opxl_font, 'name', None),
            'size': own.get('size') or base.get('size') or getattr(opxl_font, 'size', None),
            'bold': bool(own.get('bold') or (False if 'bold' in own else base.get('bold'))),
            'italic': bool(own.get('italic') or (False if 'italic' in own else base.get('italic'))),
            'underline': bool(own.get('underline') or (False if 'underline' in own else base.get('underline'))),
            'color': col,
        }

    def _safe_xf(self, cell):
        return self._xf_by_cell(cell)

    def resolve_fill(self, cell):
        xf = self._xf_by_cell(cell)
        if xf is None: return None
        sxf = self._base_xf(xf)
        if not xf['applyFill'] and sxf is not None:
            own = self.fills[sxf['fillId']] if 0 <= sxf['fillId'] < len(self.fills) else {}
            base = {}
        else:
            own = self.fills[xf['fillId']] if 0 <= xf['fillId'] < len(self.fills) else {}
            base = self.fills[sxf['fillId']] if sxf and 0 <= sxf['fillId'] < len(self.fills) else {}
        pat = own.get('pattern') or base.get('pattern')
        if pat in (None, 'none'): return {'pattern': None}
        return {'pattern': pat,
                'fg': own.get('fg') or base.get('fg'),
                'bg': own.get('bg') or base.get('bg')}

    def resolve_border(self, cell):
        xf = self._xf_by_cell(cell)
        if xf is None: return None
        sxf = self._base_xf(xf)
        if not xf['applyBorder'] and sxf is not None:
            own = self.borders[sxf['borderId']] if 0 <= sxf['borderId'] < len(self.borders) else {}
            base = {}
        else:
            own = self.borders[xf['borderId']] if 0 <= xf['borderId'] < len(self.borders) else {}
            base = self.borders[sxf['borderId']] if sxf and 0 <= sxf['borderId'] < len(self.borders) else {}
        out = {}
        for side in ('left','right','top','bottom'):
            o = own.get(side) or {}
            b = base.get(side) or {}
            out[side] = {'style': o.get('style') or b.get('style'),
                         'color': o.get('color') or b.get('color')}
        return out

def cell_address(col, row): return f"{get_column_letter(col)}{row}"
def normalize_path(p):
    try: return os.path.normpath(os.path.realpath(p))
    except: return os.path.normpath(p)
def get_sheet_names_fast(path):
    """毫秒级读取 sheet 名：直接解析 xl/workbook.xml（几十KB），
    不用 openpyxl 加载整个工作簿（300MB 文件 UI 线程会假死数十秒）。"""
    if not os.path.isfile(path): return []
    try:
        zf = zipfile.ZipFile(path)
        try:
            rels = etree.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            rid_target = {rel.get('Id'): rel.get('Target') for rel in rels}
            wb = etree.fromstring(zf.read('xl/workbook.xml'))
            NSR = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
            names = []
            for s in wb.findall(f'{NS}sheets/{NS}sheet'):
                rid = s.get(f'{NSR}id', '')
                tgt = rid_target.get(rid, '')
                if tgt.endswith('.xml') and 'worksheet' in tgt:
                    names.append(s.get('name', ''))
            if names: return names
        finally:
            zf.close()
    except Exception:
        pass
    # 兜底：openpyxl read_only
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        names = list(wb.sheetnames); wb.close(); return names
    except Exception:
        return []

def formula_text(value):
    """从单元格值提取公式文本：普通公式是 '=...' 字符串；
    数组公式在 openpyxl 中是 ArrayFormula 对象（.text 存公式文本）；否则 None。"""
    if isinstance(value, str) and value.startswith('='):
        return value
    text = getattr(value, 'text', None)
    if isinstance(text, str) and text.startswith('='):
        return text
    return None

def normalize_formula(f):
    """公式规范化比较：去首尾空格、统一大写函数名、折叠空白。"""
    if not f: return ''
    return re.sub(r'\s+', '', f).upper()

def _fmt_duration(secs):
    secs=int(secs+0.5)
    if secs<60: return f'{secs}秒'
    return f'{secs//60}分{secs%60}秒'

# 富文本解析
def _rpr_bool(rPr, tag):
    el = rPr.find(f'{NS}{tag}')
    if el is None: return None                      # 缺省=继承单元格字体，不参与比对
    return el.get('val', '1') not in ('0', 'false')

def _parse_rPr(rPr):
    font = {'name':'','size':'','bold':None,'italic':None,'underline':None,'color':''}
    if rPr is None: return font
    rf = rPr.find(f'{NS}rFont'); font['name'] = rf.get('val','') if rf is not None else ''
    sz = rPr.find(f'{NS}sz'); font['size'] = sz.get('val','') if sz is not None else ''
    font['bold'] = _rpr_bool(rPr, 'b'); font['italic'] = _rpr_bool(rPr, 'i'); font['underline'] = _rpr_bool(rPr, 'u')
    c = rPr.find(f'{NS}color'); font['color'] = (c.get('rgb','') or c.get('theme','')) if c is not None else ''
    return font

def _extract_runs_from_si(si):
    t = si.find(f'{NS}t')
    if t is not None: return [(t.text or '', None)]
    runs = []
    for r in si.findall(f'{NS}r'):
        t = r.find(f'{NS}t'); runs.append((t.text if t is not None else '', _parse_rPr(r.find(f'{NS}rPr'))))
    return runs

def parse_rich_text_from_xlsx(path):
    result = {}
    if not os.path.isfile(path): return result
    try: zf = zipfile.ZipFile(path)
    except: return result
    shared = []
    if 'xl/sharedStrings.xml' in zf.namelist():
        root = etree.fromstring(zf.read('xl/sharedStrings.xml'))
        for si in root.findall(f'{NS}si'): shared.append(_extract_runs_from_si(si))
    sheet_map = {}; rid_to_target = {}
    if 'xl/_rels/workbook.xml.rels' in zf.namelist():
        rels = etree.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        for rel in rels.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
            rid_to_target[rel.get('Id','')] = rel.get('Target','')
    if 'xl/workbook.xml' in zf.namelist():
        wb = etree.fromstring(zf.read('xl/workbook.xml'))
        sheets_elem = wb.find(f'{NS}sheets')
        if sheets_elem is not None:
            for s in sheets_elem.findall(f'{NS}sheet'):
                rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id','')
                target = rid_to_target.get(rid,'')
                if target.endswith('.xml'): sheet_map[target.split('/')[-1]] = s.get('name','')
    if not sheet_map:
        files = sorted([n for n in zf.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')],
                       key=lambda x: int(x.split('sheet')[1].split('.xml')[0]))
        for i,f in enumerate(files,1): sheet_map[f.split('/')[-1]] = f'Sheet{i}'
    for fn, sheet_name in sheet_map.items():
        if f'xl/worksheets/{fn}' not in zf.namelist(): continue
        ws = etree.fromstring(zf.read(f'xl/worksheets/{fn}'))
        data = {}
        for c in ws.findall(f'.//{NS}c'):
            ref = c.get('r',''); t = c.get('t','')
            if t == 's':
                v = c.find(f'{NS}v')
                if v is not None and v.text:
                    idx = int(v.text)
                    if 0 <= idx < len(shared) and any(f is not None for _,f in shared[idx]): data[ref] = shared[idx]
            elif t == 'inlineStr':
                is_ = c.find(f'{NS}is')
                if is_ is not None:
                    runs = _extract_runs_from_si(is_)
                    if runs and any(f is not None for _,f in runs): data[ref] = runs
        if data: result[sheet_name] = data
    zf.close()
    return result

def compare_rich_text_runs(r1, r2):
    if not r1 and not r2: return None
    if r1 is None and r2 is None: return None
    if (r1 and not r2) or (not r1 and r2):
        p1 = ''.join(t for t,_ in (r1 or [])); p2 = ''.join(t for t,_ in (r2 or []))
        return f"内容(含富文本): {p1} → {p2}" if p1 != p2 else "单元格变为富文本格式"
    p1 = ''.join(t for t,_ in r1); p2 = ''.join(t for t,_ in r2)
    if p1 != p2: return f"内容(含富文本): {p1} → {p2}"
    if len(r1) != len(r2): return f"富文本段落数不同: {len(r1)} → {len(r2)}"
    changes = []
    for i,((t1,f1),(t2,f2)) in enumerate(zip(r1,r2)):
        if f1 is not None and f2 is not None and f1 != f2:
            diff = []
            for k in ['name','size','bold','italic','underline','color']:
                v1 = f1.get(k); v2 = f2.get(k)
                # 缺省(None/空)=run继承单元格字体：显式同值vs缺省不算差异（真实差异由单元格级字体比对兜底）
                if v1 in (None, '') or v2 in (None, ''): continue
                if v1 != v2: diff.append(f"{k}: {v1}→{v2}")
            if diff: changes.append(f"段{i+1} '{t1}': {'; '.join(diff)}")
    return "富文本格式变更:\n" + "\n".join(changes) if changes else None

class DataLocator:
    def __init__(self, rules_file=None):
        self.rules = []; self.rules_file = rules_file
        if rules_file and os.path.isfile(rules_file): self.load_rules(rules_file)
    def load_rules(self, filepath):
        try:
            with open(filepath,'r',encoding='utf-8') as f: data=json.load(f)
            self.rules=data.get('rules',[]); self.rules_file=filepath; return True
        except: return False
    def locate_all(self, workbook):
        results = {}
        for rule in self.rules:
            name = rule.get('name','unnamed')
            try: results[name]=self.locate(workbook, rule)
            except Exception as e: results[name]={'error':str(e)}
        return results
    def locate(self, workbook, rule):
        sheet_name = rule.get('sheet','')
        if sheet_name not in workbook.sheetnames: return {'error':f'Sheet "{sheet_name}" not found'}
        ws = workbook[sheet_name]; mode = rule.get('mode','offset')
        if mode == 'offset': return self._locate_offset(ws, rule)
        elif mode == 'intersection': return self._locate_intersection(ws, rule)
        elif mode == 'range': return self._locate_range(ws, rule)
        elif mode == 'collect': return {'error':'collect 模式已移除（range 模式已覆盖，请改用 range）'}
        return {'error':f'Unknown mode: {mode}'}
    @staticmethod
    def _parse_area(area):
        """'G10:K15' 或 'B2' → (r1,c1,r2,c2)；非法/空 → None"""
        if not area: return None
        s = str(area).strip().upper().replace('$','')
        if not s: return None
        m = re.match(r'^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$', s)
        if not m: return None
        try:
            c1 = column_index_from_string(m.group(1)); r1 = int(m.group(2))
            c2 = column_index_from_string(m.group(3) or m.group(1)); r2 = int(m.group(4) or m.group(2))
        except Exception: return None
        return (min(r1,r2), min(c1,c2), max(r1,r2), max(c1,c2))
    def _find_anchor(self, ws, cfg):
        text = cfg.get('text','').strip(); search_in = cfg.get('search_in','all')
        area = self._parse_area(search_in)
        if area:
            r1,c1,r2,c2 = area
            for row in range(r1, min(r2,(ws.max_row or r2))+1):
                for col in range(c1, min(c2,(ws.max_column or c2))+1):
                    v=ws.cell(row,col).value
                    if v is not None and text in str(v).strip(): return (row,col)
            return None
        if search_in == 'first_row':
            for col in range(1,(ws.max_column or 1)+1):
                v=ws.cell(1,col).value
                if v is not None and text in str(v).strip(): return (1,col)
        elif search_in == 'first_col':
            for row in range(1,(ws.max_row or 1)+1):
                v=ws.cell(row,1).value
                if v is not None and text in str(v).strip(): return (row,1)
        else:
            for row in range(1,(ws.max_row or 1)+1):
                for col in range(1,(ws.max_column or 1)+1):
                    v=ws.cell(row,col).value
                    if v is not None and text in str(v).strip(): return (row,col)
        return None
    @staticmethod
    def _merge_search_in(anchor_cfg, search_in):
        if not search_in or search_in == 'all': return anchor_cfg or {}
        cfg = dict(anchor_cfg or {}); cfg.setdefault('search_in', search_in); return cfg
    def _locate_offset(self, ws, rule):
        anchor = self._find_anchor(ws, self._merge_search_in(rule.get('anchor',{}), rule.get('search_in','')))
        if not anchor: return {'error': f'Anchor "{rule.get("anchor",{}).get("text")}" not found'}
        tr = anchor[0] + rule.get('target',{}).get('row_offset',0)
        tc = anchor[1] + rule.get('target',{}).get('col_offset',0)
        if tr<1 or tc<1: return {'error':'Target out of range'}
        cell = ws.cell(tr,tc)
        return {'address':cell_address(tc,tr),'value':cell.value,'is_formula':isinstance(cell.value,str) and cell.value.startswith('=')}
    def _locate_intersection(self, ws, rule):
        ra = self._find_anchor(ws, {**rule.get('row_anchor',{}), 'search_in': rule.get('row_anchor',{}).get('search_in','all')})
        ca = self._find_anchor(ws, {**rule.get('col_anchor',{}), 'search_in': rule.get('col_anchor',{}).get('search_in','all')})
        if not ra: return {'error': f'Row anchor "{rule.get("row_anchor",{}).get("text")}" not found'}
        if not ca: return {'error': f'Col anchor "{rule.get("col_anchor",{}).get("text")}" not found'}
        cell = ws.cell(ra[0], ca[1])
        return {'row_address':cell_address(1,ra[0]),'col_address':cell_address(ca[1],1),'address':cell_address(ca[1],ra[0]),'value':cell.value,'is_formula':isinstance(cell.value,str) and cell.value.startswith('=')}
    def _range_cfg(self, ws, anchor_cfg, target_cfg, label=''):
        anchor = self._find_anchor(ws, anchor_cfg or {})
        if not anchor: return {'error': f'Anchor "{(anchor_cfg or {}).get("text","")}" not found' + (f' ({label})' if label else '')}
        target = target_cfg or {}; start_row = anchor[0] + target.get('row_offset',0); start_col = anchor[1] + target.get('col_offset',0)
        row_count = target.get('row_count',1); col_count = target.get('col_count',1); exclude = target.get('exclude',[])
        # 语义：start=锚点+offset；count 为含 start 在内的格数，负 count 向反方向延伸
        # （ro=-1,rc=-3 → start=锚点上1行，向上取3格=锚点上方连续3行；正count行为不变）
        if row_count >= 0: r1, r2 = start_row, start_row + row_count - 1
        else:              r1, r2 = start_row + row_count + 1, start_row
        if col_count >= 0: c1, c2 = start_col, start_col + col_count - 1
        else:              c1, c2 = start_col + col_count + 1, start_col
        exclude_set = set()
        for ex in exclude:
            if isinstance(ex,str):
                m = re.match(r'^([A-Z]+)(\d+)$', ex)
                if m: exclude_set.add(f"{m.group(1)}{m.group(2)}")
            elif isinstance(ex,list) and len(ex)==2:
                er=r1+ex[0] if ex[0]>=0 else r2+ex[0]+1
                ec=c1+ex[1] if ex[1]>=0 else c2+ex[1]+1
                if er>=1 and ec>=1: exclude_set.add(cell_address(ec,er))
        addresses=[]; values=[]
        for r in range(r1, r2+1):
            if r < 1 or r > (ws.max_row or 1): continue
            for c in range(c1, c2+1):
                if c < 1 or c > (ws.max_column or 1): continue
                addr = cell_address(c,r)
                if addr in exclude_set: continue
                cell = ws.cell(r,c); addresses.append(addr); values.append(cell.value)
        return {'address':addresses[0] if addresses else None,'addresses':addresses,'values':values,'range_count':len(addresses),'start':(r1,c1),'r1':r1,'r2':r2,'c1':c1,'c2':c2}
    def _locate_range(self, ws, rule):
        return self._range_cfg(ws, rule.get('anchor',{}), rule.get('target',{}))

class CheckPlugin:
    name=""; description=""
    def __init__(self, config=None): self.config = config or {}
    def check(self, old_data, new_data, context=None): raise NotImplementedError
class MeanDeviationPlugin(CheckPlugin):
    name="mean_deviation"; description="均值偏差"
    def check(self, old_data, new_data, context=None):
        results=[]; threshold=self.config.get('threshold',0.05)
        old_val=old_data.get('value') if isinstance(old_data,dict) else old_data
        new_val=new_data.get('value') if isinstance(new_data,dict) else new_data
        if old_val is None or new_val is None: return results
        try:
            old_num=float(old_val); new_num=float(new_val)
            if old_num!=0:
                dev=abs(new_num-old_num)/abs(old_num)
                if dev>threshold: results.append({'type':'均值偏差告警','desc':f'偏差 {dev:.2%} (阈值 {threshold:.0%}): {old_num:.4f} → {new_num:.4f}','severity':'warning'})
        except: pass
        return results
class ParamLockPlugin(CheckPlugin):
    name="param_lock"; description="参数锁定"
    def check(self, old_data, new_data, context=None):
        results=[]
        old_val=old_data.get('value') if isinstance(old_data,dict) else old_data
        new_val=new_data.get('value') if isinstance(new_data,dict) else new_data
        if old_val is not None and new_val is not None and old_val!=new_val:
            results.append({'type':'参数修改告警','desc':f'参数被修改: {old_val} → {new_val}','severity':'error'})
        return results
class RangeCheckPlugin(CheckPlugin):
    name="range_check"; description="范围检查"
    def check(self, old_data, new_data, context=None):
        results=[]; lsl=self.config.get('lsl'); usl=self.config.get('usl')
        new_val=new_data.get('value') if isinstance(new_data,dict) else new_data
        if new_val is None: return results
        try:
            num=float(new_val)
            if lsl is not None and num<float(lsl): results.append({'type':'低于下限','desc':f'值 {num:.4f} < LSL {lsl}','severity':'error'})
            if usl is not None and num>float(usl): results.append({'type':'超出上限','desc':f'值 {num:.4f} > USL {usl}','severity':'error'})
        except: pass
        return results
PLUGIN_REGISTRY={'mean_deviation':MeanDeviationPlugin,'param_lock':ParamLockPlugin,'range_check':RangeCheckPlugin}
class PluginManager:
    def __init__(self, config_file=None):
        self.plugins=[]; self.locator=None
        if config_file and os.path.isfile(config_file): self.load_config(config_file)
    def load_config(self, filepath):
        try:
            with open(filepath,'r',encoding='utf-8') as f: data=json.load(f)
            if 'locator_rules' in data: self.locator=DataLocator(); self.locator.rules=data['locator_rules']
            for cfg in data.get('plugins',[]):
                pname=cfg.get('type','')
                if pname in PLUGIN_REGISTRY:
                    plugin=PLUGIN_REGISTRY[pname](cfg.get('config',{})); plugin.rule_name=cfg.get('rule_name',''); plugin.description=cfg.get('description','')
                    self.plugins.append(plugin)
            return True
        except: return False
    def run_checks(self, old_wb, new_wb, log_callback=None):
        if not self.locator or not self.plugins: return []
        results=[]; old_data=self.locator.locate_all(old_wb); new_data=self.locator.locate_all(new_wb)
        for plugin in self.plugins:
            rule_name=getattr(plugin,'rule_name','')
            if not rule_name: continue
            old_val=old_data.get(rule_name); new_val=new_data.get(rule_name)
            if old_val is None and new_val is None: continue
            try:
                diffs=plugin.check(old_val,new_val)
                for diff in diffs:
                    diff['rule_name']=rule_name; diff['plugin']=plugin.name; results.append(diff)
                    if log_callback: log_callback(f" [{plugin.name}] {rule_name}: {diff['desc']}")
            except Exception as e:
                if log_callback: log_callback(f" [{plugin.name}] {rule_name} 检查失败: {e}")
        return results

class AdvancedEngine:
    name=""; description=""
    def __init__(self, config=None): self.config=config or {}
    def run(self, new_wb, rule, log_cb=None): return []

class FileNameCheckEngine(AdvancedEngine):
    name="filename_check"; description="文件名一致性"
    def run(self, new_wb, rule, log_cb=None):
        fmt=self.config.get('format_template',''); mappings=self.config.get('field_mappings',[])
        segments=self.config.get('segments',[])  # New: parsed segments with index
        fname=os.path.splitext(os.path.basename(new_wb.filename if hasattr(new_wb,'filename') and new_wb.filename else ''))[0]
        if not fname: return []
        alerts=[]
        # Mode 1: Template-based (backward compatible)
        if fmt and mappings:
            filled=fmt
            for m in mappings:
                field=m.get('field',''); sheet=m.get('sheet',''); cell_addr=m.get('cell','')
                if not field or not sheet or not cell_addr: continue
                if sheet not in new_wb.sheetnames: continue
                try:
                    col_str=''.join(ch for ch in cell_addr if ch.isalpha())
                    row_str=''.join(ch for ch in cell_addr if ch.isdigit())
                    if not col_str or not row_str: continue
                    val=new_wb[sheet].cell(int(row_str),column_index_from_string(col_str)).value
                    val_str=str(val).strip() if val is not None else ''
                    filled=filled.replace('{'+field+'}', val_str)
                except Exception as e:
                    if log_cb: log_cb(f" [filename_check] {field}: {e}")
            if filled != fname:
                first=mappings[0] if mappings else {}
                alerts.append({'sheet':first.get('sheet',''),'address':first.get('cell','A1'),'type':'文件名一致性',
                    'desc':f'文件名不匹配: 实际 "{fname}" ≠ 模板填充 "{filled}" (模板: {fmt})','advanced_check':True})
        # Mode 2: Segment-based (new)
        elif segments:
            # Split actual filename by underscore
            actual_parts = fname.split('_')
            for seg in segments:
                idx = seg.get('index', -1)
                expected = seg.get('value', '')
                sheet = seg.get('sheet', '')
                cell_addr = seg.get('cell', '')
                if idx < 0 or not sheet or not cell_addr: continue
                if sheet not in new_wb.sheetnames: continue
                try:
                    # Get expected value from report cell
                    col_str=''.join(ch for ch in cell_addr if ch.isalpha())
                    row_str=''.join(ch for ch in cell_addr if ch.isdigit())
                    if not col_str or not row_str: continue
                    cell_val = new_wb[sheet].cell(int(row_str), column_index_from_string(col_str)).value
                    cell_str = str(cell_val).strip() if cell_val is not None else ''
                    # Get actual segment from filename
                    actual_seg = actual_parts[idx] if idx < len(actual_parts) else ''
                    # Compare
                    if actual_seg != cell_str:
                        alerts.append({'sheet': sheet, 'address': cell_addr, 'type': '文件名一致性',
                            'desc': f'文件名段[{idx}]不匹配: 实际 "{actual_seg}" ≠ 报告 "{cell_str}" (期望位置: {idx})',
                            'advanced_check': True})
                except Exception as e:
                    if log_cb: log_cb(f" [filename_check] segment[{idx}]: {e}")
        return alerts

class SpecialReminderEngine(AdvancedEngine):
    name="special_reminder"; description="特殊提醒"
    def run(self, new_wb, rule, log_cb=None):
        from openpyxl.utils import get_column_letter
        desc_text=self.config.get('description','需人工复核')
        trigger=self.config.get('trigger','always')
        threshold=self.config.get('threshold')
        # 使用规则数据源定位目标区域（旧文件用于跳转）
        ds=rule.data_source; sheet=ds.get('sheet','')
        if not sheet or sheet not in new_wb.sheetnames: return []
        ws=new_wb[sheet]
        mode=ds.get('mode','offset')
        if mode=='shift': return []  # GUI已拦截，防御性跳过
        # 用 DataLocator 复用规则定位逻辑
        locator=DataLocator()
        target_cfg=ds.get('target',{})
        if mode=='offset':
            anchor_cfg=ds.get('anchor',{})
            search_in=ds.get('search_in','')
            merged=locator._merge_search_in(anchor_cfg, search_in)
            pos=locator._locate_offset(ws, {'anchor':anchor_cfg,'search_in':search_in,'target':target_cfg})
            if not pos or 'error' in pos: return []
            r1=r2=pos.get('row',1); c1=c2=pos.get('col',1)
        elif mode=='range':
            pos=locator._locate_range(ws, {'anchor':ds.get('anchor',{}),'search_in':ds.get('search_in',''),'target':target_cfg})
            if not pos or 'error' in pos: return []
            r1=pos.get('row',1); c1=pos.get('col',1)
            r2=r1+pos.get('row_count',1)-1; c2=c1+pos.get('col_count',1)-1
        elif mode=='intersection':
            pos=locator._locate_intersection(ws, ds)
            if not pos or 'error' in pos: return []
            r1=r2=pos.get('row',1); c1=c2=pos.get('col',1)
        else: return []
        r1,r2=min(r1,r2),max(r1,r2); c1,c2=min(c1,c2),max(c1,c2)
        alerts=[]
        for r in range(r1,r2+1):
            for c in range(c1,c2+1):
                cell=ws.cell(r,c); v=cell.value
                hit=False
                if trigger=='always': hit=True
                elif trigger=='nonempty': hit=v is not None and str(v).strip()!=''
                elif trigger=='gt': hit=v is not None and isinstance(v,(int,float)) and threshold is not None and v>float(threshold)
                elif trigger=='lt': hit=v is not None and isinstance(v,(int,float)) and threshold is not None and v<float(threshold)
                if hit:
                    addr=f"{get_column_letter(c)}{r}"
                    alerts.append({'sheet':sheet,'address':addr,'type':'特殊提醒',
                        'desc':f'{desc_text} [{addr}: {v}]','advanced_check':True})
        return alerts

class DataTrendEngine(AdvancedEngine):
    name="data_trend"; description="数据趋势"
    def run(self, new_wb, rule, log_cb=None):
        import math, statistics
        from openpyxl.utils import get_column_letter
        cfg=self.config
        ds=rule.data_source; cur_sheet=ds.get('sheet',''); hist_sheet=cfg.get('history_sheet','')
        if not cur_sheet or cur_sheet not in new_wb.sheetnames: return []
        if not hist_sheet or hist_sheet not in new_wb.sheetnames: return []
        # 待检数据：使用规则数据源定位
        locator=DataLocator()
        mode=ds.get('mode','offset')
        if mode=='shift': return []
        def _collect_rule_ds(ws):
            """用规则数据源收集待检值"""
            vals=[]; target_cfg=ds.get('target',{})
            if mode=='offset':
                pos=locator._locate_offset(ws, {'anchor':ds.get('anchor',{}),'search_in':ds.get('search_in',''),'target':target_cfg})
                if not pos or 'error' in pos: return vals, 'A1'
                r1=r2=pos.get('row',1); c1=c2=pos.get('col',1)
            elif mode=='range':
                pos=locator._locate_range(ws, {'anchor':ds.get('anchor',{}),'search_in':ds.get('search_in',''),'target':target_cfg})
                if not pos or 'error' in pos: return vals, 'A1'
                r1=pos.get('row',1); c1=pos.get('col',1)
                r2=r1+pos.get('row_count',1)-1; c2=c1+pos.get('col_count',1)-1
            elif mode=='intersection':
                pos=locator._locate_intersection(ws, ds)
                if not pos or 'error' in pos: return vals, 'A1'
                r1=r2=pos.get('row',1); c1=c2=pos.get('col',1)
            else: return vals, 'A1'
            r1,r2=min(r1,r2),max(r1,r2); c1,c2=min(c1,c2),max(c1,c2)
            for r in range(r1,r2+1):
                for c in range(c1,c2+1):
                    v=ws.cell(r,c).value
                    if isinstance(v,(int,float)): vals.append(float(v))
            return vals, f"{get_column_letter(c1)}{r1}"
        cur_ws=new_wb[cur_sheet]
        cur_vals, start_addr=_collect_rule_ds(cur_ws)
        # 历史数据：保留独立配置
        def _collect_hist(ws_name, acfg, tcfg):
            ws=new_wb[ws_name]; vals=[]
            anchor_text=acfg.get('text','').strip()
            ar=None
            if anchor_text:
                for r in range(1,(ws.max_row or 1)+1):
                    for c in range(1,(ws.max_column or 1)+1):
                        v=ws.cell(r,c).value
                        if v is not None and anchor_text in str(v).strip(): ar=(r,c); break
                    if ar: break
            if not ar: ar=(1,1)
            ro=tcfg.get('row_offset',0); co=tcfg.get('col_offset',0)
            rc=tcfg.get('row_count',1); cc=tcfg.get('col_count',1)
            sr=ar[0]+ro; sc=ar[1]+co
            r1=sr; r2=sr+max(rc,1)-1 if rc>=0 else sr
            c1=sc; c2=sc+max(cc,1)-1 if cc>=0 else sc
            r1,r2=min(r1,r2),max(r1,r2); c1,c2=min(c1,c2),max(c1,c2)
            for r in range(r1,r2+1):
                for c in range(c1,c2+1):
                    v=ws.cell(r,c).value
                    if isinstance(v,(int,float)): vals.append(float(v))
            return vals
        hist_vals=_collect_hist(hist_sheet, cfg.get('history_anchor',{}), cfg.get('history_target',{}))
        if not cur_vals: return []
        alerts=[]
        mt=cfg.get('metric_thresholds',{})
        cur_mean=statistics.mean(cur_vals); cur_std=statistics.pstdev(cur_vals) if len(cur_vals)>1 else 0
        cur_max=max(cur_vals); cur_min=min(cur_vals)
        sheet=cur_sheet; addr=start_addr
        def _add(metric_name, val, desc):
            alerts.append({'sheet':sheet,'address':addr,'type':'数据趋势',
                'desc':f'{metric_name}: {val} {desc}','advanced_check':True})
        mr=mt.get('mean_range',[])
        if len(mr)==2 and (cur_mean<float(mr[0]) or cur_mean>float(mr[1])):
            _add('均值',f'{cur_mean:.4f}',f'超出管制范围 [{mr[0]}~{mr[1]}]')
        sr=mt.get('stddev_range',[])
        if len(sr)==2 and (cur_std<float(sr[0]) or cur_std>float(sr[1])):
            _add('标准差',f'{cur_std:.4f}',f'超出管制范围 [{sr[0]}~{sr[1]}]')
        if 'max_limit' in mt and mt['max_limit']!='' and cur_max>float(mt['max_limit']):
            _add('最大值',f'{cur_max:.4f}',f'> 上限 {mt["max_limit"]}')
        if 'min_limit' in mt and mt['min_limit']!='' and cur_min<float(mt['min_limit']):
            _add('最小值',f'{cur_min:.4f}',f'< 下限 {mt["min_limit"]}')
        usl=mt.get('spec_usl',''); lsl=mt.get('spec_lsl','')
        cpk_min=mt.get('cpk_min','')
        if usl!='' and lsl!='' and cur_std>0:
            cpk=min((float(usl)-cur_mean)/(3*cur_std),(cur_mean-float(lsl))/(3*cur_std))
            if cpk_min!='' and cpk<float(cpk_min):
                _add('CPK',f'{cpk:.2f}',f'< 最低要求 {cpk_min}')
        if hist_vals and cfg.get('t_test',False):
            alpha=float(cfg.get('alpha',0.05))
            h_mean=statistics.mean(hist_vals); h_std=statistics.pstdev(hist_vals) if len(hist_vals)>1 else 0
            n1,n2=len(cur_vals),len(hist_vals)
            if n1>1 and n2>1 and cur_std>0 and h_std>0:
                se=math.sqrt(cur_std**2/n1 + h_std**2/n2)
                if se>0:
                    t_stat=abs(cur_mean-h_mean)/se
                    df=n1+n2-2
                    p_approx=2*math.exp(-0.717*t_stat-0.416*t_stat**2/max(df,1)) if t_stat<10 else 0.0001
                    p_approx=max(0.0001,min(1.0,p_approx))
                    if p_approx<alpha:
                        _add('t检验',f'p≈{p_approx:.4f}',f'< α={alpha}，与历史数据存在显著差异 (待检均值{cur_mean:.4f} vs 历史均值{h_mean:.4f})')
        # F检验（方差差异）
        if hist_vals and cfg.get('f_test',False):
            f_alpha=float(cfg.get('f_alpha',0.05))
            h_std=statistics.pstdev(hist_vals) if len(hist_vals)>1 else 0
            n1,n2=len(cur_vals),len(hist_vals)
            if n1>1 and n2>1 and cur_std>0 and h_std>0:
                f_stat=max(cur_std**2,1e-12)/max(h_std**2,1e-12)
                # 近似F分布p值（简化近似）
                df1,df2=n1-1,n2-1
                x=df2/(df2+df1*f_stat)
                # 不完全Beta函数近似
                if x<0.5:
                    p_f=2*x*(1+0.5*x*(df1-2)/df2)
                else:
                    p_f=2*(1-x)*(1+0.5*(1-x)*(df2-2)/df1)
                p_f=max(0.0001,min(1.0,p_f))
                if p_f<f_alpha:
                    _add('F检验',f'F={f_stat:.4f},p≈{p_f:.4f}',f'< α={f_alpha}，方差存在显著差异 (待检方差{cur_std**2:.6f} vs 历史方差{h_std**2:.6f})')
        # UCL/LCL管制限
        ucl_cfg=cfg.get('ucl_lcl',{})
        if ucl_cfg.get('enabled',False):
            sigma_level=ucl_cfg.get('sigma_level','3σ')
            if sigma_level=='特殊':
                # 特殊模式：直接使用用户输入值
                ucl_op=ucl_cfg.get('ucl_op','≥'); ucl_val=ucl_cfg.get('ucl_val','')
                lcl_op=ucl_cfg.get('lcl_op','≤'); lcl_val=ucl_cfg.get('lcl_val','')
                if ucl_val!='':
                    op_funcs={'≥':lambda a,b:a>=b,'>':lambda a,b:a>b,'=':lambda a,b:a==b,'<':lambda a,b:a<b,'≤':lambda a,b:a<=b}
                    func=op_funcs.get(ucl_op,lambda a,b:a>=b)
                    if func(cur_mean,float(ucl_val)):
                        _add('UCL',f'{cur_mean:.4f} {ucl_op} {ucl_val}',f'超出UCL管制限')
                if lcl_val!='':
                    op_funcs={'≥':lambda a,b:a>=b,'>':lambda a,b:a>b,'=':lambda a,b:a==b,'<':lambda a,b:a<b,'≤':lambda a,b:a<=b}
                    func=op_funcs.get(lcl_op,lambda a,b:a<=b)
                    if func(cur_mean,float(lcl_val)):
                        _add('LCL',f'{cur_mean:.4f} {lcl_op} {lcl_val}',f'超出LCL管制限')
            else:
                # σ倍数模式：从历史数据计算控制限
                h_mean=statistics.mean(hist_vals) if hist_vals else cur_mean
                h_std=statistics.pstdev(hist_vals) if len(hist_vals)>1 else cur_std
                n=int(sigma_level.replace('σ','')) if sigma_level.replace('σ','').isdigit() else 3
                ucl=h_mean+n*h_std; lcl=h_mean-n*h_std
                if cur_mean>ucl:
                    _add('UCL',f'{cur_mean:.4f} > {ucl:.4f}',f'超出{n}σ上管制限 (UCL={ucl:.4f})')
                elif cur_mean<lcl:
                    _add('LCL',f'{cur_mean:.4f} < {lcl:.4f}',f'超出{n}σ下管制限 (LCL={lcl:.4f})')
        return alerts

ADVANCED_ENGINE_REGISTRY={'filename_check':FileNameCheckEngine,'special_reminder':SpecialReminderEngine,'data_trend':DataTrendEngine}

class CheckItemConfig:
    def __init__(self, check_type="value", enabled=True, expect="same", options=None):
        self.check_type=check_type; self.enabled=enabled; self.expect=expect; self.options=options or {}
    def to_dict(self): return {"check_type":self.check_type,"enabled":self.enabled,"expect":self.expect,"options":self.options}
    @classmethod
    def from_dict(cls, data): return cls(data.get("check_type","value"),data.get("enabled",True),data.get("expect","same"),data.get("options",{}))
class AdvancedEngineConfig:
    def __init__(self, engine_type='', enabled=False, config=None):
        self.engine_type=engine_type; self.enabled=enabled; self.config=config or {}
    def to_dict(self): return {"engine_type":self.engine_type,"enabled":self.enabled,"config":self.config}
    @classmethod
    def from_dict(cls, data): return cls(data.get("engine_type",""),data.get("enabled",False),data.get("config",{}))
class CheckRule:
    def __init__(self, rule_name="", data_source=None, checks=None, advanced_engines=None):
        self.rule_name=rule_name; self.data_source=data_source or {}; self.checks=checks or []; self.advanced_engines=advanced_engines or []
    def to_dict(self): return {"rule_name":self.rule_name,"data_source":self.data_source,"checks":[c.to_dict() for c in self.checks],"advanced_engines":[e.to_dict() for e in self.advanced_engines]}
    @classmethod
    def from_dict(cls, data):
        obj=cls(data.get("rule_name",""),data.get("data_source",{}),[CheckItemConfig.from_dict(cd) for cd in data.get("checks",[])])
        obj.advanced_engines=[AdvancedEngineConfig.from_dict(ed) for ed in data.get("advanced_engines",[])]
        return obj
class CheckProject:
    def __init__(self, project_name="", description="", version="1.0", rules=None):
        self.project_name=project_name; self.description=description; self.version=version; self.rules=rules or []
    def to_dict(self): return {"project_name":self.project_name,"description":self.description,"version":self.version,"rules":[r.to_dict() for r in self.rules]}
    @classmethod
    def from_dict(cls, data): return cls(data.get("project_name",""),data.get("description",""),data.get("version","1.0"),[CheckRule.from_dict(rd) for rd in data.get("rules",[])])

class OpenpyxlComparer:
    def __init__(self, old_path, new_path, log_callback=None, progress_callback=None,
                 check_options=None, plugin_manager=None, progress_mode_fn=None,
                 check_project=None, stop_event=None, mode='diff', color_tolerance=0):
        self.old_path=old_path; self.new_path=new_path
        self.log=log_callback if log_callback else print
        self.progress=progress_callback if progress_callback else lambda v,s: None
        self.progress_mode=progress_mode_fn if progress_mode_fn else lambda m: None
        self.stop_event=stop_event if stop_event else threading.Event()
        self.mode=mode
        self.diffs=[]; self.sheet_diffs=[]; self.stats={'total_cells':0,'diff_cells':0,'added_sheets':[],'removed_sheets':[],'images_diff':0}
        self.old_rich={}; self.new_rich={}
        self.check_options=dict(DEFAULT_CHECK_OPTIONS); self.check_options.update(check_options or {})
        self.plugin_manager=plugin_manager; self.check_project=check_project
        self.sheet_order=[]; self._log_buffer=[]; self._last_gui_update=0
        self.color_tolerance = color_tolerance
        self.old_cache = None
        self.new_cache = None

    def _color_pair_equal(self, c1, c2, context='text'):
        # 颜色统一解析为 RGB 后精确判定：旧单元格颜色用旧工作簿缓存解析，
        # 新单元格颜色用新工作簿缓存解析（各自的主题表/自定义调色板独立，不能交叉）。
        if c1 is None and c2 is None: return True
        r_old = self.old_cache.resolve_color(c1, context) if self.old_cache is not None else None
        r_new = self.new_cache.resolve_color(c2, context) if self.new_cache is not None else None
        if r_old and r_new:
            ro, rn = r_old.upper(), r_new.upper()
            if ro == rn: return True
            # 调色板量化吸附：一侧为 indexed 时，文件可能把直接RGB吸附到最近的自定义调色板色
            # （视觉同色，hex 差 20~45，如 B4EBB4<->CCFFCC、D9D9D9<->C0C0C0），按欧氏距离容忍；
            # 真实颜色差异（红<->绿/蓝）距离 300+，远超阈值，不受影响。
            def _is_indexed(c):
                if isinstance(c, dict): return c.get('type') == 'indexed'
                return getattr(c, 'type', None) == 'indexed'
            if (_is_indexed(c1) or _is_indexed(c2)) and rgb_euclidean(ro, rn) <= 60:
                return True
            # 仅当某一侧为 theme+tint 回算颜色时，允许 1 通道级舍入噪声
            # （ECMA tint 经浮点 HLS 回 8bit 在极端 tint 下固有 +-1 抖动，非数据容差）
            def _is_tinted(c):
                if isinstance(c, dict):
                    return c.get('type') == 'theme' and abs(float(c.get('tint') or 0)) > 1e-9
                ct = getattr(c, 'type', None)
                return ct == 'theme' and abs(float(getattr(c, 'tint', 0) or 0)) > 1e-9
            if (_is_tinted(c1) or _is_tinted(c2)) and rgb_channel_close(ro, rn):
                return True
            return False
        # 仅一侧可解析：无法可靠判定，退回存储签名比较（不猜测）
        if isinstance(c1, dict) and isinstance(c2, dict): return c1 == c2
        return color_signature(c1) == color_signature(c2)

    def _flush_log(self, force=False):
        now=time.time()
        if not force and (now-self._last_gui_update)<0.15: return
        if self._log_buffer:
            for m in self._log_buffer:  # 时间戳/着色统一由 log() 处理
                try: self.log(m)
                except: pass
            self._log_buffer.clear(); self._last_gui_update=now
    def _buf_log(self, msg): self._log_buffer.append(msg)
    def _with_heartbeat(self, label, fn):
        stop_event=threading.Event(); t0=time.time()
        def _heartbeat():
            while not stop_event.is_set():
                stop_event.wait(1.0)
                if stop_event.is_set(): break
                try:
                    self._buf_log(f"⏳ {label}... 已耗时 {time.time()-t0:.0f}s"); self._flush_log(force=True)
                except: pass
        hb=threading.Thread(target=_heartbeat,daemon=True); self.progress_mode('indeterminate'); hb.start()
        try: result=fn()
        finally: stop_event.set(); hb.join(timeout=2); self.progress_mode('determinate')
        return result
    def run(self):
        start_time=time.time(); old_wb,new_wb=self._load_workbooks()
        if not old_wb or not new_wb: return False
        try:
            self.old_wb_ref=old_wb; self.new_wb_ref=new_wb
            self.sheet_order=list(old_wb.sheetnames)
            self._run_diff_mode(old_wb,new_wb)
            before=len(self.diffs); self.diffs=merge_adjacent_diffs(self.diffs)
            merged=before-len(self.diffs)
            if merged>0: self._buf_log(f"相邻同类差异已合并：{before} → {len(self.diffs)} 条（合并 {merged} 条）"); self._flush_log(force=True)
            before_rt=len(self.diffs); self.diffs=dedup_rich_text_diffs(self.diffs)
            if before_rt-len(self.diffs)>0: self._buf_log(f"富文本重复差异已过滤：{before_rt-len(self.diffs)} 条"); self._flush_log(force=True)
        finally:
            if self.old_cache: self.old_cache.close()
            if self.new_cache: self.new_cache.close()
        self.progress(95,"生成报告..."); self._flush_log(force=True)
        total_time=time.time()-start_time; self.progress(100,"对比完成")
        self._buf_log(f"对比阶段耗时: {_fmt_duration(total_time)} | 差异: {self.stats['diff_cells']} 处单元格, {len(self.sheet_diffs)} 处Sheet"); self._flush_log(force=True)
        return True
    def _load_workbooks(self):
        try:
            def _load():
                self.progress(5,"正在加载旧版文件..."); self._flush_log(force=True)
                old_wb=load_workbook(self.old_path,data_only=False); self._buf_log(f"旧版加载完成: {len(old_wb.sheetnames)} 个sheet"); self._flush_log(force=True)
                self.progress(15,"正在加载新版文件..."); self._flush_log(force=True)
                new_wb=load_workbook(self.new_path,data_only=False); self._buf_log(f"新版加载完成: {len(new_wb.sheetnames)} 个sheet"); self._flush_log(force=True)
                # 懒加载标记：data_only=True 副本在 shift 规则首次遇到公式格时才加载
                self.old_wb_values = None; self.new_wb_values = None
                self.progress(20,"正在解析富文本..."); self._flush_log(force=True)
                self.old_rich=parse_rich_text_from_xlsx(self.old_path); self.new_rich=parse_rich_text_from_xlsx(self.new_path)
                self._buf_log(f"富文本解析完成：旧版 {sum(len(v) for v in self.old_rich.values())} 个，新版 {sum(len(v) for v in self.new_rich.values())} 个"); self._flush_log(force=True)
                self.old_cache = WorkbookStyleCache(self.old_path)
                self.new_cache = WorkbookStyleCache(self.new_path)
                self._buf_log(f"样式解析完成：旧版主题色{len(self.old_cache.theme)}组、自定义格式{len(self.old_cache.num_fmts)}个；新版主题色{len(self.new_cache.theme)}组、自定义格式{len(self.new_cache.num_fmts)}个")
                self._flush_log(force=True)
                return old_wb,new_wb
            return self._with_heartbeat("加载中", _load)
        except Exception as e:
            self._buf_log(f"加载工作簿失败: {e}"); self._flush_log(force=True); return None,None
    def _run_diff_mode(self, old_wb, new_wb):
        self._compare_sheets(old_wb,new_wb); total=len(old_wb.sheetnames); start_time=time.time()
        for idx,sheet_name in enumerate(old_wb.sheetnames,1):
            if self.stop_event.is_set(): self._buf_log(f"用户请求停止，已跳过剩余 {total-idx+1} 个sheet"); self._flush_log(force=True); raise KeyboardInterrupt
            pct=25+int(55*idx/total); self.progress(pct,f"对比 {sheet_name}... ({idx}/{total})"); self._flush_log(force=True)
            if sheet_name in new_wb.sheetnames: self._compare_worksheet(old_wb[sheet_name],new_wb[sheet_name],sheet_name)
            self._buf_log(f"已完成 {sheet_name} ({idx}/{total})，累计耗时 {time.time()-start_time:.0f}s"); self._flush_log(force=True)
        if self.plugin_manager and self.plugin_manager.plugins:
            self.progress(85,"执行数据检查插件..."); self._buf_log(f"执行 {len(self.plugin_manager.plugins)} 个检查插件..."); self._flush_log(force=True)
            for diff in self.plugin_manager.run_checks(old_wb,new_wb,self._buf_log):
                self.diffs.append({'sheet':'🔍 数据检查','address':diff.get('rule_name',''),'type':diff['type'],'desc':diff['desc']})
        if self.check_project:
            self._run_advanced_engines(new_wb)
    def _run_advanced_engines(self, new_wb):
        if not self.check_project: return
        for rule in self.check_project.rules:
            for ecfg in rule.advanced_engines:
                if not ecfg.enabled: continue
                cls=ADVANCED_ENGINE_REGISTRY.get(ecfg.engine_type)
                if not cls: continue
                engine=cls(ecfg.config)
                try:
                    for alert in engine.run(new_wb, rule, self._buf_log):
                        self.diffs.append({'sheet':alert.get('sheet',''),'address':alert.get('address',''),
                            'type':alert.get('type','高级检查'),'desc':alert.get('desc',''),
                            'advanced_check':True,'rule_name':rule.rule_name})
                    self._buf_log(f" [高级检查] {ecfg.engine_type} 完成")
                except Exception as e:
                    self._buf_log(f" [高级检查] {ecfg.engine_type} 失败: {e}")
        self._flush_log(force=True)
    def _apply_rule_filter(self, diffs, old_wb, new_wb):
        diff_type_map={'内容变化':'value','公式变化':'formula','字体变化':'font','填充变化':'fill','边框变化':'border','对齐变化':'alignment','数字格式变化':'number_format','合并新增':'merged_cells','合并删除':'merged_cells','行高变化':'row_height','列宽变化':'col_width','图片新增':'images','图片变动':'images','图片尺寸变化':'images','条件格式新增':'conditional_format','条件格式删除':'conditional_format','条件格式修改':'conditional_format','条件格式变化':'conditional_format','富文本变化':'rich_text','单元格新增':'value','单元格删除':'value'}
        rule_addr_map={}; shift_new_map={}; shift_old_map={}; locator=DataLocator()
        for rule in self.check_project.rules:
            ds=rule.data_source
            if ds.get('mode')=='shift':
                if ds.get('old_sheet') or ds.get('old_anchor'):
                    self._buf_log(f"规则[{rule.rule_name}] 为旧版双区域shift结构，已不支持，请重新编辑该规则"); self._flush_log(force=True); continue
                sheet=ds.get('sheet','')
                if sheet not in old_wb.sheetnames or sheet not in new_wb.sheetnames: continue
                hdr_t=ds.get('header_target',{}); rows=ds.get('rows','')
                o_ac=locator._merge_search_in(ds.get('anchor',{}), ds.get('search_in',''))
                o_loc=locator._range_cfg(old_wb[sheet], o_ac, hdr_t, '标题行范围')
                n_loc=locator._range_cfg(new_wb[sheet], o_ac, hdr_t, '标题行范围')
                if 'error' in o_loc or 'error' in n_loc:
                    self._buf_log(f"规则[{rule.rule_name}] 标题区域定位失败: {o_loc.get('error') or n_loc.get('error')}"); self._flush_log(force=True); continue
                rowset=self._parse_row_spec(rows)
                if not rowset:
                    self._buf_log(f"规则[{rule.rule_name}] 垂直范围无效或为空: {rows!r}，跳过该规则"); self._flush_log(force=True); continue
                o_start=o_loc['start']; n_start=n_loc['start']; hrc=hdr_t.get('row_count',1); hcc=hdr_t.get('col_count',1)
                # v3.30 固定偏移：扫描旧报告标题区有数据的列，每列 + shift_offset = 新报告对应列
                shift_offset = ds.get('shift_offset', 0)
                pairs = []
                old_data_cols = []
                for c in range(o_loc.get('c1',o_start[1]), o_loc.get('c2',o_start[1]+hcc-1)+1):
                    has_data = False
                    for dr in range(o_loc.get('r1',o_start[0]), o_loc.get('r2',o_start[0]+hrc-1)+1):
                        v = old_wb[sheet].cell(dr, c).value
                        if v is not None and str(v).strip() != '':
                            has_data = True; break
                    if has_data:
                        old_data_cols.append(c)
                        nc = c + shift_offset
                        if nc >= 1:
                            pairs.append((c, nc))
                if not pairs:
                    self._buf_log(f"规则[{rule.rule_name}] 旧报告标题区域无数据列，跳过该规则"); self._flush_log(force=True); continue
                plog = '; '.join(f"{get_column_letter(oc)}→{get_column_letter(nc)}" for oc, nc in pairs)
                self._buf_log(f"规则[{rule.rule_name}] 固定偏移={shift_offset}，配对 {len(pairs)} 列: {plog}")
                no_data_cols = [c for c in range(o_loc.get('c1',o_start[1]), o_loc.get('c2',o_start[1]+hcc-1)+1) if c not in set(old_data_cols)]
                if no_data_cols:
                    self._buf_log(f"  旧报告标题区空列(跳过): {','.join(get_column_letter(c) for c in no_data_cols)}")
                self._flush_log(force=True)
                ows=old_wb[sheet]; nws=new_wb[sheet]
                for oc,nc in pairs:
                    for r in sorted(rowset):
                        oa=cell_address(oc,r); na=cell_address(nc,r)
                        entry=(rule,sheet,oa,sheet,na)
                        shift_old_map.setdefault((sheet,oa),[]).append(entry)
                        shift_new_map.setdefault((sheet,na),[]).append(entry)
                    # v3.28 标题行自身（机台名/表头等 hrc 覆盖行，新旧按行序对应）也挂配对键：
                    # 列搬移后标题格走配对复核，避免同地址把两台不同机台对撞误报
                    for dr in range(o_loc.get('r1',o_start[0]), o_loc.get('r2',o_start[0]+hrc-1)+1):
                        ndr = dr - o_loc.get('r1',o_start[0]) + n_loc.get('r1',n_start[0])
                        oa=cell_address(oc,dr); na=cell_address(nc,ndr)
                        entry=(rule,sheet,oa,sheet,na)
                        shift_old_map.setdefault((sheet,oa),[]).append(entry)
                        shift_new_map.setdefault((sheet,na),[]).append(entry)
                    # 列宽 diff 挂在 {列}1：旧列宽对旧位置、新列宽对新位置
                    cw_entry=(rule,sheet,cell_address(oc,1),sheet,cell_address(nc,1))
                    shift_old_map.setdefault((sheet,cell_address(oc,1)),[]).append(cw_entry)
                    shift_new_map.setdefault((sheet,cell_address(nc,1)),[]).append(cw_entry)
            else:
                sheet=ds.get('sheet','')
                if sheet not in old_wb.sheetnames or sheet not in new_wb.sheetnames: continue
                locator.rules=[ds]
                old_data=locator.locate_all(old_wb).get(ds.get('name','')); new_data=locator.locate_all(new_wb).get(ds.get('name',''))
                if not old_data or not new_data: continue
                addresses=old_data.get('addresses') or [old_data.get('address')] if isinstance(old_data,dict) else None
                if not addresses: continue
                for addr in addresses:
                    if addr: rule_addr_map.setdefault((sheet,addr),[]).append(rule)
        for d in diffs:
            if d['sheet']=='🔍 数据检查': continue
            check_type=diff_type_map.get(d['type'])
            if not check_type and not d.get('advanced_check'): continue
            key=(d['sheet'],d['address'])
            shift_hits=(shift_old_map if d['type']=='单元格删除' else shift_new_map).get(key,[]); hits=rule_addr_map.get(key,[])
            if not shift_hits and not hits: continue
            s_all,s_any,s_fail=True,False,[]
            _seen_sr=set()
            for rule,o_sheet,o_addr,n_sheet,n_addr in shift_hits:
                _k=(rule.rule_name,n_addr)
                if _k in _seen_sr: continue
                _seen_sr.add(_k)
                s_any=True
                old_ws=old_wb[o_sheet]; new_ws=new_wb[n_sheet]
                oc=column_index_from_string(''.join(ch for ch in o_addr if ch.isalpha())); orow=int(''.join(ch for ch in o_addr if ch.isdigit()))
                nc=column_index_from_string(''.join(ch for ch in n_addr if ch.isalpha())); nrow=int(''.join(ch for ch in n_addr if ch.isdigit()))
                old_cell=old_ws.cell(row=orow,column=oc); new_cell=new_ws.cell(row=nrow,column=nc)
                for check in rule.checks:
                    if not check.enabled: continue
                    ct=check.check_type
                    # 列宽/行高 diff 是整列/整行维度载体，只跑对应检查项，避免误伤其他属性
                    if d['type']=='列宽变化' and ct!='col_width': continue
                    if d['type']=='行高变化' and ct!='row_height': continue
                    if ct in ('fill','font','border','number_format','row_height','col_width','alignment') and d.get('com_style'):
                        diff = self._compare_style_by_com(ct, d['com_style']['old'], d['com_style']['new'], check.options)
                        desc = diff if diff else self._com_style_same_desc(ct, d['com_style']['new'])
                    else:
                        diff=self._compare_by_check_type(ct,old_cell,new_cell,check.options,old_ws,new_ws,o_addr,o_sheet,new_address=n_addr,new_sheet_name=n_sheet)
                        desc=self._build_diff_desc(ct,old_cell,new_cell,diff,False,self)
                    ok=(check.expect=='same' and diff is None) or (check.expect=='different' and diff is not None)
                    if not ok:
                        s_all=False
                        type_name=TYPE_DISPLAY.get(ct, ct)
                        s_fail.append(f"{type_name} {check.expect}: {desc}")
            if s_any:
                d['rule_name']=shift_hits[0][0].rule_name; d['rule_expect']='AND'
                d['rule_diff_desc']='\n'.join(s_fail)
                if s_all: d['rule_pass']=True
                continue
            a_all,a_any,a_fail=True,False,[]
            for rule in hits:
                a_any=True
                old_ws=old_wb[d['sheet']]; new_ws=new_wb[d['sheet']]
                col_str=''.join(ch for ch in d['address'] if ch.isalpha()); row_str=''.join(ch for ch in d['address'] if ch.isdigit())
                if not col_str or not row_str: continue
                col=column_index_from_string(col_str); row=int(row_str)
                old_cell=old_ws.cell(row=row,column=col); new_cell=new_ws.cell(row=row,column=col)
                for check in rule.checks:
                    if not check.enabled: continue
                    ct=check.check_type
                    # 列宽/行高 diff 是整列/整行维度载体，只跑对应检查项，避免误伤其他属性
                    if d['type']=='列宽变化' and ct!='col_width': continue
                    if d['type']=='行高变化' and ct!='row_height': continue
                    if ct in ('fill','font','border','number_format','row_height','col_width','alignment') and d.get('com_style'):
                        diff = self._compare_style_by_com(ct, d['com_style']['old'], d['com_style']['new'], check.options)
                        desc = diff if diff else self._com_style_same_desc(ct, d['com_style']['new'])
                    else:
                        diff=self._compare_by_check_type(ct,old_cell,new_cell,check.options,old_ws,new_ws,d['address'],d['sheet'])
                        desc=self._build_diff_desc(ct,old_cell,new_cell,diff,False,self)
                    ok=(check.expect=='same' and diff is None) or (check.expect=='different' and diff is not None)
                    if not ok:
                        a_all=False
                        type_name=TYPE_DISPLAY.get(ct, ct)
                        a_fail.append(f"{type_name} {check.expect}: {desc}")
            if a_any:
                d['rule_name']=hits[0].rule_name; d['rule_expect']='AND'
                d['rule_diff_desc']='\n'.join(a_fail)
                if a_all: d['rule_pass']=True
        # ---- 未命中规则的样式类 diff：用 COM 数据重跑对比 ----
        style_types = {'fill','font','border','number_format','row_height','col_width','alignment'}
        type_map = {v:k for k,v in ExcelCOMVerifier.STYLE_TYPE_MAP.items()}
        for d in diffs:
            if d.get('rule_name'): continue
            if d.get('advanced_check'): continue
            if d['sheet'] == '🔍 数据检查': continue
            ct = type_map.get(d.get('type',''))
            if not ct or ct not in style_types: continue
            if not d.get('com_style'): continue
            diff_desc = self._compare_style_by_com(ct, d['com_style']['old'], d['com_style']['new'])
            if diff_desc is None:
                d['rule_pass'] = True
                d['rule_name'] = '样式显示一致'
                d['rule_diff_desc'] = f"[{TYPE_DISPLAY.get(ct, ct)}] {self._com_style_same_desc(ct, d['com_style']['new'])}"
            else:
                d['rule_name'] = '常规'
                d['rule_diff_desc'] = f"[{TYPE_DISPLAY.get(ct, ct)}] {diff_desc}"
        # 列宽/行高 diff 的 desc 统一用 COM 显示值（XML stored 宽度含 padding，与 Excel 显示不一致）
        for d in diffs:
            if d.get('type') not in ('行高变化','列宽变化') or not d.get('com_style'): continue
            _ct = 'row_height' if d['type']=='行高变化' else 'col_width'
            _dd = self._compare_style_by_com(_ct, d['com_style']['old'], d['com_style']['new'])
            d['desc'] = _dd if _dd else self._com_style_same_desc(_ct, d['com_style']['new'])
        # 保存 shift 配对表，供 debug 导出读配对旧格
        self._last_shift_old_map = shift_old_map; self._last_shift_new_map = shift_new_map

    @staticmethod
    def _build_diff_desc(check_type, old_cell, new_cell, diff_result, is_exempted, comparer=None):
        if diff_result is not None: desc = diff_result
        elif check_type == 'formula': desc = f"公式: {formula_text(old_cell.value)} → {formula_text(new_cell.value)}"
        elif comparer and check_type in ('value','number_format'): desc = comparer._format_pair(check_type, old_cell, new_cell)
        elif check_type == 'rich_text': desc = "富文本: 一致"
        elif check_type == 'merged_cells': desc = "合并单元格: 一致"
        elif check_type == 'images': desc = "图片: 一致"
        elif check_type == 'conditional_format': desc = "条件格式: 一致"
        elif check_type == 'font': desc = f"字体: {old_cell.font.name}; 字号: {old_cell.font.size}"
        elif check_type == 'fill': desc = "填充: 一致"
        elif check_type == 'border': desc = "边框: 一致"
        elif check_type == 'alignment': desc = "对齐: 一致"
        elif check_type == 'row_height': desc = "行高: 一致"
        elif check_type == 'col_width':
            cl = get_column_letter(old_cell.column)
            desc = f"列宽({cl}): 一致"
        else: desc = f"{OpenpyxlComparer._fmt_val(old_cell.value)} → {OpenpyxlComparer._fmt_val(new_cell.value)}"
        return desc
    @staticmethod
    def _fmt_val(v):
        # 浮点值显示：10位有效数字去浮点尾数；小数截合理位，大数不会变科学计数法
        return f"{v:.10g}" if isinstance(v,float) else str(v)
    def _format_pair(self, check_type, old_cell, new_cell):
        ov, nv = old_cell.value, new_cell.value
        if check_type == 'formula': return f"公式: {formula_text(ov)} → {formula_text(nv)}"
        if check_type == 'number_format':
            nf1=self._get_actual_number_format(old_cell,'old'); nf2=self._get_actual_number_format(new_cell,'new')
            return f"数字格式: {format_numfmt_readable(nf1)} → {format_numfmt_readable(nf2)}"
        return f"{self._fmt_val(ov)} → {self._fmt_val(nv)}"
    @staticmethod
    def _parse_row_spec(spec):
        """'3,5-7,13-14' → {3,5,6,7,13,14}；非法返回空集合"""
        rows=set()
        for part in str(spec or '').split(','):
            part=part.strip()
            if not part: continue
            m=re.match(r'^(\d+)(?:-(\d+))?$', part)
            if not m: return set()
            a=int(m.group(1)); b=int(m.group(2) or a)
            if b<a: a,b=b,a
            if a<1: return set()
            rows.update(range(a,b+1))
        return rows
    def _normalize_number_format(self, fmt):
        if fmt is None: return ''
        s=re.sub(r'\s','',fmt); s=re.sub(r'\\(.)',r'\1',s); return s.lower()
    def _format_numfmt_readable(self, fmt):
        """数字格式码 -> 可读描述（委托模块级函数）"""
        return format_numfmt_readable(fmt)
    def _get_actual_number_format(self, cell, which):
        cache = self.old_cache if which == 'old' else self.new_cache
        if cache is None:
            try: return cell.number_format or 'General'
            except Exception: return 'General'
        return cache.resolve_numfmt(cell)
    def _fmt_decimals(self,fmt):
        # 数字格式码 -> 显示小数位；非数值格式返回 -1
        s=str(fmt or 'General')
        segs=[x for x in s.split(';') if x.strip()] or [s]
        seg=segs[0]
        base=seg.split(']')[-1]
        if re.search(r'[ymdhs]',base,re.I): return -1          # 日期时间，不放宽
        m=re.findall(r'\d',seg.split('.')[1].split(';')[0]) if '.' in seg else []
        if '%' in seg:
            return -1                                             # 百分比：不做 round 豁免，直接用存储值对比
        if m: return len(m)                                     # 0.0000 -> 4位
        return 0                                                # 整数显示（#,##0 等）
    @staticmethod
    def _value_decimals(v):
        # 数值本身的小数位数（以待测端粘贴值的精度定义 round 位数）
        if isinstance(v,bool) or not isinstance(v,(int,float)): return None
        f=v if isinstance(v,float) else float(v)
        s=repr(f)
        if 'e' in s or 'E' in s:
            m,e=s.split('e') if 'e' in s else s.split('E')
            exp=int(e)
            d=len(m.split('.')[1]) if '.' in m else 0
            return max(0,d-exp)
        if '.' in s:
            frac=s.split('.')[1]
            return 0 if frac.strip('0')=='' else len(frac)
        return 0
    def _display_equivalent(self,old_cell,new_cell,old_sheet,new_sheet,old_val=None,new_val=None):
        # 两格值严格不等时的浮点尾差豁免：
        # shift 场景旧端常为公式（缓存值带完整浮点尾数），新端是粘贴值。
        # 规则：仅当双方小数位数悬殊（max>=10，一边为浮点展开长尾）时，按【较少者】N 把双方 round N 位再比；
        #       N 由数值本身定义，不依赖数字格式（百分比/公式格格式不再挡豁免）；
        #       新格为日期/时间显示格式时不放宽（避免日期序列数误豁免）。
        v1=old_val if old_val is not None else old_cell.value
        v2=new_val if new_val is not None else new_cell.value
        if isinstance(v1,bool) or isinstance(v2,bool): return False
        if not (isinstance(v1,(int,float)) and isinstance(v2,(int,float))): return False
        try:
            f2=self._get_actual_number_format(new_cell,'new')
            if self._fmt_decimals(f2)<0 and re.search(r'[ymdhs]',str(f2 or '').split(';')[0].split(']')[-1],re.I):
                return False                                    # 新格按日期/时间显示：不放宽
            n1=self._value_decimals(v1); n2=self._value_decimals(v2)
            ns=[x for x in (n1,n2) if x is not None]
            if ns and max(ns)>=10:
                # 位数悬殊才 round 截齐；两边都是正常精度时位数不同即真差异（如 199.8 vs 199.85）
                n=min(ns)
                if round(v1,n)==round(v2,n): return True
            # 兜底：相对容差 1e-9（round 未覆盖的极端 ulp 尾差）
            return abs(v1-v2)<=1e-9*max(abs(v1),abs(v2),1.0)
        except Exception: return False
    @staticmethod
    def _try_number(s):
        """字符串转数字，无法转换则返回原值"""
        try: return int(s)
        except ValueError:
            try: return float(s)
            except ValueError: return s
    def _compare_worksheet(self, old_ws,new_ws,sheet_name):
        opts=self.check_options
        _,_,old_rmax,old_cmax=self._real_data_range(old_ws); _,_,new_rmax,new_cmax=self._real_data_range(new_ws)
        max_row=min(max(old_rmax,new_rmax),500000); max_col=min(max(old_cmax,new_cmax),500)
        if max_row==0 or max_col==0: return
        self._buf_log(f" Sheet大小: {max_row}行 x {max_col}列")
        check_value=opts.get('value',True); check_formula=opts.get('formula',True); check_font=opts.get('font',True); check_fill=opts.get('fill',True); check_border=opts.get('border',True); check_align=opts.get('alignment',True); check_nf=opts.get('number_format',True)
        # 预扫描本 sheet 的数组/共享公式区域（区域内成员格只存缓存值，须按锚点公式比对）
        if self.old_cache is not None: self.old_cache.get_formula_refs(sheet_name)
        if self.new_cache is not None: self.new_cache.get_formula_refs(sheet_name)
        skip_value = not (check_value or check_formula)
        batch_size=200; row_count=0
        for row_idx in range(1,max_row+1):
            if self.stop_event.is_set(): raise KeyboardInterrupt
            old_row_data=[]; new_row_data=[]
            if row_idx<=old_rmax:
                rows=list(old_ws.iter_rows(min_row=row_idx,max_row=row_idx,min_col=1,max_col=max_col,values_only=False)); old_row_data=rows[0] if rows else []
            if row_idx<=new_rmax:
                rows=list(new_ws.iter_rows(min_row=row_idx,max_row=row_idx,min_col=1,max_col=max_col,values_only=False)); new_row_data=rows[0] if rows else []
            for col_idx in range(1,max_col+1):
                old_cell=old_row_data[col_idx-1] if col_idx-1<len(old_row_data) else None
                new_cell=new_row_data[col_idx-1] if col_idx-1<len(new_row_data) else None
                if old_cell is None and new_cell is None: continue
                if old_cell is None or new_cell is None:
                    old_val=old_cell.value if old_cell else None; new_val=new_cell.value if new_cell else None
                    if old_val is None and new_val is None: continue
                    addr=cell_address(col_idx,row_idx)
                    if old_cell is None: self.diffs.append({'sheet':sheet_name,'address':addr,'type':'单元格新增','desc':f'新增: {new_val}'})
                    else: self.diffs.append({'sheet':sheet_name,'address':addr,'type':'单元格删除','desc':f'删除: {old_val}'})
                    self.stats['diff_cells']+=1; continue
                old_v=old_cell.value; new_v=new_cell.value; addr=cell_address(col_idx,row_idx)
                if old_v is None and new_v is None:
                    if not (check_font or check_fill or check_border or check_align or check_nf): continue
                # 数组/共享公式成员格：XML 里只有缓存值，缓存值差异是计算结果噪声，
                # 必须按锚点公式文本判定（与普通公式"公式相同不报值差"的逻辑一致）
                in_array=False
                if self.old_cache is not None and self.new_cache is not None:
                    of_in = self.old_cache.find_array_formula(sheet_name,row_idx,col_idx)
                    nf_in = self.new_cache.find_array_formula(sheet_name,row_idx,col_idx)
                    of_anchor = formula_text(old_v); nf_anchor = formula_text(new_v)
                    # 锚点格自身公式优先于区域扫描（openpyxl 直接读 <f> 文本最可靠）；
                    # 扫描结果仅用于成员格（成员格 openpyxl 读到的是缓存值）
                    of_txt = of_anchor or of_in; nf_txt = nf_anchor or nf_in
                    if of_txt or nf_txt:
                        in_array=True
                        if of_txt and nf_txt:
                            if normalize_formula(of_txt)!=normalize_formula(nf_txt):
                                self.diffs.append({'sheet':sheet_name,'address':addr,'type':'公式变化','desc':f"数组公式: {of_txt} → {nf_txt}"}); self.stats['diff_cells']+=1
                        elif of_txt and not nf_txt:
                            self.diffs.append({'sheet':sheet_name,'address':addr,'type':'公式变化','desc':f"公式状态: 是 → 否（原数组公式区域）"}); self.stats['diff_cells']+=1
                        else:
                            self.diffs.append({'sheet':sheet_name,'address':addr,'type':'公式变化','desc':f"公式状态: 否 → 是（新数组公式区域）"}); self.stats['diff_cells']+=1
                if not in_array and not skip_value:
                    val_diff=self._get_cell_diff(old_cell,new_cell)
                    if val_diff:
                        self.diffs.append({'sheet':sheet_name,'address':addr,'type':'内容变化','desc':val_diff}); self.stats['diff_cells']+=1; continue
                if in_array and (check_value or check_formula):
                    # 数组公式区域内：公式已判过，缓存值差异一律跳过，继续检查格式
                    pass
                elif check_formula and not check_value and not in_array:
                    of=old_cell.value if isinstance(old_cell.value,str) and old_cell.value.startswith('=') else None
                    nf=new_cell.value if isinstance(new_cell.value,str) and new_cell.value.startswith('=') else None
                    if of!=nf:
                        self.diffs.append({'sheet':sheet_name,'address':addr,'type':'公式变化','desc':f"公式: {of} → {nf}"}); self.stats['diff_cells']+=1; continue
                if check_font:
                    fdiff=self._cmp_font(old_cell,new_cell)
                    if fdiff: self.diffs.append({'sheet':sheet_name,'address':addr,'type':'字体变化','desc':fdiff})
                if check_fill:
                    ffdiff=self._cmp_fill(old_cell,new_cell)
                    if ffdiff: self.diffs.append({'sheet':sheet_name,'address':addr,'type':'填充变化','desc':ffdiff})
                if check_border:
                    bdiff=self._cmp_border(old_cell,new_cell)
                    if bdiff: self.diffs.append({'sheet':sheet_name,'address':addr,'type':'边框变化','desc':bdiff})
                if check_align:
                    adiff=self._cmp_alignment(old_cell.alignment,new_cell.alignment)
                    if adiff: self.diffs.append({'sheet':sheet_name,'address':addr,'type':'对齐变化','desc':adiff})
                if check_nf:
                    nf1=self._get_actual_number_format(old_cell,'old')
                    nf2=self._get_actual_number_format(new_cell,'new')
                    if self._normalize_number_format(nf1)!=self._normalize_number_format(nf2):
                        self.diffs.append({'sheet':sheet_name,'address':addr,'type':'数字格式变化','desc':f'数字格式: {self._format_numfmt_readable(nf1)} → {self._format_numfmt_readable(nf2)}'})
                row_count+=1
                if row_count%batch_size==0:
                    self.progress(25+int(55*(row_idx/max_row)),f" {sheet_name}: {row_idx}/{max_row}行..."); self._flush_log(force=True)
        if opts.get('rich_text',True):
            old_sr=self.old_rich.get(sheet_name,{}); new_sr=self.new_rich.get(sheet_name,{})
            for ref in set(old_sr)|set(new_sr):
                diff=compare_rich_text_runs(old_sr.get(ref),new_sr.get(ref))
                if diff: self.diffs.append({'sheet':sheet_name,'address':ref,'type':'富文本变化','desc':diff})
        if opts.get('merged_cells',True): self._compare_merged_cells(old_ws,new_ws,sheet_name)
        if opts.get('row_height',True) or opts.get('col_width',True): self._compare_row_col_dimensions(old_ws,new_ws,sheet_name)
        if opts.get('images',True): self._compare_images(old_ws,new_ws,sheet_name)
        if opts.get('conditional_format',True): self._compare_conditional_formats(old_ws,new_ws,sheet_name)
    def _get_cell_diff(self,c1,c2):
        def value_to_str(v):
            if v is None: return ''
            if hasattr(v,'text'): return str(v.text)   # ArrayFormula 对象
            return str(v)
        v1=c1.value; v2=c2.value; opts=self.check_options
        f1=formula_text(v1); f2=formula_text(v2)
        if opts.get('formula',True):
            if f1 and f2 and normalize_formula(f1)==normalize_formula(f2): return None
        n1=value_to_str(v1); n2=value_to_str(v2)
        if bool(f1)!=bool(f2): return f"公式状态: {'是' if f1 else '否'} → {'是' if f2 else '否'}"
        if f1 and f2 and normalize_formula(f1)!=normalize_formula(f2): return f"公式: {f1} → {f2}"
        if n1!=n2:
            # 数值尾差豁免：两边都是纯数值（非公式）时，按新端数值小数位 round 后一致则不报
            if not f1 and not f2 and isinstance(v1,(int,float)) and not isinstance(v1,bool) and isinstance(v2,(int,float)) and not isinstance(v2,bool):
                if self._display_equivalent(c1,c2,'','',v1,v2): return None
            return f"{n1[:120]} → {n2[:120]}"
        return None
    def _cmp_font(self, old_cell, new_cell):
        changes=[]
        f1 = self.old_cache.resolve_font(old_cell, old_cell.font) if self.old_cache else None
        f2 = self.new_cache.resolve_font(new_cell, new_cell.font) if self.new_cache else None
        if f1 is None or f2 is None: return None
        n1=f1.get('name'); n2=f2.get('name')
        if n1 and n2 and n1!=n2:
            if not self.old_cache._font_names_equivalent(n1, n2):
                changes.append(f"字体: {n1}→{n2}")
        s1=f1.get('size') or 11; s2=f2.get('size') or 11
        if s1!=s2: changes.append(f"字号: {s1}→{s2}")
        b1=bool(f1.get('bold')); b2=bool(f2.get('bold'))
        if b1!=b2: changes.append(f"加粗: {b1}→{b2}")
        i1=bool(f1.get('italic')); i2=bool(f2.get('italic'))
        if i1!=i2: changes.append(f"斜体: {i1}→{i2}")
        u1=bool(f1.get('underline')); u2=bool(f2.get('underline'))
        if u1!=u2: changes.append(f"下划线: {u1}→{u2}")
        if not self._color_pair_equal(f1.get('color'), f2.get('color'), 'text'):
            changes.append(f"颜色: {self._fmt_color_dict(f1.get('color'))}→{self._fmt_color_dict(f2.get('color'))}")
        return '; '.join(changes) if changes else None

    def _fmt_color_dict(self, cd):
        if cd is None: return '无'
        t = cd.get('type')
        if t == 'rgb': return f"RGB:{cd.get('rgb')}"
        if t == 'theme': return f"Theme:{cd.get('theme')}(tint={cd.get('tint',0):.2f})"
        if t == 'indexed': return f"Indexed:{cd.get('idx')}"
        if t == 'auto': return '自动'
        return '未知颜色'
    def _cmp_fill(self, old_cell, new_cell):
        fl1 = self.old_cache.resolve_fill(old_cell) if self.old_cache else None
        fl2 = self.new_cache.resolve_fill(new_cell) if self.new_cache else None
        if fl1 is None or fl2 is None: return None
        t1 = fl1.get('pattern') or 'none'; t2 = fl2.get('pattern') or 'none'
        if t1 == 'none' and t2 == 'none': return None
        if t1 != t2:
            return f"填充类型: {t1}→{t2}"
        if not self._color_pair_equal(fl1.get('fg'), fl2.get('fg'), 'bg'):
            return f"填充色: {self._fmt_color_dict(fl1.get('fg'))}→{self._fmt_color_dict(fl2.get('fg'))}"
        return None
    def _cmp_border(self, old_cell, new_cell):
        bd1 = self.old_cache.resolve_border(old_cell) if self.old_cache else None
        bd2 = self.new_cache.resolve_border(new_cell) if self.new_cache else None
        if bd1 is None or bd2 is None: return None
        parts=[]; side_names={'left':'左','right':'右','top':'上','bottom':'下'}
        style_names={'thin':'细线','medium':'中等线','dashed':'虚线','dotted':'点线','double':'双线','thick':'粗线','dashDot':'点划线','dashDotDot':'双点划线','slantDashDot':'斜点划线','mediumDashed':'中等虚线','mediumDashDot':'中等点划线','mediumDashDotDot':'中等双点划线','hair':'发丝线'}
        for side in ['left','right','top','bottom']:
            s1=bd1.get(side,{}) or {}; s2=bd2.get(side,{}) or {}
            st1=s1.get('style'); st2=s2.get('style')
            color_same=self._color_pair_equal(s1.get('color'), s2.get('color'), 'text')
            if st1!=st2 or not color_same:
                def style_desc(st): return style_names.get(st,st or '无')
                parts.append(f"{side_names[side]}: {style_desc(st1)}/{self._fmt_color_dict(s1.get('color'))}→{style_desc(st2)}/{self._fmt_color_dict(s2.get('color'))}")
        return '; '.join(parts) if parts else None
    def _cmp_alignment(self,a1,a2):
        changes=[]
        h1=a1.horizontal or ''; h2=a2.horizontal or ''
        if h1!=h2: changes.append(f"水平: {h1 or '默认'}→{h2 or '默认'}")
        v1=a1.vertical or ''; v2=a2.vertical or ''
        if v1!=v2: changes.append(f"垂直: {v1 or '默认'}→{v2 or '默认'}")
        w1=a1.wrap_text or False; w2=a2.wrap_text or False
        if w1!=w2: changes.append(f"自动换行: {w1}→{w2}")
        return '; '.join(changes) if changes else None
    def _compare_row_col_dimensions(self,old_ws,new_ws,sheet_name):
        opts=self.check_options
        old_dims = self.old_cache.get_dims(sheet_name) if self.old_cache else {'rows':{},'cols':{},'def_h':15.0,'def_w':8.43}
        new_dims = self.new_cache.get_dims(sheet_name) if self.new_cache else {'rows':{},'cols':{},'def_h':15.0,'def_w':8.43}
        if opts.get('row_height',True):
            # 仅比较 customHeight=1 的显式行高；自动行高（ht 仅为缓存值，Excel 会重算）不报差异
            old_custom = {r:v[0] for r,v in old_dims['rows'].items() if v[1]}
            new_custom = {r:v[0] for r,v in new_dims['rows'].items() if v[1]}
            for row_idx in sorted(set(old_custom)|set(new_custom)):
                oh = old_custom.get(row_idx); nh = new_custom.get(row_idx)
                if oh is not None and nh is not None:
                    if abs(oh-nh) < 0.01: continue
                    desc=f'行高: {oh:g} → {nh:g}'
                elif oh is not None:
                    desc=f'行高: {oh:g} → 自动行高'
                else:
                    desc=f'行高: 自动行高 → {nh:g}'
                self.diffs.append({'sheet':sheet_name,'address':f"A{row_idx}",'type':'行高变化','desc':desc})
        if opts.get('col_width',True):
            old_cw = {c:v[0] for c,v in old_dims['cols'].items() if v[1]}
            new_cw = {c:v[0] for c,v in new_dims['cols'].items() if v[1]}
            for ci in sorted(set(old_cw)|set(new_cw)):
                ow = old_cw.get(ci); nw = new_cw.get(ci)
                letter = get_column_letter(ci)
                if ow is not None and nw is not None:
                    if abs(ow-nw) < 0.01: continue
                    desc=f'列宽({letter}): {ow:g} → {nw:g}'
                elif ow is not None:
                    desc=f'列宽({letter}): {ow:g} → 自动列宽'
                else:
                    desc=f'列宽({letter}): 自动列宽 → {nw:g}'
                self.diffs.append({'sheet':sheet_name,'address':cell_address(ci,1),'type':'列宽变化','desc':desc})
    def _compare_merged_cells(self,old_ws,new_ws,sheet_name):
        old_merged=set(str(m) for m in old_ws.merged_cells.ranges); new_merged=set(str(m) for m in new_ws.merged_cells.ranges)
        for addr in new_merged-old_merged: self.diffs.append({'sheet':sheet_name,'address':addr.split(':')[0],'type':'合并新增','desc':f'新增合并区域 {addr}'})
        for addr in old_merged-new_merged: self.diffs.append({'sheet':sheet_name,'address':addr.split(':')[0],'type':'合并删除','desc':f'删除合并区域 {addr}'})
    def _get_images_from_ws(self,ws):
        images=[]
        if hasattr(ws,'_images') and ws._images:
            for img in ws._images:
                try:
                    anchor=img.anchor
                    if anchor and hasattr(anchor,'_from'):
                        col=anchor._from.col+1; row=anchor._from.row+1; images.append((cell_address(col,row),img.width,img.height))
                except: pass
            if images: return sorted(images,key=lambda x:x[0])
        if hasattr(ws,'_drawing') and ws._drawing:
            for anchor in ws._drawing.anchors:
                if hasattr(anchor,'image'):
                    img=anchor.image
                    col=(anchor._from.col if hasattr(anchor,'_from') else 0)+1
                    row=(anchor._from.row if hasattr(anchor,'_from') else 0)+1
                    images.append((cell_address(col,row),img.width,img.height))
            return sorted(images,key=lambda x:x[0])
        return images
    def _compare_images(self,old_ws,new_ws,sheet_name):
        old_imgs=self._get_images_from_ws(old_ws); new_imgs=self._get_images_from_ws(new_ws)
        if not old_imgs and not new_imgs: return
        old_set={addr:(w,h) for addr,w,h in old_imgs}; new_set={addr:(w,h) for addr,w,h in new_imgs}
        old_addrs=set(old_set); new_addrs=set(new_set)
        added=new_addrs-old_addrs; removed=old_addrs-new_addrs; common=old_addrs&new_addrs
        changed=[]
        for addr in common:
            w1,h1=old_set[addr]; w2,h2=new_set[addr]
            if abs(w1-w2)/max(w1,w2)>0.01 or abs(h1-h2)/max(h1,h2)>0.01:
                if abs(w1-w2)>1 or abs(h1-h2)>1: changed.append((addr,w1,h1,w2,h2))
        if len(added)+len(removed)+len(changed)>0:
            self.stats['images_diff']+=len(added)+len(removed)+len(changed)
            for addr in sorted(added):
                w,h=new_set[addr]; self.diffs.append({'sheet':sheet_name,'address':addr,'type':'图片新增','desc':f'新增图片 ({w:.0f}x{h:.0f})'})
            for addr in sorted(removed):
                w,h=old_set[addr]; self.diffs.append({'sheet':sheet_name,'address':addr,'type':'图片变动','desc':f'图片移动 ({w:.0f}x{h:.0f})'})
            for addr,w1,h1,w2,h2 in sorted(changed,key=lambda x:x[0]):
                self.diffs.append({'sheet':sheet_name,'address':addr,'type':'图片尺寸变化','desc':f'图片尺寸: {w1:.0f}x{h1:.0f} → {w2:.0f}x{h2:.0f}'})
    def _compare_conditional_formats(self,old_ws,new_ws,sheet_name):
        # 按精确 sqref 配对：同范围比规则内容；仅新有=新增，仅旧有=删除
        #（首地址分组+双重循环会把不同范围交叉配对，产生"范围A→范围B"误报）
        old_map={str(cf.sqref):cf for cf in old_ws.conditional_formatting}
        new_map={str(cf.sqref):cf for cf in new_ws.conditional_formatting}
        for rng in sorted(set(old_map)|set(new_map)):
            addr=rng.split()[0].split(':')[0]
            old_cf=old_map.get(rng); new_cf=new_map.get(rng)
            if old_cf is None:
                details=self._cf_rules_brief(new_cf.rules)
                self.diffs.append({'sheet':sheet_name,'address':addr,'type':'条件格式新增','desc':f"新增规则（范围 {rng}）：{details}",'short_desc':f"新增规则（{rng}）"})
            elif new_cf is None:
                details=self._cf_rules_brief(old_cf.rules)
                self.diffs.append({'sheet':sheet_name,'address':addr,'type':'条件格式删除','desc':f"删除规则（范围 {rng}）：{details}",'short_desc':f"删除规则（{rng}）"})
            else:
                od=self._cf_rule_diffs(old_cf.rules,new_cf.rules)
                if od:
                    self.diffs.append({'sheet':sheet_name,'address':addr,'type':'条件格式变化','desc':f"规则变化（范围 {rng}）：{od}",'short_desc':f"规则变化（{rng}）"})
    # Excel 内置条件格式标准色 -> 中文色名
    _CF_COLOR_NAMES={'FFC7CE':'浅红','9C0006':'深红','C6EFCE':'浅绿','006100':'深绿','FFEB9C':'浅黄','9C6500':'深黄',
                     'FF0000':'红色','00B050':'绿色','FFFF00':'黄色','0070C0':'蓝色','FFFFFF':'白色','000000':'黑色'}

    def _cf_color_desc(self, c):
        """dxf 颜色对象 -> 可读描述（内置标准色显示中文名）；不可读返回 ''"""
        if c is None: return ''
        try:
            rgb = getattr(c, 'rgb', None)
            if isinstance(rgb, str):
                s=rgb.strip()
                # 合法 rgb 只含 hex 字符且长度为 6(RRGGBB) 或 8(AARRGGBB)；含其它字符的是 openpyxl 解析失败的错误串
                if len(s) not in (6,8) or not all(ch in '0123456789ABCDEF' for ch in s.upper()): return ''
                hx=s[-6:].upper()
                if hx=='000000' and (len(s)==6 or s[:2]=='00'): return ''   # 未设色的占位（00000000 等），非真实黑色
                name=self._CF_COLOR_NAMES.get(hx)
                return f"{name}({hx})" if name else hx
        except Exception: pass
        return ''

    def _cf_dxf_brief(self, dxf):
        """条件格式 dxf 样式 -> 简要中文（填充/字色/加粗/斜体）"""
        if dxf is None: return '无格式'
        parts=[]
        try:
            if getattr(dxf, 'fill', None) is not None:
                fg = self._cf_color_desc(getattr(dxf.fill, 'fgColor', None)) or self._cf_color_desc(getattr(dxf.fill, 'start_color', None)) or self._cf_color_desc(getattr(dxf.fill, 'bgColor', None)) or self._cf_color_desc(getattr(dxf.fill, 'end_color', None))
                if fg: parts.append('填充'+fg)
        except Exception: pass
        try:
            if getattr(dxf, 'font', None) is not None:
                fc = self._cf_color_desc(getattr(dxf.font, 'color', None))
                if fc: parts.append('字色'+fc)
                if getattr(dxf.font, 'b', False): parts.append('加粗')
                if getattr(dxf.font, 'i', False): parts.append('斜体')
        except Exception: pass
        return '/'.join(parts) if parts else '无格式'

    _CF_TYPE_MAP={'cellIs':'单元格值','expression':'公式','containsText':'包含文本','notContainsText':'不含文本',
                  'containsBlanks':'包含空值','notContainsBlanks':'不含空值','containsErrors':'包含错误','notContainsErrors':'不含错误',
                  'duplicateValues':'重复值','uniqueValues':'唯一值','top10':'前N项','aboveAverage':'高于均值','belowAverage':'低于均值',
                  'colorScale':'色阶','dataBar':'数据条','iconSet':'图标集'}
    _CF_OP_MAP={'greaterThan':'大于','lessThan':'小于','equal':'等于','notEqual':'不等于','greaterThanOrEqual':'大于等于',
                'lessThanOrEqual':'小于等于','between':'介于','notBetween':'不介于','containsText':'包含'}

    def _cf_rule_summaries(self, rules):
        out=[]
        for r in rules:
            try:
                f=getattr(r,'formula',None)
                if isinstance(f,(list,tuple)): f=','.join(str(x) for x in f)
                rt=getattr(r,'type','?') or '?'; op=getattr(r,'operator',None)
                cond=self._CF_TYPE_MAP.get(rt,rt)
                # containsText/notContainsText 的 operator 与类型语义重复，不追加后缀
                if op and op in self._CF_OP_MAP and rt not in ('containsText','notContainsText'):
                    cond+='（'+self._CF_OP_MAP[op]+'）'
                dxf_obj=getattr(r,'dxf',None)
                has_fmt=bool(dxf_obj is not None and (getattr(dxf_obj,'fill',None) is not None or getattr(dxf_obj,'font',None) is not None))
                flabel = '阈值' if rt=='cellIs' else ('文本' if rt in ('containsText','notContainsText') else '公式')
                out.append({'cond':cond,
                            'formula':str(f) if f is not None else '',
                            'flabel':flabel,
                            'dxf':self._cf_dxf_brief(dxf_obj),'has_fmt':has_fmt,
                            'stop':bool(getattr(r,'stopIfTrue',False))})
            except Exception:
                out.append({'cond':'?','formula':'','flabel':'公式','dxf':'?','has_fmt':False,'stop':False})
        return out

    def _cf_rule_brief_one(self, s):
        """单条规则摘要 -> 可读明细：[条件类型] 公式xxx 填充浅红/字色深红/加粗；格式为空时不写"""
        bits=[]
        if s.get('formula'): bits.append(f"{s.get('flabel','公式')}{s['formula']}")
        dxf=s.get('dxf','')
        if dxf and dxf not in ('无格式','?'): bits.append(dxf)
        tail=('；'+'；'.join(bits)) if bits else ''
        stop='，为真则停止' if s.get('stop') else ''
        return f"[{s.get('cond','?')}]{tail}{stop}"

    def _cf_rules_brief(self, rules):
        """规则列表 -> 可读明细串（用于新增/删除时展示规则内容）"""
        summaries=self._cf_rule_summaries(rules)
        if not summaries: return '无规则'
        return '；'.join(f"规则{i+1}{self._cf_rule_brief_one(s)}" for i,s in enumerate(summaries))

    def _cf_rule_diffs(self, old_rules, new_rules):
        """逐条比对两侧规则，输出变化前→变化后的可读差异；无差异返回 ''"""
        o=self._cf_rule_summaries(old_rules); n=self._cf_rule_summaries(new_rules)
        parts=[]
        for i in range(max(len(o),len(n))):
            ro=o[i] if i<len(o) else None; rn=n[i] if i<len(n) else None
            if ro is None:
                parts.append(f"新增规则{i+1}{self._cf_rule_brief_one(rn)}"); continue
            if rn is None:
                parts.append(f"删除规则{i+1}{self._cf_rule_brief_one(ro)}"); continue
            if ro==rn: continue
            sub=[]
            if ro['cond']!=rn['cond']: sub.append(f"条件：{ro['cond']} → {rn['cond']}")
            if ro['formula']!=rn['formula']: sub.append(f"{rn.get('flabel','公式')}：{ro['formula'] or '无'} → {rn['formula'] or '无'}")
            # 新版有填充/字体对象但读不出颜色（openpyxl 解析主题色失败）时，跳过格式对比避免误报
            if ro['dxf']!=rn['dxf'] and not (rn['has_fmt'] and rn['dxf'] in ('无格式','?')):
                sub.append(f"格式：{ro['dxf']} → {rn['dxf']}")
            if ro['stop']!=rn['stop']: sub.append('停止：开 → 关' if ro['stop'] else '停止：关 → 开')
            parts.append(f"规则{i+1}（{rn['cond']}）：" + '；'.join(sub))
        return '；'.join(parts)

    def _compare_style_by_com(self, check_type, old_style, new_style, options=None):
        """基于 COM 采集的样式数据（dict）做对比。
        返回 None=一致，字符串=差异描述。"""
        if check_type == 'fill':
            changes = []
            oc = old_style.get('fill_color'); nc = new_style.get('fill_color')
            if oc and nc and oc != nc:
                changes.append(f"颜色: RGB:{oc} → RGB:{nc}")
            op = old_style.get('fill_pattern'); np_ = new_style.get('fill_pattern')
            if op is not None and np_ is not None and op != np_:
                # 两边颜色同为白色(FFFFFF)时实心≈无填充，图案差异不影响显示，豁免
                if not (oc and nc and oc == nc and str(oc).upper().endswith('FFFFFF')):
                    changes.append(f"图案: {op} → {np_}")
            return "; ".join(changes) if changes else None

        elif check_type == 'font':
            changes = []
            on = old_style.get('font_name'); nn = new_style.get('font_name')
            if on and nn and on != nn:
                changes.append(f"字体: {on} → {nn}")
            os_ = old_style.get('font_size'); ns = new_style.get('font_size')
            if os_ is not None and ns is not None and os_ != ns:
                changes.append(f"字号: {os_} → {ns}")
            for prop, label in [('font_bold','加粗'), ('font_italic','斜体'), ('font_underline','下划线')]:
                ov = old_style.get(prop); nv = new_style.get(prop)
                if ov is not None and nv is not None and bool(ov) != bool(nv):
                    changes.append(f"{label}: {bool(ov)} → {bool(nv)}")
            oc = old_style.get('font_color'); nc = new_style.get('font_color')
            if oc and nc and oc != nc:
                changes.append(f"颜色: RGB:{oc} → RGB:{nc}")
            return "; ".join(changes) if changes else None

        elif check_type == 'border':
            changes = []
            edge_names = {'left':'左', 'right':'右', 'top':'顶', 'bottom':'底'}
            for ename, elabel in edge_names.items():
                os = old_style.get(f'border_{ename}_style')
                ns = new_style.get(f'border_{ename}_style')
                if os is not None and ns is not None:
                    if os != ns:
                        changes.append(f"{elabel}边框样式: {os} → {ns}")
                    if os != -4142:
                        oc = old_style.get(f'border_{ename}_color')
                        ncc = new_style.get(f'border_{ename}_color')
                        if oc and ncc and oc != ncc:
                            changes.append(f"{elabel}边框颜色: RGB:{oc} → RGB:{ncc}")
                        ow = old_style.get(f'border_{ename}_weight')
                        nw = new_style.get(f'border_{ename}_weight')
                        if ow is not None and nw is not None and abs(ow - nw) > 0.01:
                            changes.append(f"{elabel}边框粗细: {ow} → {nw}")
            return "; ".join(changes) if changes else None

        elif check_type == 'number_format':
            of = old_style.get('num_format'); nf = new_style.get('num_format')
            if of is None or nf is None: return None
            if of != nf:
                return f"数字格式: {of} → {nf}"
            return None

        elif check_type == 'row_height':
            oh = old_style.get('row_height'); nh = new_style.get('row_height')
            if oh is None or nh is None: return None
            if abs(oh - nh) > 0.01:
                return f"行高: {oh:g} → {nh:g}"
            return None

        elif check_type == 'col_width':
            ow = old_style.get('col_width'); nw = new_style.get('col_width')
            if ow is None or nw is None: return None
            if abs(ow - nw) > 0.01:
                return f"列宽: {ow:g} → {nw:g}"
            return None

        elif check_type == 'alignment':
            changes = []
            prop_labels = {'h_align':'水平对齐', 'v_align':'垂直对齐', 'wrap_text':'自动换行'}
            for prop, label in prop_labels.items():
                ov = old_style.get(prop); nv = new_style.get(prop)
                if ov is not None and nv is not None and ov != nv:
                    changes.append(f"{label}: {ov} → {nv}")
            return "; ".join(changes) if changes else None

        return None

    def _com_style_same_desc(self, check_type, style):
        """用COM数据生成一致时的描述（两边相同，只列一边的值）"""
        if check_type == 'col_width':
            w = style.get('col_width')
            return f"列宽: {w:g}" if w is not None else "列宽: 一致"
        if check_type == 'row_height':
            h = style.get('row_height')
            return f"行高: {h:g}" if h is not None else "行高: 一致"
        if check_type == 'number_format':
            nf = style.get('num_format')
            return f"数字格式: {nf}" if nf else "数字格式: 一致"
        if check_type == 'font':
            parts = []
            if style.get('font_name'): parts.append(f"字体: {style['font_name']}")
            if style.get('font_size') is not None: parts.append(f"字号: {style['font_size']:g}")
            if style.get('font_bold'): parts.append("加粗")
            if style.get('font_italic'): parts.append("斜体")
            return "; ".join(parts) if parts else "字体: 一致"
        if check_type == 'fill':
            c = style.get('fill_color')
            return f"填充: RGB:{c}" if c else "填充: 一致"
        if check_type == 'border':
            return "边框: 显示一致"
        if check_type == 'alignment':
            return "对齐: 显示一致"
        return "显示一致"

    def _compare_by_check_type(self, check_type, old_cell, new_cell, options=None, old_ws=None, new_ws=None, address=None, sheet_name='', new_address=None, new_sheet_name=''):
        if new_address is None: new_address=address
        if not new_sheet_name: new_sheet_name=sheet_name
        if check_type=='value':
            # 数组/共享公式成员格只有缓存值，公式一致则视为值一致
            if self.old_cache is not None and self.new_cache is not None and address:
                of=self.old_cache.find_array_formula(sheet_name,old_cell.row,old_cell.column) or formula_text(old_cell.value)
                nf=self.new_cache.find_array_formula(new_sheet_name,new_cell.row,new_cell.column) or formula_text(new_cell.value)
                if of and nf and normalize_formula(of)==normalize_formula(nf): return None
            # 公式格：懒加载 data_only 副本取缓存值（首次遇到才加载）
            ov = old_cell.value; nv = new_cell.value
            if isinstance(ov, str) and ov.startswith('='):
                if self.old_wb_values is None:
                    def _load_old_vals(): return load_workbook(self.old_path, data_only=True)
                    self.old_wb_values = self._with_heartbeat("加载旧版缓存值副本", _load_old_vals)
                    self._buf_log(f"✓ 旧版缓存加载完成"); self._flush_log(force=True)
                try:
                    ov = self.old_wb_values[sheet_name].cell(old_cell.row, old_cell.column).value
                    if isinstance(ov, str): ov = self._try_number(ov)
                except Exception: pass
            if isinstance(nv, str) and nv.startswith('='):
                if self.new_wb_values is None:
                    def _load_new_vals(): return load_workbook(self.new_path, data_only=True)
                    self.new_wb_values = self._with_heartbeat("加载新版缓存值副本", _load_new_vals)
                    self._buf_log(f"✓ 新版缓存加载完成"); self._flush_log(force=True)
                try:
                    nv = self.new_wb_values[new_sheet_name].cell(new_cell.row, new_cell.column).value
                    if isinstance(nv, str): nv = self._try_number(nv)
                except Exception: pass
            if ov != nv:
                if self._display_equivalent(old_cell,new_cell,sheet_name,new_sheet_name,ov,nv): return None
                return f"{self._fmt_val(ov)} → {self._fmt_val(nv)}"
        elif check_type=='formula':
            of=formula_text(old_cell.value); nf=formula_text(new_cell.value)
            # 数组/共享公式成员格：回退到锚点公式区域匹配
            if of is None and self.old_cache is not None:
                of=self.old_cache.find_array_formula(sheet_name,old_cell.row,old_cell.column)
            if nf is None and self.new_cache is not None:
                nf=self.new_cache.find_array_formula(new_sheet_name,new_cell.row,new_cell.column)
            if normalize_formula(of)!=normalize_formula(nf): return f"公式: {of} → {nf}"
        elif check_type=='rich_text':
            if old_ws and new_ws:
                old_rich=self.old_rich.get(sheet_name,{}).get(address); new_rich=self.new_rich.get(new_sheet_name,{}).get(new_address)
                return compare_rich_text_runs(old_rich,new_rich)
        elif check_type=='font': return self._cmp_font(old_cell,new_cell)
        elif check_type=='fill': return self._cmp_fill(old_cell,new_cell)
        elif check_type=='border': return self._cmp_border(old_cell,new_cell)
        elif check_type=='alignment': return self._cmp_alignment(old_cell.alignment,new_cell.alignment)
        elif check_type=='number_format':
            nf1=self._get_actual_number_format(old_cell,'old')
            nf2=self._get_actual_number_format(new_cell,'new')
            if self._normalize_number_format(nf1)!=self._normalize_number_format(nf2): return f"数字格式: {self._format_numfmt_readable(nf1)} → {self._format_numfmt_readable(nf2)}"
        elif check_type=='merged_cells':
            om=str(old_cell.coordinate) if any(str(old_cell.coordinate) in str(m) for m in old_ws.merged_cells.ranges) else None
            nm=str(new_cell.coordinate) if any(str(new_cell.coordinate) in str(m) for m in new_ws.merged_cells.ranges) else None
            if om!=nm: return f"合并区域: {om} → {nm}"
        elif check_type=='row_height':
            od=self.old_cache.get_dims(sheet_name) if self.old_cache else {'rows':{}}
            nd=self.new_cache.get_dims(new_sheet_name) if self.new_cache else {'rows':{}}
            oi=od['rows'].get(old_cell.row); ni=nd['rows'].get(new_cell.row)
            oh=oi[0] if oi and oi[1] else None; nh=ni[0] if ni and ni[1] else None
            if oh is None and nh is None: return None
            if oh is not None and nh is not None:
                if abs(oh-nh)<0.01: return None
                return f"行高: {oh:g} → {nh:g}"
            return f"行高: {f'{oh:g}' if oh is not None else '自动'} → {f'{nh:g}' if nh is not None else '自动'}"
        elif check_type=='col_width':
            cl=get_column_letter(old_cell.column)
            od=self.old_cache.get_dims(sheet_name) if self.old_cache else {'cols':{}}
            nd=self.new_cache.get_dims(new_sheet_name) if self.new_cache else {'cols':{}}
            oi=od['cols'].get(old_cell.column); ni=nd['cols'].get(new_cell.column)
            ow=oi[0] if oi and oi[1] else None; nw=ni[0] if ni and ni[1] else None
            if ow is None and nw is None: return None
            if ow is not None and nw is not None:
                if abs(ow-nw)<0.01: return None
                return f"列宽({cl}): {ow:g} → {nw:g}"
            return f"列宽({cl}): {f'{ow:g}' if ow is not None else '自动'} → {f'{nw:g}' if nw is not None else '自动'}"
        elif check_type=='images':
            old_imgs=self._get_images_from_ws(old_ws); new_imgs=self._get_images_from_ws(new_ws)
            old_has=any(addr==address for addr,_,_ in old_imgs); new_has=any(addr==new_address for addr,_,_ in new_imgs)
            if old_has!=new_has: return f"图片存在: {old_has} → {new_has}"
        elif check_type=='conditional_format':
            # 条件格式按单元格独立比规则内容（类型/阈值/停止/格式），所有模式都不比范围：
            # 范围平移/扩展时原格规则不变即 same；被移出/新扩展的格子自身有 diff 时会自报删除/新增
            old_rng,old_rules=self._get_cfs_for_cell(old_ws,address)
            new_rng,new_rules=self._get_cfs_for_cell(new_ws,new_address)
            od=self._cf_rule_diffs(old_rules,new_rules)
            return f"规则变化: {od}" if od else None
        return None
    def _get_cfs_for_cell(self, ws, address):
        """返回 (范围描述, 该单元格命中的所有规则列表)；未命中返回 (None, [])。
        用 MultiCellRange 正确处理多区域 sqref（如 A1:B2 D10:E10）。"""
        ranges=[]; rules=[]
        for cf in ws.conditional_formatting:
            sq=str(cf.sqref)
            try:
                hit = address in MultiCellRange(sq)
            except Exception:
                hit = address in sq
            if hit:
                ranges.append(sq)
                try: rules.extend(list(cf.rules))
                except Exception: pass
        return ('；'.join(ranges) if ranges else None), rules
    def _compare_sheets(self,old_wb,new_wb):
        old_names=set(old_wb.sheetnames); new_names=set(new_wb.sheetnames)
        for name in sorted(new_names-old_names,key=lambda n:list(new_names).index(n)):
            self.sheet_diffs.append({'name':name,'type':'新增','desc':f'新版新增 Sheet: {name}'})
        for name in sorted(old_names-new_names,key=lambda n:list(old_names).index(n)):
            self.sheet_diffs.append({'name':name,'type':'删除','desc':f'旧版有但新版无: {name}'})
    def _row_has_data(self,ws,row,cols):
        return any(ws.cell(row,c).value is not None for c in range(1,cols+1))
    def _col_has_data(self,ws,col,rows):
        return any(ws.cell(r,col).value is not None for r in range(1,rows+1))
    def _real_data_range(self,ws):
        max_row=ws.max_row or 0; max_col=ws.max_column or 0
        if max_row==0 or max_col==0: return 0,0,1,1
        if max_row<=50000 and max_col<=200: return 1,1,max_row,max_col
        check_cols=min(max_col,500); lo,hi=1,max_row; real_max_row=1
        if max_row>100000 and not self._row_has_data(ws,100000,check_cols): hi=100000
        if hi>20000 and not self._row_has_data(ws,20000,check_cols): hi=20000
        if hi>5000 and not self._row_has_data(ws,5000,check_cols): hi=5000
        while lo<=hi:
            mid=(lo+hi)//2
            if self._row_has_data(ws,mid,check_cols): real_max_row=mid; lo=mid+1
            else: hi=mid-1
        check_rows=min(real_max_row,500); col_upper=min(max_col,500)
        if col_upper>200 and not self._col_has_data(ws,200,check_rows): col_upper=200
        if col_upper>50 and not self._col_has_data(ws,50,check_rows): col_upper=50
        real_max_col=1; lo,hi=1,col_upper
        while lo<=hi:
            mid=(lo+hi)//2
            if self._col_has_data(ws,mid,check_rows): real_max_col=mid; lo=mid+1
            else: hi=mid-1
        return 1,1,real_max_row,real_max_col

class ExcelCOMVerifier:
    """Excel COM 显示层数据采集器。
    利用已打开的 Excel 实例读取单元格渲染后的终值，写入 diff['com_style']。
    不做任何豁免/结论判定，仅提供真实显示数据供规则引擎比对。"""

    COM_STYLE_TYPES = {'填充变化','字体变化','边框变化','数字格式变化','行高变化','列宽变化','对齐变化'}
    # 样式类 check_type → type 名称映射（规则引擎用）
    STYLE_TYPE_MAP = {
        'fill': '填充变化', 'font': '字体变化', 'border': '边框变化',
        'number_format': '数字格式变化', 'row_height': '行高变化',
        'col_width': '列宽变化', 'alignment': '对齐变化'
    }

    def __init__(self, old_path, new_path, log_callback=None, progress_fn=None, progress_mode_fn=None):
        self.old_path = os.path.abspath(old_path)
        self.new_path = os.path.abspath(new_path)
        self.log = log_callback if log_callback else print
        self._progress = progress_fn if progress_fn else (lambda v,s=None: None)
        self._progress_mode = progress_mode_fn if progress_mode_fn else (lambda m: None)
        self.excel_app = None
        self.old_wb = None
        self.new_wb = None
        self._opened_old = False
        self._opened_new = False
        self._created_excel = False

    def collect_style_data(self, diffs, old_font_equiv=None, new_font_equiv=None):
        """采集样式类 diff 的 COM 真实显示数据，写入 diff['com_style'] = {'old':{},'new':{}}。
        不做任何豁免判定，仅提供数据。
        old_font_equiv / new_font_equiv: 字体等价表列表（每组为 set），用于字体名归一化。
        返回 (success_count, fail_count)。"""
        target = [d for d in diffs if d.get('sheet') and d.get('address')
                  and d.get('sheet') != '🔍 数据检查' and not d.get('advanced_check')]
        if not target:
            self.log("COM数据采集：无样式类差异项")
            self._progress(100,'COM采集跳过')
            return 0, 0

        # 按 sheet 聚合，减少 COM 跨 sheet 切换开销
        by_sheet = {}
        for d in target:
            by_sheet.setdefault(d['sheet'], []).append(d)

        self.log(f"COM数据采集：{len(target)} 条差异，分布在 {len(by_sheet)} 个 sheet，开始连接 Excel...")
        self._progress_indet('正在连接 Excel...')

        if not self._connect_excel():
            self.log("COM数据采集：无法连接 Excel，跳过")
            self._progress_det(100); return 0, len(target)
        if not self._ensure_workbooks():
            self.log("COM数据采集：无法打开工作簿，跳过")
            self._release_workbooks()
            self._progress_det(100); return 0, len(target)

        success = 0
        fail = 0
        total = len(target)
        done = 0
        try: self._progress_mode('determinate')
        except Exception: pass
        self._progress(0,'COM数据采集中...')

        for sheet, sheet_diffs in by_sheet.items():
            try:
                old_ws = self.old_wb.Worksheets(sheet)
                new_ws = self.new_wb.Worksheets(sheet)
            except Exception as e:
                self.log(f"  COM采集：sheet [{sheet}] 访问失败: {e}")
                fail += len(sheet_diffs)
                done += len(sheet_diffs)
                continue

            for d in sheet_diffs:
                addr = d.get('address', '')
                if ':' in addr:
                    addr = addr.split(':')[0]
                if not re.match(r'^[A-Za-z]+[0-9]+$', addr):
                    fail += 1; done += 1; continue
                try:
                    old_r = old_ws.Range(addr)
                    new_r = new_ws.Range(addr)
                    old_style = self._read_cell_style(old_r, old_font_equiv)
                    new_style = self._read_cell_style(new_r, new_font_equiv)
                    d['com_style'] = {'old': old_style, 'new': new_style}
                    success += 1
                except Exception:
                    fail += 1
                done += 1
                if done % 50 == 0 or done == total:
                    try: self._progress(int(100*done/total), f'COM采集 {done}/{total}')
                    except Exception: pass

        self._release_workbooks()
        self.log(f"COM数据采集完成：成功 {success} 条，失败 {fail} 条")
        try: self._progress(100,'COM采集完成')
        except Exception: pass
        return success, fail

    def _read_cell_style(self, rng, font_equiv=None):
        """从 COM Range 读取完整样式属性，返回 dict。
        font_equiv: 字体等价表，用于归一化字体名"""
        s = {}
        # --- 填充 ---
        try:
            s['fill_color'] = '%06X' % (int(rng.Interior.Color) & 0xFFFFFF)
        except Exception:
            s['fill_color'] = None
        try:
            s['fill_pattern'] = rng.Interior.Pattern
        except Exception:
            s['fill_pattern'] = None

        # --- 字体 ---
        try:
            raw_name = rng.Font.Name
            s['font_name_raw'] = raw_name
            s['font_name'] = self._normalize_font_name(raw_name, font_equiv)
        except Exception:
            s['font_name'] = None; s['font_name_raw'] = None
        for prop, key in [('Size','font_size'), ('Bold','font_bold'),
                          ('Italic','font_italic'), ('Underline','font_underline')]:
            try: s[key] = getattr(rng.Font, prop)
            except Exception: s[key] = None
        try:
            s['font_color'] = '%06X' % (int(rng.Font.Color) & 0xFFFFFF)
        except Exception:
            s['font_color'] = None

        # --- 边框（左/右/顶/底）---
        edge_ids = {'left':7, 'right':8, 'top':9, 'bottom':10}
        for ename, eid in edge_ids.items():
            try:
                b = rng.Borders(eid)
                s[f'border_{ename}_style'] = b.LineStyle
                try:
                    s[f'border_{ename}_color'] = '%06X' % (int(b.Color) & 0xFFFFFF)
                except Exception:
                    s[f'border_{ename}_color'] = None
                try:
                    s[f'border_{ename}_weight'] = b.Weight
                except Exception:
                    s[f'border_{ename}_weight'] = None
            except Exception:
                s[f'border_{ename}_style'] = None
                s[f'border_{ename}_color'] = None
                s[f'border_{ename}_weight'] = None

        # --- 数字格式 ---
        try:
            s['num_format'] = rng.NumberFormat
        except Exception:
            s['num_format'] = None

        # --- 行高 ---
        try:
            s['row_height'] = float(rng.RowHeight)
        except Exception:
            s['row_height'] = None

        # --- 列宽 ---
        try:
            s['col_width'] = float(rng.EntireColumn.ColumnWidth)
        except Exception:
            s['col_width'] = None

        # --- 对齐 ---
        for prop, key in [('HorizontalAlignment','h_align'),
                          ('VerticalAlignment','v_align'),
                          ('WrapText','wrap_text')]:
            try: s[key] = getattr(rng, prop)
            except Exception: s[key] = None

        return s

    @staticmethod
    def _normalize_font_name(name, font_equiv):
        """用字体等价表归一化字体名。
        找到 name 所在的等价组，返回组内第一个名字作为标准名。"""
        if not name or not font_equiv:
            return name
        for group in font_equiv:
            if name in group:
                return sorted(group)[0]
        return name

    def _progress_det(self, v):
        """切 determinate 并设值（0-100）"""
        try:
            self._progress_mode('determinate')
            self._progress(v,'高级审核中...')
        except Exception: pass

    def _progress_indet(self, label=''):
        try: self._progress_mode('indeterminate')
        except Exception: pass

    def _connect_excel(self):
        try:
            self.excel_app = win32com.client.GetActiveObject("Excel.Application")
            self.log(f"高级审核：已连接 Excel 实例（{self.excel_app.Caption}）")
            return True
        except Exception:
            pass
        try:
            self.excel_app = win32com.client.DispatchEx("Excel.Application")
            self._created_excel = True
            self.excel_app.Visible = False
            self.excel_app.DisplayAlerts = False
            self.log("高级审核：已创建新 Excel 实例")
            return True
        except Exception as e:
            self.log(f"高级审核：无法启动 Excel - {e}")
            return False

    def _ensure_workbooks(self):
        try:
            for wb in self.excel_app.Workbooks:
                if normalize_path(wb.FullName) == self.old_path:
                    self.old_wb = wb; break
            if self.old_wb is None:
                self.old_wb = self.excel_app.Workbooks.Open(self.old_path, ReadOnly=True)
                self._opened_old = True
                self.log(f"高级审核：已打开参考报告")

            for wb in self.excel_app.Workbooks:
                if normalize_path(wb.FullName) == self.new_path:
                    self.new_wb = wb; break
            if self.new_wb is None:
                self.new_wb = self.excel_app.Workbooks.Open(self.new_path, ReadOnly=True)
                self._opened_new = True
                self.log(f"高级审核：已打开待检报告")
            return True
        except Exception as e:
            self.log(f"高级审核：打开工作簿失败 - {e}")
            return False

    def _release_workbooks(self):
        try:
            if self._opened_old and self.old_wb:
                self.old_wb.Close(SaveChanges=False)
            if self._opened_new and self.new_wb:
                self.new_wb.Close(SaveChanges=False)
        except Exception:
            pass
        if self._created_excel and self.excel_app:
            try: self.excel_app.Quit()
            except Exception: pass
            self.log("高级审核：已关闭隐藏 Excel 实例")
        self.old_wb = None
        self.new_wb = None
        self.excel_app = None
        self._created_excel = False

    def _verify_single(self, diff):
        """已废弃：COM 不再做判定，仅通过 collect_style_data 采集数据。"""
        return False

    def _ole_to_rgb(self, ole_long):
        v = int(ole_long) & 0xFFFFFF
        return (v & 0xFF, (v >> 8) & 0xFF, (v >> 16) & 0xFF)

    def _cmp_fill_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        changes = []
        try:
            oc = self._ole_to_rgb(old_r.Interior.Color)
            nc = self._ole_to_rgb(new_r.Interior.Color)
            if oc != nc:
                changes.append(f"颜色: RGB:{oc:06X} → RGB:{nc:06X}")
        except Exception:
            pass
        try:
            op = old_r.Interior.Pattern
            np_ = new_r.Interior.Pattern
            if op != np_: changes.append(f"图案: {op} → {np_}")
        except Exception:
            pass
        return "; ".join(changes) if changes else ""
    def _cmp_font_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        changes = []
        try:
            oc = self._ole_to_rgb(old_r.Font.Color)
            nc = self._ole_to_rgb(new_r.Font.Color)
            if oc != nc:
                changes.append(f"颜色: RGB:{oc:06X} → RGB:{nc:06X}")
        except Exception:
            pass
        for prop, label in [('Name','字体'), ('Size','字号'), ('Bold','加粗'), ('Italic','斜体'), ('Underline','下划线')]:
            try:
                ov = getattr(old_r.Font, prop)
                nv = getattr(new_r.Font, prop)
                if ov != nv: changes.append(f"{label}: {ov} → {nv}")
            except Exception:
                pass
        return "; ".join(changes) if changes else ""
    def _cmp_border_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        changes = []
        edge_names = {7:'左', 8:'右', 9:'顶', 10:'底'}
        for edge_id in (7, 8, 9, 10):
            try:
                ob = old_r.Borders(edge_id)
                nb = new_r.Borders(edge_id)
                if ob.LineStyle != nb.LineStyle:
                    changes.append(f"{edge_names.get(edge_id,edge_id)}边框样式: {ob.LineStyle} → {nb.LineStyle}")
                if ob.LineStyle != -4142:
                    try:
                        oc = self._ole_to_rgb(ob.Color)
                        nc = self._ole_to_rgb(nb.Color)
                        if oc != nc: changes.append(f"{edge_names.get(edge_id,edge_id)}边框颜色: RGB:{oc:06X} → RGB:{nc:06X}")
                    except Exception:
                        pass
                    try:
                        if abs(ob.Weight - nb.Weight) > 0.01:
                            changes.append(f"{edge_names.get(edge_id,edge_id)}边框粗细: {ob.Weight} → {nb.Weight}")
                    except Exception:
                        pass
            except Exception:
                pass
        return "; ".join(changes) if changes else ""
    def _cmp_numfmt_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        try:
            of = old_r.NumberFormat
            nf = new_r.NumberFormat
            if of != nf: return f"数字格式: {format_numfmt_readable(of)} → {format_numfmt_readable(nf)}"
        except Exception:
            pass
        return ""
    def _cmp_rowheight_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        try:
            oh = old_r.RowHeight
            nh = new_r.RowHeight
            if abs(oh - nh) > 0.01: return f"行高: {oh} → {nh}"
        except Exception:
            pass
        return ""
    def _cmp_colwidth_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        try:
            ow = old_r.EntireColumn.Width
            nw = new_r.EntireColumn.Width
            if abs(ow - nw) > 0.01: return f"列宽: {ow} → {nw}"
        except Exception:
            pass
        return ""
    def _cmp_align_com(self, old_r, new_r):
        """返回差异描述或空字符串(一致)"""
        changes = []
        prop_labels = {'HorizontalAlignment':'水平对齐', 'VerticalAlignment':'垂直对齐', 'WrapText':'自动换行'}
        for prop, label in prop_labels.items():
            try:
                ov = getattr(old_r, prop)
                nv = getattr(new_r, prop)
                if ov != nv: changes.append(f"{label}: {ov} → {nv}")
            except Exception:
                pass
        return "; ".join(changes) if changes else ""
class CheckOptionsDialog(tb.Toplevel):
    def __init__(self,parent,current_options,com_on=True):
        super().__init__(parent); self.title("检测项目设置"); self.resizable(False,False); self.transient(parent); self.grab_set(); self.result=None; self.vars={}
        main=tb.Frame(self,padding=15); main.pack(fill='both',expand=True)
        # 高级审核区（置顶，内容检测之上）
        adv=tb.Labelframe(main,text="高级审核",padding=(8,5)); adv.grid(row=0,column=0,columnspan=3,sticky='ew',padx=5,pady=(0,8))
        self.com_var=tk.BooleanVar(value=bool(com_on))
        tb.Checkbutton(adv,text="开启高级审核（COM 显示层复核）",variable=self.com_var,bootstyle="round-toggle").pack(anchor='w')
        tb.Label(adv,text="高级审核需保持正常开启目标Excel以调用COM进行显示层高级审核",bootstyle="secondary",wraplength=520).pack(anchor='w',pady=(4,0))
        # 三个检测分组（内容/格式/结构），左右宽度不变
        for col,(gname,keys) in enumerate(CHECK_OPTION_GROUPS):
            lf=tb.Labelframe(main,text=gname,padding=(8,4)); lf.grid(row=1,column=col,sticky='nsew',padx=5,pady=5)
            for key in keys:
                var=tk.BooleanVar(value=current_options.get(key,True)); self.vars[key]=var
                tb.Checkbutton(lf,text=CHECK_OPTION_LABELS[key],variable=var,bootstyle="round-toggle").pack(anchor='w',pady=1)
        main.columnconfigure(0,weight=1); main.columnconfigure(1,weight=1); main.columnconfigure(2,weight=1)
        btn=tb.Frame(main); btn.grid(row=2,column=0,columnspan=3,pady=(12,0))
        tb.Button(btn,text="全选",width=8,command=lambda:[v.set(True) for v in self.vars.values()]).pack(side='left',padx=(0,5))
        tb.Button(btn,text="全不选",width=8,command=lambda:[v.set(False) for v in self.vars.values()]).pack(side='left',padx=(0,5))
        tb.Button(btn,text="取消",width=8,command=self._on_cancel).pack(side='right',padx=(5,0))
        tb.Button(btn,text="确定",bootstyle=PRIMARY,width=8,command=self._on_ok).pack(side='right')
        center_window(self,parent); self.wait_window()
    def _on_ok(self):
        self.result={k:v.get() for k,v in self.vars.items()}; self.com_on=self.com_var.get(); self.destroy()
    def _on_cancel(self):
        self.result=None; self.destroy()

class ComCheckDialog(tb.Toplevel):
    """高级审核开启前确认窗：Excel 过期激活警告 + COM 通道自检"""
    def __init__(self,parent):
        super().__init__(parent); self.title("高级审核确认"); self.resizable(False,False); self.transient(parent); self.grab_set()
        self.result=None; self._spawned_app=None
        main=tb.Frame(self,padding=15); main.pack(fill='both',expand=True)
        tb.Label(main,text="⚠ 为防止 Excel 软件过期导致激活弹窗阻止高级 COM 操作，\n请确认 Excel 已处于开启且可交互状态（无激活/登录弹窗遮挡）。",
                 bootstyle="warning",justify='left').pack(anchor='w',pady=(0,8))
        self.status=tb.Label(main,text="尚未检查 COM 通道",bootstyle="secondary"); self.status.pack(anchor='w',pady=4)
        bf=tb.Frame(main); bf.pack(fill='x',pady=(10,0))
        tb.Button(bf,text="检查COM通道",bootstyle="info-outline",width=14,command=self._check_com).pack(side='left',padx=(0,8))
        tb.Button(bf,text="取消",width=10,command=self._cancel).pack(side='right')
        tb.Button(bf,text="确认开启",bootstyle="success-outline",width=12,command=self._ok).pack(side='right',padx=(0,8))
        center_window(self,parent); self.wait_window()
    def _check_com(self):
        try:
            app=win32com.client.GetActiveObject("Excel.Application"); self._spawned=False
            try: n=app.Workbooks.Count
            except Exception: n='?'
            self.status.configure(text=f"√ COM通道正常（已连接 Excel，{n} 个工作簿打开）",bootstyle="success")
        except Exception:
            try:
                app=win32com.client.DispatchEx("Excel.Application"); app.Visible=True; self._spawned_app=app
                self.status.configure(text="√ COM通道正常（已新启动 Excel，请确认其无激活弹窗）",bootstyle="success")
            except Exception as e:
                self.status.configure(text=f"× COM不可用：{e}",bootstyle="danger")
    def _ok(self):
        if self._spawned_app is not None:
            try: self._spawned_app.Quit()
            except Exception: pass
        self.result=True; self.destroy()
    def _cancel(self):
        if self._spawned_app is not None:
            try: self._spawned_app.Quit()
            except Exception: pass
        self.result=False; self.destroy()

class ProjectListDialog(tb.Toplevel):
    def __init__(self,parent,project_dir):
        super().__init__(parent); self.title("选择规则配置"); self.geometry("450x400"); self.resizable(False,False); self.transient(parent); self.result=None; self.project_dir=project_dir
        try: self._build_ui()
        except Exception as e: messagebox.showerror("错误",f"加载配置列表失败: {e}"); self.destroy(); return
        center_window(self,parent)
        self.lift(); self.focus_force()
        try: self.grab_set()
        except Exception: pass
        self.wait_window()
    def _build_ui(self):
        frame=tb.Frame(self,padding=10); frame.pack(fill='both',expand=True)
        files=[]
        try:
            if not os.path.isdir(self.project_dir):
                tb.Label(frame,text=f"目录不存在: {self.project_dir}").pack(pady=20); tb.Button(frame,text="关闭",command=self.destroy).pack(pady=5); return
            files=[f for f in os.listdir(self.project_dir) if f.endswith('.json')]
        except: files=[]
        if not files:
            tb.Label(frame,text="程序目录下没有配置文件").pack(pady=20); tb.Button(frame,text="关闭",command=self.destroy).pack(pady=5); return
        lf=tb.Frame(frame); lf.pack(side='top',fill='both',expand=True)
        self.listbox=tk.Listbox(lf,font=("微软雅黑",10)); self.listbox.pack(side='left',fill='both',expand=True)
        sb=ttk.Scrollbar(lf,orient='vertical',command=self.listbox.yview); sb.pack(side='right',fill='y'); self.listbox.configure(yscrollcommand=sb.set)
        self.listbox.bind('<Double-Button-1>',lambda e:self._load())
        for f in sorted(files): self.listbox.insert('end',os.path.splitext(f)[0])
        btn=tb.Frame(frame); btn.pack(side='bottom',fill='x',pady=5)
        tb.Button(btn,text="确定",bootstyle=PRIMARY,command=self._load).pack(side='left',padx=5)
        tb.Button(btn,text="取消",bootstyle="outline",command=self.destroy).pack(side='right',padx=5)
    def _load(self):
        sel=self.listbox.curselection()
        if not sel: messagebox.showwarning("提示","请先选择一个配置"); return
        self.result=os.path.join(self.project_dir,self.listbox.get(sel[0])+".json"); self.destroy()

class CheckProjectDialog(tb.Toplevel):
    def __init__(self,parent,old_path,new_path,check_project=None):
        super().__init__(parent); self.title("检查项目集配置"); self.geometry("1100x700"); self.transient(parent); self.parent=parent; self.old_path=old_path; self.new_path=new_path; self.project=check_project or CheckProject(); self.result=None; self._build_ui(); self._refresh_rule_list(); center_window(self,parent)
        self.lift(); self.focus_force()
        try: self.grab_set()
        except Exception: pass
        self.wait_window()
    def _build_ui(self):
        info=tb.Labelframe(self,text="项目信息",padding=8); info.pack(fill='x',padx=10,pady=5)
        tb.Label(info,text="项目名称:").grid(row=0,column=0,sticky='w',padx=(0,5)); self.project_name_var=tk.StringVar(value=self.project.project_name); tb.Entry(info,textvariable=self.project_name_var,width=30).grid(row=0,column=1,sticky='w')
        tb.Label(info,text="版本:").grid(row=0,column=2,sticky='w',padx=(15,5)); self.version_var=tk.StringVar(value=self.project.version); tb.Entry(info,textvariable=self.version_var,width=10).grid(row=0,column=3,sticky='w')
        tb.Label(info,text="描述:").grid(row=0,column=4,sticky='w',padx=(15,5)); self.desc_var=tk.StringVar(value=self.project.description); tb.Entry(info,textvariable=self.desc_var,width=30).grid(row=0,column=5,sticky='w')
        lf=tb.Labelframe(self,text="规则列表",padding=8); lf.pack(fill='both',expand=True,padx=10,pady=5)
        self.tree=tb.Treeview(lf,columns=('rule_name','data_source','checks'),show='headings',height=15)
        self.tree.heading('rule_name',text='规则名称'); self.tree.heading('data_source',text='数据源'); self.tree.heading('checks',text='检查项数')
        self.tree.column('rule_name',width=200); self.tree.column('data_source',width=300); self.tree.column('checks',width=80,anchor='center')
        self.tree.pack(side='left',fill='both',expand=True)
        sb=tb.Scrollbar(lf,orient='vertical',command=self.tree.yview); sb.pack(side='right',fill='y'); self.tree.configure(yscrollcommand=sb.set)
        btn=tb.Frame(self,padding=5); btn.pack(fill='x',padx=10,pady=5)
        tb.Button(btn,text="添加规则",bootstyle=PRIMARY,command=self.add_rule).pack(side='left',padx=5)
        tb.Button(btn,text="复制规则",bootstyle="secondary",command=self.copy_rule).pack(side='left',padx=5)
        tb.Button(btn,text="编辑规则",bootstyle=INFO,command=self.edit_rule).pack(side='left',padx=5)
        tb.Button(btn,text="删除规则",bootstyle=DANGER,command=self.delete_rule).pack(side='left',padx=5)
        tb.Button(btn,text="上移",bootstyle="outline",command=lambda:self.move_rule(-1)).pack(side='left',padx=5)
        tb.Button(btn,text="下移",bootstyle="outline",command=lambda:self.move_rule(1)).pack(side='left',padx=5)
        tb.Separator(btn,orient='vertical').pack(side='left',fill='y',padx=10)
        tb.Button(btn,text="保存项目",bootstyle=SUCCESS,command=self.save_project).pack(side='left',padx=5)
        tb.Button(btn,text="加载项目",bootstyle="outline",command=self.load_project).pack(side='left',padx=5)
        tb.Button(btn,text="应用",bootstyle=INFO,command=self.apply_project).pack(side='right',padx=5)
    def _refresh_rule_list(self):
        self.tree.delete(*self.tree.get_children())
        for idx,rule in enumerate(self.project.rules):
            ds=rule.data_source
            if ds.get('mode')=='shift': ds_str=f"[搬移] {ds.get('sheet','?')} · 行 {ds.get('rows','?')}"
            else: ds_str=f"{ds.get('sheet','?')} | {ds.get('anchor',{}).get('text','?')}"
            self.tree.insert('','end',iid=str(idx),values=(rule.rule_name,ds_str,len(rule.checks)))
    def add_rule(self):
        if not os.path.isfile(self.old_path) and not os.path.isfile(self.new_path): messagebox.showwarning("提示","请先在主界面选择有效的Excel文件"); return
        dlg=RuleEditorDialog(self,self.old_path,self.new_path)
        if dlg.result is not None: self.project.rules.append(dlg.result); self._refresh_rule_list()
    def copy_rule(self):
        sel=self.tree.selection()
        if not sel: messagebox.showwarning("提示","请先选择要复制的规则"); return
        idx=int(sel[0]); orig=self.project.rules[idx]; new=copy.deepcopy(orig); new.rule_name=orig.rule_name+"_副本"; self.project.rules.append(new); self._refresh_rule_list(); self.tree.selection_set(str(len(self.project.rules)-1))
    def edit_rule(self):
        sel=self.tree.selection()
        if not sel: messagebox.showwarning("提示","请先选择要编辑的规则"); return
        idx=int(sel[0]); dlg=RuleEditorDialog(self,self.old_path,self.new_path,rule=self.project.rules[idx])
        if dlg.result is not None: self.project.rules[idx]=dlg.result; self._refresh_rule_list()
    def delete_rule(self):
        sel=self.tree.selection()
        if not sel: messagebox.showwarning("提示","请先选择要删除的规则"); return
        idx=int(sel[0])
        if messagebox.askyesno("确认","确定删除该规则？"): self.project.rules.pop(idx); self._refresh_rule_list()
    def move_rule(self,direction):
        sel=self.tree.selection()
        if not sel: return
        idx=int(sel[0]); ni=idx+direction
        if 0<=ni<len(self.project.rules): self.project.rules[idx],self.project.rules[ni]=self.project.rules[ni],self.project.rules[idx]; self._refresh_rule_list(); self.tree.selection_set(str(ni))
    def save_project(self):
        self.project.project_name=self.project_name_var.get(); self.project.version=self.version_var.get(); self.project.description=self.desc_var.get()
        name=self.project.project_name.strip()
        if not name: messagebox.showwarning("提示","项目名称不能为空"); return
        safe_name=re.sub(r'[\\/:*?"<>|]','_',name).strip()
        if not safe_name: messagebox.showwarning("提示","项目名称包含非法字符"); return
        if safe_name!=name:
            self.project.project_name=safe_name; self.project_name_var.set(safe_name); name=safe_name
        filepath=os.path.join(PROGRAM_DIR,f"{name}.json")
        if os.path.exists(filepath) and not messagebox.askyesno("提示",f"文件已存在，是否覆盖？\n{filepath}"): return
        payload=json.dumps(self.project.to_dict(),ensure_ascii=False,indent=2)
        try:
            with open(filepath,'w',encoding='utf-8') as f:
                f.write(payload); f.flush()
                try: os.fsync(f.fileno())
                except Exception: pass
            # 回读校验，确认文件真正落盘且 JSON 完整
            with open(filepath,'r',encoding='utf-8') as f: json.load(f)
        except PermissionError:
            messagebox.showerror("保存失败",f"没有写入权限或文件被其他程序占用：\n{filepath}\n\n请关闭已打开的同名文件，或以管理员身份运行。")
            return
        except Exception as e:
            messagebox.showerror("保存失败",f"配置写入失败：{e}\n目标路径：{filepath}")
            return
        messagebox.showinfo("成功",f"项目已保存到：\n{filepath}")
    def load_project(self):
        dlg=ProjectListDialog(self,PROGRAM_DIR)
        if dlg.result:
            try:
                with open(dlg.result,'r',encoding='utf-8') as f: data=json.load(f)
                self.project=CheckProject.from_dict(data); self.project_name_var.set(self.project.project_name); self.version_var.set(self.project.version); self.desc_var.set(self.project.description); self._refresh_rule_list(); messagebox.showinfo("成功","项目已加载")
            except Exception as e: messagebox.showerror("错误",f"加载失败：{str(e)}")
    def apply_project(self):
        self.project.project_name=self.project_name_var.get(); self.project.version=self.version_var.get(); self.project.description=self.desc_var.get(); self.result=self.project; self.destroy()

class FileNameCheckConfigDialog(tb.Toplevel):
    def __init__(self, parent, config, sheets):
        super().__init__(parent); self.title("文件名一致性配置"); self.geometry("600x550"); self.transient(parent); self.result=None
        self.sheets=sheets; self.segments_data=config.get('segments',[])
        main=tb.Frame(self,padding=15); main.pack(fill='both',expand=True)
        # Mode selection
        mode_frame=tb.Frame(main); mode_frame.pack(fill='x',pady=(0,8))
        self.mode_var=tk.StringVar(value='segment' if self.segments_data else 'template')
        tb.Radiobutton(mode_frame,text="模板模式",variable=self.mode_var,value='template',command=self._switch_mode).pack(side='left',padx=4)
        tb.Radiobutton(mode_frame,text="分段模式（推荐）",variable=self.mode_var,value='segment',command=self._switch_mode).pack(side='left',padx=4)
        # Template mode frame
        self.template_frame=tb.Frame(main)
        tb.Label(self.template_frame,text="模板格式 (用 {字段名} 占位):").pack(anchor='w')
        self.fmt_var=tk.StringVar(value=config.get('format_template','')); tb.Entry(self.template_frame,textvariable=self.fmt_var,width=50).pack(fill='x',pady=(2,8))
        tb.Label(self.template_frame,text="字段映射 (字段名 → Sheet + 单元格):",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(4,0))
        mf=tb.Frame(self.template_frame); mf.pack(fill='both',expand=True,pady=4)
        cols=('field','sheet','cell'); self.tv=tb.Treeview(mf,columns=cols,show='headings',height=6)
        for c,w in [('field',120),('sheet',180),('cell',80)]: self.tv.heading(c,text=c); self.tv.column(c,width=w)
        self.tv.pack(side='left',fill='both',expand=True); sb=tb.Scrollbar(mf,orient='vertical',command=self.tv.yview); sb.pack(side='right',fill='y'); self.tv.configure(yscrollcommand=sb.set)
        for m in config.get('field_mappings',[]): self.tv.insert('','end',values=(m.get('field',''),m.get('sheet',''),m.get('cell','')))
        bf=tb.Frame(self.template_frame); bf.pack(fill='x',pady=4)
        tb.Label(bf,text="字段:").pack(side='left'); self.f_var=tk.StringVar(); tb.Entry(bf,textvariable=self.f_var,width=10).pack(side='left',padx=2)
        tb.Label(bf,text="Sheet:").pack(side='left',padx=(8,0)); self.s_var=tk.StringVar(); self.s_cb=tb.Combobox(bf,textvariable=self.s_var,width=14,values=sheets); self.s_cb.pack(side='left',padx=2)
        tb.Label(bf,text="单元格:").pack(side='left',padx=(8,0)); self.c_var=tk.StringVar(); tb.Entry(bf,textvariable=self.c_var,width=6).pack(side='left',padx=2)
        tb.Button(bf,text="添加",width=5,command=self._add).pack(side='left',padx=4)
        tb.Button(bf,text="删除",width=5,command=self._del).pack(side='left',padx=2)
        # Segment mode frame
        self.segment_frame=tb.Frame(main)
        tb.Label(self.segment_frame,text="示例文件名（按 _ 自动拆分）:",font=('微软雅黑',9,'bold')).pack(anchor='w')
        sample_frame=tb.Frame(self.segment_frame); sample_frame.pack(fill='x',pady=4)
        self.sample_var=tk.StringVar(value=config.get('sample_filename','')); tb.Entry(sample_frame,textvariable=self.sample_var,width=50).pack(side='left',fill='x',expand=True)
        tb.Button(sample_frame,text="解析",width=6,command=self._parse_filename).pack(side='left',padx=(8,0))
        tb.Label(self.segment_frame,text="分段列表（勾选需要检查的段）:",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(8,0))
        seg_frame=tb.Frame(self.segment_frame); seg_frame.pack(fill='both',expand=True,pady=4)
        seg_cols=('idx','segment','sheet','cell'); self.seg_tv=tb.Treeview(seg_frame,columns=seg_cols,show='headings',height=10)
        for c,w,h in [('idx',40),('segment',180),('sheet',150),('cell',80)]:
            self.seg_tv.heading(c,text=c); self.seg_tv.column(c,width=w)
        self.seg_tv.pack(side='left',fill='both',expand=True)
        seg_sb=tb.Scrollbar(seg_frame,orient='vertical',command=self.seg_tv.yview); seg_sb.pack(side='right',fill='y')
        self.seg_tv.configure(yscrollcommand=seg_sb.set)
        # Load existing segments
        for seg in self.segments_data:
            self.seg_tv.insert('','end',values=(seg.get('index',''),seg.get('value',''),seg.get('sheet',''),seg.get('cell','')))
        edit_frame=tb.Frame(self.segment_frame); edit_frame.pack(fill='x',pady=4)
        tb.Label(edit_frame,text="Sheet:").pack(side='left'); self.seg_s_var=tk.StringVar(); self.seg_s_cb=tb.Combobox(edit_frame,textvariable=self.seg_s_var,width=14,values=sheets); self.seg_s_cb.pack(side='left',padx=2)
        tb.Label(edit_frame,text="单元格:").pack(side='left',padx=(8,0)); self.seg_c_var=tk.StringVar(); tb.Entry(edit_frame,textvariable=self.seg_c_var,width=6).pack(side='left',padx=2)
        tb.Button(edit_frame,text="设置映射",width=7,command=self._set_seg_mapping).pack(side='left',padx=8)
        tb.Button(edit_frame,text="删除",width=5,command=self._del_seg).pack(side='left',padx=2)
        tb.Label(self.segment_frame,text="说明: 解析后勾选的段会与实际文件名对应位置的片段比较，不匹配则报警。",foreground='gray',wraplength=500,justify='left').pack(anchor='w',pady=(4,0))
        # Initial mode (pack content BEFORE buttons so buttons stay at bottom)
        self._switch_mode()
        # Buttons
        btn=tb.Frame(main); btn.pack(fill='x',pady=(8,0))
        tb.Button(btn,text="确定",bootstyle=PRIMARY,width=8,command=self._ok).pack(side='right',padx=5)
        tb.Button(btn,text="取消",width=8,command=self.destroy).pack(side='right')
        center_window(self,parent); self.grab_set(); self.wait_window()
    def _switch_mode(self):
        if self.mode_var.get()=='template':
            self.segment_frame.pack_forget()
            self.template_frame.pack(fill='both',expand=True)
        else:
            self.template_frame.pack_forget()
            self.segment_frame.pack(fill='both',expand=True)
    def _add(self):
        f,s,c=self.f_var.get().strip(),self.s_var.get().strip(),self.c_var.get().strip()
        if f and s and c: self.tv.insert('','end',values=(f,s,c)); self.f_var.set(''); self.c_var.set('')
    def _del(self):
        for sel in self.tv.selection(): self.tv.delete(sel)
    def _parse_filename(self):
        sample=self.sample_var.get().strip()
        if not sample: messagebox.showwarning("提示","请输入示例文件名"); return
        # Clear existing
        for iid in self.seg_tv.get_children(): self.seg_tv.delete(iid)
        # Split by underscore
        parts=sample.split('_')
        for idx,part in enumerate(parts):
            self.seg_tv.insert('','end',values=(idx,part,'',''))
    def _set_seg_mapping(self):
        sel=self.seg_tv.selection()
        if not sel: messagebox.showwarning("提示","请先选择要设置的分段"); return
        s=self.seg_s_var.get().strip(); c=self.seg_c_var.get().strip()
        if not s or not c: messagebox.showwarning("提示","请输入Sheet和单元格"); return
        for iid in sel:
            vals=self.seg_tv.item(iid)['values']
            self.seg_tv.item(iid,values=(vals[0],vals[1],s,c))
    def _del_seg(self):
        for sel in self.seg_tv.selection(): self.seg_tv.delete(sel)
    def _ok(self):
        if self.mode_var.get()=='template':
            mappings=[]
            for iid in self.tv.get_children():
                v=self.tv.item(iid)['values']; mappings.append({'field':str(v[0]),'sheet':str(v[1]),'cell':str(v[2])})
            self.result={'format_template':self.fmt_var.get().strip(),'field_mappings':mappings,'segments':[]}
        else:
            segments=[]
            for iid in self.seg_tv.get_children():
                v=self.seg_tv.item(iid)['values']
                idx=int(v[0]) if str(v[0]).isdigit() else -1
                segments.append({'index':idx,'value':str(v[1]),'sheet':str(v[2]),'cell':str(v[3])})
            self.result={'format_template':'','field_mappings':[],'segments':segments,'sample_filename':self.sample_var.get().strip()}
        self.destroy()

class SpecialReminderConfigDialog(tb.Toplevel):
    def __init__(self, parent, config, sheets):
        super().__init__(parent); self.title("特殊提醒配置"); self.geometry("480x340"); self.transient(parent); self.result=None; self.sheets=sheets
        main=tb.Frame(self,padding=15); main.pack(fill='both',expand=True)
        tb.Label(main,text="数据源：自动套用当前规则的数据源配置",foreground='gray',wraplength=440,justify='left').pack(anchor='w',pady=(0,8))
        tb.Label(main,text="提醒描述:").pack(anchor='w'); self.desc_var=tk.StringVar(value=config.get('description','需人工复核')); tb.Entry(main,textvariable=self.desc_var,width=45).pack(fill='x',pady=2)
        tb.Label(main,text="触发条件:").pack(anchor='w',pady=(8,0))
        tf=tb.Frame(main); tf.pack(fill='x',pady=2)
        self.trigger_var=tk.StringVar(value=config.get('trigger','always'))
        for val,txt in [('always','始终触发'),('nonempty','任意非空'),('gt','值 >'),('lt','值 <')]:
            tb.Radiobutton(tf,text=txt,variable=self.trigger_var,value=val).pack(side='left',padx=6)
        tf2=tb.Frame(main); tf2.pack(fill='x',pady=(4,0))
        tb.Label(tf2,text="阈值:").pack(side='left')
        self.thresh_var=tk.StringVar(value=str(config.get('threshold',''))); tb.Entry(tf2,textvariable=self.thresh_var,width=10).pack(side='left',padx=(4,0))
        btn=tb.Frame(main); btn.pack(fill='x',pady=(16,0))
        tb.Button(btn,text="确定",bootstyle=PRIMARY,width=8,command=self._ok).pack(side='right',padx=5)
        tb.Button(btn,text="取消",width=8,command=self.destroy).pack(side='right')
        center_window(self,parent); self.grab_set(); self.wait_window()
    def _ok(self):
        try: th=float(self.thresh_var.get()) if self.thresh_var.get().strip() else None
        except: th=None
        self.result={'description':self.desc_var.get().strip(),'trigger':self.trigger_var.get(),'threshold':th}
        self.destroy()

class DataTrendConfigDialog(tb.Toplevel):
    def __init__(self, parent, config, sheets):
        super().__init__(parent); self.title("数据趋势配置"); self.geometry("520x680"); self.transient(parent); self.result=None; self.sheets=sheets
        main=tb.Frame(self,padding=15); main.pack(fill='both',expand=True)
        canvas=tk.Canvas(main,highlightthickness=0); sb=ttk.Scrollbar(main,orient='vertical',command=canvas.yview)
        sf=tb.Frame(canvas); sf.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=sf,anchor="nw",width=480); canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left",fill="both",expand=True); sb.pack(side="right",fill="y")
        # 提示
        tb.Label(sf,text="待检数据源：自动套用当前规则的数据源配置",foreground='gray',wraplength=440,justify='left').pack(anchor='w',pady=(0,4))
        # 历史数据
        tb.Label(sf,text="▸ 历史数据范围（同文件）",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(8,2))
        tb.Label(sf,text="Sheet:").pack(anchor='w'); self.hs_var=tk.StringVar(value=config.get('history_sheet','')); tb.Combobox(sf,textvariable=self.hs_var,width=20,values=sheets).pack(fill='x',pady=1)
        tb.Label(sf,text="锚点:").pack(anchor='w'); self.ha_var=tk.StringVar(value=config.get('history_anchor',{}).get('text','')); tb.Entry(sf,textvariable=self.ha_var,width=30).pack(fill='x',pady=1)
        hf=tb.Frame(sf); hf.pack(fill='x',pady=1)
        tb.Label(hf,text="行偏移:").pack(side='left'); self.hro=tk.StringVar(value=str(config.get('history_target',{}).get('row_offset',0))); tb.Entry(hf,textvariable=self.hro,width=5).pack(side='left',padx=2)
        tb.Label(hf,text="列偏移:").pack(side='left'); self.hco=tk.StringVar(value=str(config.get('history_target',{}).get('col_offset',0))); tb.Entry(hf,textvariable=self.hco,width=5).pack(side='left',padx=2)
        tb.Label(hf,text="行数:").pack(side='left'); self.hrc=tk.StringVar(value=str(config.get('history_target',{}).get('row_count',1))); tb.Entry(hf,textvariable=self.hrc,width=5).pack(side='left',padx=2)
        tb.Label(hf,text="列数:").pack(side='left'); self.hcc=tk.StringVar(value=str(config.get('history_target',{}).get('col_count',1))); tb.Entry(hf,textvariable=self.hcc,width=5).pack(side='left',padx=2)
        # 管制阈值
        mt=config.get('metric_thresholds',{})
        tb.Label(sf,text="▸ 管制阈值",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(8,2))
        mf=tb.Frame(sf); mf.pack(fill='x',pady=1)
        tb.Label(mf,text="均值范围:").pack(side='left'); self.mn_lo=tk.StringVar(value=str(mt.get('mean_range',['',''])[0])); tb.Entry(mf,textvariable=self.mn_lo,width=6).pack(side='left',padx=1)
        tb.Label(mf,text="~").pack(side='left'); self.mn_hi=tk.StringVar(value=str(mt.get('mean_range',['',''])[1])); tb.Entry(mf,textvariable=self.mn_hi,width=6).pack(side='left',padx=1)
        tb.Label(mf,text="标准差:").pack(side='left',padx=(8,0)); self.sd_lo=tk.StringVar(value=str(mt.get('stddev_range',['',''])[0])); tb.Entry(mf,textvariable=self.sd_lo,width=6).pack(side='left',padx=1)
        tb.Label(mf,text="~").pack(side='left'); self.sd_hi=tk.StringVar(value=str(mt.get('stddev_range',['',''])[1])); tb.Entry(mf,textvariable=self.sd_hi,width=6).pack(side='left',padx=1)
        mf2=tb.Frame(sf); mf2.pack(fill='x',pady=1)
        tb.Label(mf2,text="最大值≤:").pack(side='left'); self.max_l=tk.StringVar(value=str(mt.get('max_limit',''))); tb.Entry(mf2,textvariable=self.max_l,width=8).pack(side='left',padx=2)
        tb.Label(mf2,text="最小值≥:").pack(side='left',padx=(8,0)); self.min_l=tk.StringVar(value=str(mt.get('min_limit',''))); tb.Entry(mf2,textvariable=self.min_l,width=8).pack(side='left',padx=2)
        # 规格与CPK
        tb.Label(sf,text="▸ 规格与CPK",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(8,2))
        sf2=tb.Frame(sf); sf2.pack(fill='x',pady=1)
        tb.Label(sf2,text="USL:").pack(side='left'); self.usl=tk.StringVar(value=str(mt.get('spec_usl',''))); tb.Entry(sf2,textvariable=self.usl,width=8).pack(side='left',padx=2)
        tb.Label(sf2,text="LSL:").pack(side='left',padx=(8,0)); self.lsl=tk.StringVar(value=str(mt.get('spec_lsl',''))); tb.Entry(sf2,textvariable=self.lsl,width=8).pack(side='left',padx=2)
        tb.Label(sf2,text="CPK≥:").pack(side='left',padx=(8,0)); self.cpk=tk.StringVar(value=str(mt.get('cpk_min',''))); tb.Entry(sf2,textvariable=self.cpk,width=6).pack(side='left',padx=2)
        # 显著性检验
        tb.Label(sf,text="▸ 显著性检验",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(8,2))
        tf=tb.Frame(sf); tf.pack(fill='x',pady=1)
        self.ttest_var=tk.BooleanVar(value=config.get('t_test',False)); tb.Checkbutton(tf,text="t 检验(均值)",variable=self.ttest_var,bootstyle="round-toggle").pack(side='left')
        tb.Label(tf,text="α:").pack(side='left',padx=(12,0)); self.alpha_var=tk.StringVar(value=str(config.get('alpha',0.05))); tb.Entry(tf,textvariable=self.alpha_var,width=6).pack(side='left',padx=2)
        tf2=tb.Frame(sf); tf2.pack(fill='x',pady=1)
        self.ftest_var=tk.BooleanVar(value=config.get('f_test',False)); tb.Checkbutton(tf2,text="F 检验(方差)",variable=self.ftest_var,bootstyle="round-toggle").pack(side='left')
        tb.Label(tf2,text="α:").pack(side='left',padx=(12,0)); self.falpha_var=tk.StringVar(value=str(config.get('f_alpha',0.05))); tb.Entry(tf2,textvariable=self.falpha_var,width=6).pack(side='left',padx=2)
        # UCL/LCL
        ucl_cfg=config.get('ucl_lcl',{})
        tb.Label(sf,text="▸ UCL/LCL 管制限",font=('微软雅黑',9,'bold')).pack(anchor='w',pady=(8,2))
        uf=tb.Frame(sf); uf.pack(fill='x',pady=1)
        self.ucl_enable_var=tk.BooleanVar(value=ucl_cfg.get('enabled',False)); tb.Checkbutton(uf,text="启用",variable=self.ucl_enable_var,bootstyle="round-toggle").pack(side='left')
        tb.Label(uf,text="σ倍数:").pack(side='left',padx=(8,0)); self.sigma_var=tk.StringVar(value=ucl_cfg.get('sigma_level','3σ'))
        tb.Combobox(uf,textvariable=self.sigma_var,values=['1σ','2σ','3σ','4σ','5σ','6σ','特殊'],width=6).pack(side='left',padx=2)
        self.sigma_var.trace_add('write',self._toggle_ucl_special)
        # UCL
        uf2=tb.Frame(sf); uf2.pack(fill='x',pady=1)
        tb.Label(uf2,text="UCL:").pack(side='left')
        self.ucl_op_var=tk.StringVar(value=ucl_cfg.get('ucl_op','≥')); tb.Combobox(uf2,textvariable=self.ucl_op_var,values=['≥','>','=','<','≤'],width=3).pack(side='left',padx=2)
        self.ucl_val_var=tk.StringVar(value=str(ucl_cfg.get('ucl_val',''))); self.ucl_val_entry=tb.Entry(uf2,textvariable=self.ucl_val_var,width=8); self.ucl_val_entry.pack(side='left',padx=2)
        # LCL
        uf3=tb.Frame(sf); uf3.pack(fill='x',pady=1)
        tb.Label(uf3,text="LCL:").pack(side='left')
        self.lcl_op_var=tk.StringVar(value=ucl_cfg.get('lcl_op','≤')); tb.Combobox(uf3,textvariable=self.lcl_op_var,values=['≥','>','=','<','≤'],width=3).pack(side='left',padx=2)
        self.lcl_val_var=tk.StringVar(value=str(ucl_cfg.get('lcl_val',''))); self.lcl_val_entry=tb.Entry(uf3,textvariable=self.lcl_val_var,width=8); self.lcl_val_entry.pack(side='left',padx=2)
        # 初始化特殊模式状态
        self._toggle_ucl_special(None,None,None)
        btn=tb.Frame(main); btn.pack(fill='x',pady=(8,0))
        tb.Button(btn,text="确定",bootstyle=PRIMARY,width=8,command=self._ok).pack(side='right',padx=5)
        tb.Button(btn,text="取消",width=8,command=self.destroy).pack(side='right')
        center_window(self,parent); self.grab_set(); self.wait_window()
    def _toggle_ucl_special(self,*args):
        is_special=self.sigma_var.get()=='特殊'
        state='normal' if is_special else 'disabled'
        try: self.ucl_val_entry.configure(state=state)
        except: pass
        try: self.lcl_val_entry.configure(state=state)
        except: pass
    def _to_f(self, v):
        try: return float(v) if v.strip() else ''
        except: return ''
    def _ok(self):
        mt={'mean_range':[self._to_f(self.mn_lo.get()),self._to_f(self.mn_hi.get())],
            'stddev_range':[self._to_f(self.sd_lo.get()),self._to_f(self.sd_hi.get())],
            'max_limit':self._to_f(self.max_l.get()),'min_limit':self._to_f(self.min_l.get()),
            'spec_usl':self._to_f(self.usl.get()),'spec_lsl':self._to_f(self.lsl.get()),'cpk_min':self._to_f(self.cpk.get())}
        ucl_lcl={'enabled':self.ucl_enable_var.get(),'sigma_level':self.sigma_var.get(),
            'ucl_op':self.ucl_op_var.get(),'ucl_val':self._to_f(self.ucl_val_var.get()),
            'lcl_op':self.lcl_op_var.get(),'lcl_val':self._to_f(self.lcl_val_var.get())}
        self.result={'history_sheet':self.hs_var.get().strip(),'history_anchor':{'text':self.ha_var.get().strip()},
            'history_target':{'row_offset':int(self.hro.get() or 0),'col_offset':int(self.hco.get() or 0),
                'row_count':int(self.hrc.get() or 1),'col_count':int(self.hcc.get() or 1)},
            'metric_thresholds':mt,'t_test':self.ttest_var.get(),'alpha':float(self.alpha_var.get() or 0.05),
            'f_test':self.ftest_var.get(),'f_alpha':float(self.falpha_var.get() or 0.05),
            'ucl_lcl':ucl_lcl}
        self.destroy()

class RuleEditorDialog(tb.Toplevel):
    def __init__(self,parent,old_path,new_path,rule=None):
        super().__init__(parent); self.title("编辑规则"); self.geometry("1100x800"); self.transient(parent); self.parent=parent; self.old_path=old_path; self.new_path=new_path; self.result=None; self.rule=rule or CheckRule(); self._build_ui(); self._load_rule_data(); center_window(self,parent)
        self.lift(); self.focus_force()
        try: self.grab_set()
        except Exception: pass
        self.wait_window()
    def _get_sheet_names(self):
        for p in [self.old_path,self.new_path]:
            if os.path.isfile(p):
                sheets=get_sheet_names_fast(p)
                if sheets: return sheets
        return []
    def _build_ui(self):
        main=tb.Frame(self); main.pack(fill='both',expand=True,padx=10,pady=5)
        left=tb.Frame(main,width=440); left.pack(side='left',fill='y',padx=(0,10)); left.pack_propagate(False)
        # 按钮先固定在底部，保证任何模式下都可见
        btn=tb.Frame(left); btn.pack(side='bottom',fill='x',pady=(5,2)); tb.Button(btn,text="确定",bootstyle=PRIMARY,width=10,command=self.on_ok).pack(side='left',padx=5); tb.Button(btn,text="取消",bootstyle="outline",width=10,command=self.on_cancel).pack(side='right',padx=5)
        # 数据源配置放入可滚动区域，参数再多也不会挤压按钮
        ds_outer=tb.Frame(left); ds_outer.pack(side='top',fill='both',expand=True,pady=5)
        ds_canvas=tk.Canvas(ds_outer,highlightthickness=0,width=400,bg='#ffffff')
        ds_sb=ttk.Scrollbar(ds_outer,orient="vertical",command=ds_canvas.yview)
        ds=tb.Labelframe(ds_canvas,text="数据源配置",padding=10)
        ds.bind("<Configure>",lambda e:ds_canvas.configure(scrollregion=ds_canvas.bbox("all")))
        ds_canvas.create_window((0,0),window=ds,anchor="nw",width=396)
        ds_canvas.configure(yscrollcommand=ds_sb.set)
        ds_canvas.pack(side="left",fill="both",expand=True); ds_sb.pack(side="right",fill="y")
        def _ds_wheel(event):
            ds_canvas.yview_scroll(int(-1*(event.delta/120)),"units")
        ds_canvas.bind("<Enter>",lambda e:ds_canvas.bind_all("<MouseWheel>",_ds_wheel))
        ds_canvas.bind("<Leave>",lambda e:ds_canvas.unbind_all("<MouseWheel>"))
        tb.Label(ds,text="规则名称:").pack(anchor='w'); self.rule_name_var=tk.StringVar(); tb.Entry(ds,textvariable=self.rule_name_var,width=40).pack(fill='x',pady=2)
        tb.Label(ds,text="Sheet:").pack(anchor='w'); sr=tb.Frame(ds); sr.pack(fill='x',pady=2); self.sheet_var=tk.StringVar(); self.sheet_cb=tb.Combobox(sr,textvariable=self.sheet_var,width=30); self.sheet_cb.pack(side='left',fill='x',expand=True); tb.Button(sr,text="抓取",bootstyle="outline",width=5,command=self.fetch_sheets).pack(side='right',padx=(3,0))
        # 打开编辑器时自动毫秒级加载 sheet 名（纯 zipfile 解析，不加载整个工作簿，不卡 UI）
        try:
            auto_sheets=self._get_sheet_names()
            if auto_sheets: self.sheet_cb['values']=auto_sheets
        except Exception: pass
        tb.Label(ds,text="锚点文字:").pack(anchor='w'); self.anchor_text_var=tk.StringVar(); self.anchor_entry=tb.Entry(ds,textvariable=self.anchor_text_var,width=40); self.anchor_entry.pack(fill='x',pady=2)
        tb.Label(ds,text="搜索范围(例A1:B2，留空=全表):").pack(anchor='w'); self.search_in_var=tk.StringVar(value=''); self.search_in_cb=tb.Entry(ds,textvariable=self.search_in_var,width=38); self.search_in_cb.pack(fill='x',pady=2)
        tb.Label(ds,text="模式:").pack(anchor='w'); self.mode_var=tk.StringVar(value='offset'); self.mode_cb=tb.Combobox(ds,textvariable=self.mode_var,values=['offset','intersection','range','shift'],width=38); self.mode_cb.pack(fill='x',pady=2); self.mode_cb.bind('<<ComboboxSelected>>',lambda e:self._build_param_fields())
        self.param_frame=tb.Frame(ds); self.param_frame.pack(fill='x',pady=2); self._build_param_fields()
        right=tb.Frame(main); right.pack(side='right',fill='both',expand=True)
        cf=tb.Labelframe(right,text="检查项（可多选）",padding=10); cf.pack(fill='both',expand=True)
        canvas=tk.Canvas(cf,highlightthickness=0,bg='#ffffff'); sb=ttk.Scrollbar(cf,orient="vertical",command=canvas.yview); sf=tb.Frame(canvas)
        _cw=canvas.create_window((0,0),window=sf,anchor="nw"); canvas.bind("<Configure>",lambda e:canvas.itemconfigure(_cw,width=e.width)); sf.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all"))); canvas.configure(yscrollcommand=sb.set); canvas.pack(side="left",fill="both",expand=True); sb.pack(side="right",fill="y")
        self.check_vars={}; self.expect_vars={}
        for gname,keys in CHECK_OPTION_GROUPS:
            lf=tb.Labelframe(sf,text=f"{gname}（共{len(keys)}项）",padding=(8,5)); lf.pack(fill='x',pady=(0,8),padx=5)
            for key in keys:
                row=tb.Frame(lf); row.pack(fill='x',pady=1); lc=tb.Frame(row); lc.pack(side='left',fill='x',expand=True); var=tk.BooleanVar(value=True); self.check_vars[key]=var; tb.Checkbutton(lc,text=CHECK_OPTION_LABELS[key],variable=var,bootstyle="round-toggle").pack(side='left',anchor='w'); rc=tb.Frame(row); rc.pack(side='right'); tb.Label(rc,text="期望:").pack(side='left',padx=(10,2)); ev=tk.StringVar(value='same'); tb.Combobox(rc,textvariable=ev,values=['same','different'],width=8).pack(side='left'); self.expect_vars[key]=ev
        adv_lf=tb.Labelframe(sf,text="高级检查（仅待检文件）",padding=(8,5)); adv_lf.pack(fill='x',pady=(0,8),padx=5)
        self.adv_engine_vars={}; self.adv_engine_configs={}
        for etype,elabel in [('filename_check','文件名一致性'),('special_reminder','特殊提醒'),('data_trend','数据趋势')]:
            arow=tb.Frame(adv_lf); arow.pack(fill='x',pady=1)
            ev=tk.BooleanVar(value=False); self.adv_engine_vars[etype]=ev
            self.adv_engine_configs[etype]={}
            tb.Checkbutton(arow,text=elabel,variable=ev,bootstyle="round-toggle").pack(side='left',anchor='w')
            tb.Button(arow,text="配置",width=5,command=lambda t=etype,l=elabel:self._open_engine_config(t,l)).pack(side='right',padx=(5,0))
    def _open_engine_config(self, engine_type, label):
        cfg=self.adv_engine_configs.get(engine_type,{})
        sheets=[]
        for p in [self.old_path,self.new_path]:
            if os.path.isfile(p):
                try: sheets=get_sheet_names_fast(p); break
                except: pass
        # 特殊提醒和数据趋势不支持shift模式
        if engine_type in ('special_reminder','data_trend') and self.mode_var.get()=='shift':
            messagebox.showwarning("不支持搬移模式","特殊提醒和数据趋势不允许使用 shift（搬移）模式，请改为 offset / range / intersection")
            return
        if engine_type=='filename_check':
            dlg=FileNameCheckConfigDialog(self,cfg,sheets); 
        elif engine_type=='special_reminder':
            dlg=SpecialReminderConfigDialog(self,cfg,sheets)
        else:
            dlg=DataTrendConfigDialog(self,cfg,sheets)
        if dlg.result is not None: self.adv_engine_configs[engine_type]=dlg.result
    def fetch_sheets(self):
        if not os.path.isfile(self.old_path) and not os.path.isfile(self.new_path): messagebox.showwarning("提示","请先选择有效的Excel文件路径"); return
        self.config(cursor='watch')
        try:
            sheets=self._get_sheet_names()
        finally:
            self.config(cursor='')
        if sheets:
            self.sheet_cb['values']=sheets
            if not self.sheet_var.get() and sheets: self.sheet_var.set(sheets[0])
        else: messagebox.showwarning("提示","无法读取Sheet列表，请确认文件格式或文件未被占用")
    def _build_param_fields(self):
        for w in self.param_frame.winfo_children(): w.destroy()
        mode=self.mode_var.get()
        # 切换到 shift 模式时，拦截已启用的特殊提醒/数据趋势
        if mode=='shift':
            for etype in ('special_reminder','data_trend'):
                if self.adv_engine_vars.get(etype) and self.adv_engine_vars[etype].get():
                    etype_label = '特殊提醒' if etype=='special_reminder' else '数据趋势'
                    messagebox.showwarning("不支持搬移模式", f"已启用【{etype_label}】，不允许切换到 shift（搬移）模式，请改为 offset / range / intersection")
                    self.mode_var.set('offset'); mode='offset'; break
        if mode=='offset':
            tb.Label(self.param_frame,text="行偏移(负数向上):").pack(anchor='w'); self.offset_row_var=tk.StringVar(value='0'); tb.Entry(self.param_frame,textvariable=self.offset_row_var,width=10).pack(fill='x',pady=2); tb.Label(self.param_frame,text="列偏移(负数向左):").pack(anchor='w'); self.offset_col_var=tk.StringVar(value='0'); tb.Entry(self.param_frame,textvariable=self.offset_col_var,width=10).pack(fill='x',pady=2)
        elif mode=='intersection':
            tb.Label(self.param_frame,text="行锚点文字:").pack(anchor='w'); self.row_anchor_text_var=tk.StringVar(); tb.Entry(self.param_frame,textvariable=self.row_anchor_text_var,width=20).pack(fill='x',pady=2); tb.Label(self.param_frame,text="列锚点文字:").pack(anchor='w'); self.col_anchor_text_var=tk.StringVar(); tb.Entry(self.param_frame,textvariable=self.col_anchor_text_var,width=20).pack(fill='x',pady=2)
        elif mode=='range':
            tb.Label(self.param_frame,text="行偏移(负数向上):").pack(anchor='w'); self.range_row_offset_var=tk.StringVar(value='0'); tb.Entry(self.param_frame,textvariable=self.range_row_offset_var,width=10).pack(fill='x',pady=2); tb.Label(self.param_frame,text="列偏移(负数向左):").pack(anchor='w'); self.range_col_offset_var=tk.StringVar(value='0'); tb.Entry(self.param_frame,textvariable=self.range_col_offset_var,width=10).pack(fill='x',pady=2); tb.Label(self.param_frame,text="行数:").pack(anchor='w'); self.range_row_count_var=tk.StringVar(value='1'); tb.Entry(self.param_frame,textvariable=self.range_row_count_var,width=10).pack(fill='x',pady=2); tb.Label(self.param_frame,text="列数:").pack(anchor='w'); self.range_col_count_var=tk.StringVar(value='1'); tb.Entry(self.param_frame,textvariable=self.range_col_count_var,width=10).pack(fill='x',pady=2); tb.Label(self.param_frame,text="排除单元格(逗号分隔，如A5,C7或[1,0]):").pack(anchor='w'); self.range_exclude_var=tk.StringVar(value=''); tb.Entry(self.param_frame,textvariable=self.range_exclude_var,width=30).pack(fill='x',pady=2)
        elif mode=='shift':
            tb.Label(self.param_frame,text="标题行范围（相对一级锚点）",font=('',9,'bold')).pack(anchor='w',pady=(2,0))
            self.sh_ro_var=tk.StringVar(value='0'); self.sh_co_var=tk.StringVar(value='0'); self.sh_rc_var=tk.StringVar(value='1'); self.sh_cc_var=tk.StringVar(value='20')
            for _lbl,_var in (("行偏移(负数向上):",self.sh_ro_var),("列偏移(负数向左):",self.sh_co_var),("行数:",self.sh_rc_var),("列数:",self.sh_cc_var)):
                _r=tb.Frame(self.param_frame); _r.pack(fill='x',pady=1); tb.Label(_r,text=_lbl).pack(side='left'); tb.Entry(_r,textvariable=_var,width=8).pack(side='left',padx=4)
            tb.Label(self.param_frame,text="垂直范围(行号，如3,5-7,13-14):").pack(anchor='w',pady=(4,1)); self.sh_rows_var=tk.StringVar(value=''); tb.Entry(self.param_frame,textvariable=self.sh_rows_var,width=24).pack(fill='x',pady=1)
            tb.Label(self.param_frame,text="列偏移值(旧列号+偏移=新列号):").pack(anchor='w',pady=(4,1)); self.sh_offset_var=tk.StringVar(value='0'); tb.Entry(self.param_frame,textvariable=self.sh_offset_var,width=8).pack(anchor='w',pady=1)
            tb.Label(self.param_frame,text="两份报告共用顶部Sheet/锚点/搜索范围。旧报告标题区有数据的列按固定偏移映射到新报告列（如偏移+2则J→L、K→M），垂直范围内对应单元格逐格比对，勾选检查项一致即豁免。",foreground='gray',wraplength=360,justify='left').pack(fill='x',pady=2)
    def _load_rule_data(self):
        self.rule_name_var.set(self.rule.rule_name); ds=self.rule.data_source; self.sheet_var.set(ds.get('sheet','')); self.anchor_text_var.set(ds.get('anchor',{}).get('text','')); self.search_in_var.set('' if ds.get('search_in','all') in ('all','') else ds.get('search_in','')); self.mode_var.set(ds.get('mode','offset')); self._build_param_fields()
        target=ds.get('target',{}); mode=self.mode_var.get()
        if mode=='offset': self.offset_row_var.set(str(target.get('row_offset',0))); self.offset_col_var.set(str(target.get('col_offset',0)))
        elif mode=='range':
            self.range_row_offset_var.set(str(target.get('row_offset',0))); self.range_col_offset_var.set(str(target.get('col_offset',0))); self.range_row_count_var.set(str(target.get('row_count',1))); self.range_col_count_var.set(str(target.get('col_count',1)))
            ex=target.get('exclude',[]); self.range_exclude_var.set(','.join(str(x) if isinstance(x,str) else f"[{x[0]},{x[1]}]" for x in ex))
        elif mode=='shift':
            ht=ds.get('header_target',{}); self.sh_ro_var.set(str(ht.get('row_offset',0))); self.sh_co_var.set(str(ht.get('col_offset',0))); self.sh_rc_var.set(str(ht.get('row_count',1))); self.sh_cc_var.set(str(ht.get('col_count',20)))
            self.sh_rows_var.set(ds.get('rows','')); self.sh_offset_var.set(str(ds.get('shift_offset',0)))
        elif mode=='intersection':
            ra=ds.get('row_anchor',{}); ca=ds.get('col_anchor',{}); self.row_anchor_text_var.set(ra.get('text','').strip()); self.col_anchor_text_var.set(ca.get('text','').strip())
        # 刷新检查项开关和期望值：先全部重置为关闭，再按规则逐项打开
        for key in self.check_vars:
            self.check_vars[key].set(False)
            self.expect_vars[key].set('same')
        for check in self.rule.checks:
            if check.check_type in self.check_vars:
                self.check_vars[check.check_type].set(bool(check.enabled))
                self.expect_vars[check.check_type].set(check.expect or 'same')
        for etype in self.adv_engine_vars:
            self.adv_engine_vars[etype].set(False); self.adv_engine_configs[etype]={}
        for ae in getattr(self.rule,'advanced_engines',[]):
            if ae.engine_type in self.adv_engine_vars:
                self.adv_engine_vars[ae.engine_type].set(ae.enabled)
                self.adv_engine_configs[ae.engine_type]=dict(ae.config)
    def _to_int(self, var, default=0):
        try: return int(str(var.get()).strip())
        except Exception: return default
    def on_ok(self):
        si=self.search_in_var.get().strip().upper().replace('$','')
        try: ds_si=si if DataLocator._parse_area(si) else 'all'
        except Exception: ds_si='all'
        self.rule.rule_name=self.rule_name_var.get(); ds={'name':self.rule_name_var.get(),'sheet':self.sheet_var.get(),'anchor':{'text':self.anchor_text_var.get()},'search_in':ds_si,'mode':self.mode_var.get()}
        if self.mode_var.get()=='offset': ds['target']={'row_offset':self._to_int(self.offset_row_var),'col_offset':self._to_int(self.offset_col_var)}
        elif self.mode_var.get()=='intersection': ds['row_anchor']={'text':self.row_anchor_text_var.get().strip(),'search_in':ds_si}; ds['col_anchor']={'text':self.col_anchor_text_var.get().strip(),'search_in':ds_si}
        elif self.mode_var.get()=='range': ds['target']={'row_offset':self._to_int(self.range_row_offset_var),'col_offset':self._to_int(self.range_col_offset_var),'row_count':self._to_int(self.range_row_count_var,1) or 1,'col_count':self._to_int(self.range_col_count_var,1) or 1,'exclude':self._parse_exclude(self.range_exclude_var.get())}
        elif self.mode_var.get()=='shift':
            ds['header_target']={'row_offset':self._to_int(self.sh_ro_var),'col_offset':self._to_int(self.sh_co_var),'row_count':max(1,self._to_int(self.sh_rc_var,1)),'col_count':max(1,self._to_int(self.sh_cc_var,20))}; ds['rows']=self.sh_rows_var.get().strip(); ds['shift_offset']=self._to_int(self.sh_offset_var)
        self.rule.data_source=ds; self.rule.checks=[]
        for key,var in self.check_vars.items():
            if var.get(): self.rule.checks.append(CheckItemConfig(check_type=key,enabled=True,expect=self.expect_vars[key].get()))
        self.rule.advanced_engines=[]
        for etype,var in self.adv_engine_vars.items():
            if var.get(): self.rule.advanced_engines.append(AdvancedEngineConfig(engine_type=etype,enabled=True,config=self.adv_engine_configs.get(etype,{})))
        self.result=self.rule; self.destroy()
    def on_cancel(self): self.result=None; self.destroy()
    def _parse_exclude(self,text):
        if not text.strip(): return []
        parts=[p.strip() for p in text.split(',') if p.strip()]; exclude=[]
        for p in parts:
            if re.match(r'^[A-Za-z]+\d+$',p): exclude.append(p)
            elif re.match(r'^\[\d+,\d+\]$',p): inner=p[1:-1].split(','); exclude.append([int(inner[0]),int(inner[1])])
        return exclude

class DiffViewer:
    def __init__(self,root):
        self.root=root; self.root.title(f"MBO PBO报告检查工具 {VERSION}"); self.root.geometry("1100x750")
        pythoncom.CoInitialize()  # 主线程初始化 COM，确保 GetActiveObject 等可用
        self.old_path=tk.StringVar(); self.new_path=tk.StringVar(); self.topmost=tk.BooleanVar(value=False)
        self.check_options=dict(DEFAULT_CHECK_OPTIONS); self.plugin_manager=None; self.config_file=None; self.stop_event=threading.Event(); self.check_project=None; self.old_sheet_order=[]
        self.color_tolerance=tk.IntVar(value=0)
        self.com_verify=tk.BooleanVar(value=True)
        toolbar=tb.Frame(root,padding=5); toolbar.pack(fill='x')
        top=tb.Frame(toolbar); top.grid(row=0,column=0,rowspan=2,sticky='ns',padx=(2,2),pady=1)
        top_in=tb.Frame(top); top_in.pack(anchor='center')
        tb.Label(top_in,text="置顶").pack(anchor='center'); tb.Checkbutton(top_in,variable=self.topmost,command=self.toggle_topmost,bootstyle="round-toggle").pack(anchor='center')
        self.start_btn=tb.Button(toolbar,text="常规差异\n逐行对比",bootstyle=INFO,width=10,command=self.start_compare); self.start_btn.grid(row=0,column=1,rowspan=2,sticky='nsew',padx=2,pady=1)
        self.config_btn=tb.Button(toolbar,text="高级\n审核",bootstyle="outline-primary",width=7,command=self.load_check_project); self.config_btn.grid(row=0,column=2,rowspan=2,sticky='nsew',padx=2,pady=1)
        self.project_btn=tb.Button(toolbar,text="高级审核\n规则配置",bootstyle="outline-primary",width=9,command=self.open_project_dialog); self.project_btn.grid(row=0,column=3,rowspan=2,sticky='nsew',padx=2,pady=1)
        self.settings_btn=tb.Button(toolbar,text="常规差异\n检测设置",bootstyle="outline",width=9,command=self.open_check_options); self.settings_btn.grid(row=0,column=4,rowspan=2,sticky='nsew',padx=2,pady=1)
        tb.Separator(toolbar,orient='vertical').grid(row=0,column=5,rowspan=2,sticky='ns',padx=8)
        path_frame=tb.Frame(toolbar); path_frame.grid(row=0,column=6,rowspan=2,sticky='nsew',padx=(0,10)); path_frame.columnconfigure(1,weight=1)
        # 状态徽章 + Entry + 浏览按钮（Canvas 圆角徽章）
        self.old_status=self._make_badge(path_frame); self.old_status.grid(row=0,column=0,padx=(0,4),pady=2); self.old_status.bind('<Button-1>',self._on_status_click)
        self.old_entry=tb.Entry(path_frame,foreground='#adb5bd'); self.old_entry.grid(row=0,column=1,sticky='ew',pady=2); tb.Button(path_frame,text="浏览",bootstyle="outline-primary",width=6,command=lambda:self.browse(self.old_path)).grid(row=0,column=2,padx=(5,0),pady=2)
        self.new_status=self._make_badge(path_frame); self.new_status.grid(row=1,column=0,padx=(0,4),pady=2); self.new_status.bind('<Button-1>',self._on_status_click)
        self.new_entry=tb.Entry(path_frame,foreground='#adb5bd'); self.new_entry.grid(row=1,column=1,sticky='ew',pady=2); tb.Button(path_frame,text="浏览",bootstyle="outline-primary",width=6,command=lambda:self.browse(self.new_path)).grid(row=1,column=2,padx=(5,0),pady=2)
        self._set_badge(self.old_status,False); self._set_badge(self.new_status,False)
        self._setup_path_placeholder()
        toolbar.columnconfigure(6,weight=1)
        self.progress=tb.Progressbar(root,mode='determinate',bootstyle="success"); self.progress.pack(fill='x',padx=5,pady=(0,5))
        tree_frame=tb.Frame(root,padding=(5,0)); tree_frame.pack(fill='both',expand=True); tree_frame.columnconfigure(0,weight=1); tree_frame.rowconfigure(0,weight=1)
        self.tree=tb.Treeview(tree_frame,columns=('action','address','type'),show='tree headings',bootstyle=PRIMARY)
        self.tree.heading('#0',text='Sheet / 差异项'); self.tree.heading('action',text='收起',command=self._toggle_all_nodes); self.tree.heading('address',text='位置'); self.tree.heading('type',text='类型')
        self.tree.column('#0',width=340,minwidth=200); self.tree.column('action',width=60,minwidth=60,anchor='center',stretch=False); self.tree.column('address',width=130,minwidth=0,anchor='center',stretch=False); self.tree.column('type',width=130,minwidth=0,anchor='center',stretch=False)
        self.tree.tag_configure('sheet',foreground='blue'); self.tree.tag_configure('warning_sheet',foreground='red'); self.tree.tag_configure('twisty',foreground='#0d6efd'); self.tree.tag_configure('advanced_check',foreground='#F39C12',font=('微软雅黑',9,'italic'))
        try:
            self.root.option_add('*TScrollbar.width',22)
            _st=ttk.Style(); _st.configure('Vertical.TScrollbar',width=22)
        except Exception: pass
        sb=tb.Scrollbar(tree_frame,orient='vertical',command=self.tree.yview,bootstyle=ROUND)
        # 加宽滚动条：直接对 ttkbootstrap 实际生效的样式名设置宽度，避免被主题默认值覆盖
        try:
            _ws=ttk.Style()
            real_style=sb.cget('style') or 'Vertical.TScrollbar'
            for _sn in {real_style,'Vertical.TScrollbar'}:
                try: _ws.configure(_sn,width=30,arrowsize=16)
                except Exception: pass
        except Exception: pass
        self.tree.configure(yscrollcommand=sb.set); self.tree.grid(row=0,column=0,sticky='nsew'); sb.grid(row=0,column=1,sticky='ns')
        self.tree.bind('<<TreeviewSelect>>',self.on_tree_select); self.tree.bind('<Double-1>',self.on_tree_double_click); self.tree.bind('<Button-1>',self.on_tree_click)
        bottom=tb.Frame(root,padding=5); bottom.pack(fill='x'); bottom.columnconfigure(0,weight=0); bottom.columnconfigure(1,weight=1)
        df=tb.Labelframe(bottom,text="差异详情",padding=5,bootstyle=INFO); df.grid(row=0,column=0,sticky='nsew',padx=(0,3)); df.columnconfigure(0,weight=1); df.rowconfigure(0,weight=1); self.detail=tk.Text(df,width=42,height=7,wrap='word',font=("微软雅黑",9),bg='#ffffff',fg='#212529',relief='flat',highlightthickness=1,highlightbackground='#dee2e6',highlightcolor='#0d6efd',padx=5,pady=5); self.detail.grid(row=0,column=0,sticky='nsew'); detail_sb=tb.Scrollbar(df,orient='vertical',command=self.detail.yview,bootstyle="round"); self.detail.configure(yscrollcommand=detail_sb.set); detail_sb.grid(row=0,column=1,sticky='ns')
        lf=tb.Labelframe(bottom,text="日志",padding=5,bootstyle=SECONDARY); lf.grid(row=0,column=1,sticky='nsew',padx=(3,0)); lf.columnconfigure(0,weight=1); lf.rowconfigure(0,weight=1); self.log_text=tk.Text(lf,height=7,wrap='word',font=("微软雅黑",9),bg='#ffffff',fg='#212529',relief='flat',highlightthickness=1,highlightbackground='#dee2e6',highlightcolor='#0d6efd',padx=5,pady=5); self.log_text.grid(row=0,column=0,sticky='nsew'); log_sb=tb.Scrollbar(lf,orient='vertical',command=self.log_text.yview,bootstyle="round"); self.log_text.configure(yscrollcommand=log_sb.set); log_sb.grid(row=0,column=1,sticky='ns')
        # 日志着色 tag
        self.log_text.tag_configure('ts',foreground='#9aa0a6',font=("微软雅黑",8))
        self.log_text.tag_configure('log_ok',foreground='#198754')
        self.log_text.tag_configure('log_bad',foreground='#dc3545')
        self.log_text.tag_configure('log_bold',font=("微软雅黑",9,'bold'))
        self.log_text.tag_configure('log_red_bold',foreground='#dc3545',font=("微软雅黑",9,'bold'))
        self.log_text.tag_configure('log_blue_bold',foreground='#0d6efd',font=("微软雅黑",9,'bold'))
        self.log_text.tag_configure('log_hb',foreground='#6c757d')

        # Sheet列压缩控制：type/address 压完后才允许压缩 #0
        def _on_root_resize(event):
            try:
                aw = self.tree.column('address', 'width')
                tw = self.tree.column('type', 'width')
                if aw <= 1 and tw <= 1:
                    if self.tree.column('#0', 'minwidth') != 80:
                        self.tree.column('#0', minwidth=80)
                else:
                    if self.tree.column('#0', 'minwidth') != 200:
                        self.tree.column('#0', minwidth=200)
            except Exception:
                pass
        self.root.bind('<Configure>', _on_root_resize, add='+')
        self.diff_items=[]; self.result_data=None; self._modal_busy=False
        self.old_entry.bind('<Enter>',lambda e:self._show_path_tip(e,self.old_path.get()))
        self.old_entry.bind('<Leave>',lambda e:self._hide_path_tip())
        self.new_entry.bind('<Enter>',lambda e:self._show_path_tip(e,self.new_path.get()))
        self.new_entry.bind('<Leave>',lambda e:self._hide_path_tip())
        self._tip_win=None
    def _show_path_tip(self,event,path):
        if not path: return
        self._hide_path_tip()
        try:
            name=os.path.basename(path)
            tw=tk.Toplevel(self.root); tw.wm_overrideredirect(True)
            tw.wm_geometry(f'+{event.x_root+12}+{event.y_root+20}')
            tk.Label(tw,text=name,background='#ffffe0',relief='solid',borderwidth=1,font=('微软雅黑',9)).pack()
            self._tip_win=tw
        except Exception: pass
    def _hide_path_tip(self):
        if self._tip_win is not None:
            try: self._tip_win.destroy()
            except Exception: pass
            self._tip_win=None
    def _make_badge(self,parent):
        """创建圆角徽章 Canvas（26x26，背景跟随 ttk Frame 主题色）"""
        try:
            bg = ttk.Style().lookup('TFrame','background') or '#ffffff'
        except Exception:
            bg = '#ffffff'
        c=tk.Canvas(parent,width=26,height=26,bg=bg,highlightthickness=0,cursor='hand2')
        return c
    def _draw_badge(self,canvas,text,color):
        """在 Canvas 上绘制圆角矩形+符号"""
        canvas.delete('all')
        r=7; x1,y1,x2,y2=2,2,24,24
        # 圆角矩形：4 个角用 arc，中间用矩形填充
        canvas.create_arc(x1,y1,x1+2*r,y1+2*r,start=90,extent=90,fill=color,outline=color,style='pieslice')
        canvas.create_arc(x2-2*r,y1,x2,y1+2*r,start=0,extent=90,fill=color,outline=color,style='pieslice')
        canvas.create_arc(x2-2*r,y2-2*r,x2,y2,start=270,extent=90,fill=color,outline=color,style='pieslice')
        canvas.create_arc(x1,y2-2*r,x1+2*r,y2,start=180,extent=90,fill=color,outline=color,style='pieslice')
        canvas.create_rectangle(x1+r,y1,x2-r,y2,fill=color,outline=color)
        canvas.create_rectangle(x1,y1+r,x2,y2-r,fill=color,outline=color)
        canvas.create_text(13,13,text=text,fill='white',font=('Segoe UI Symbol',11))
    def _set_badge(self,status_widget,ok):
        """切换徽章状态：ok=True 绿色✔，False 橙色!（flatly 色系）"""
        if ok: self._draw_badge(status_widget,'\u2714','#18BC9C')
        else:  self._draw_badge(status_widget,'!','#F39C12')
    def _setup_path_placeholder(self):
        """Entry placeholder + 状态徽章联动（不绑定 textvariable，手动同步避免占位文字回写 var 误触发徽章）"""
        def _bind(entry, var, badge, ph):
            def refresh_badge():
                self._set_badge(badge, bool(var.get()))
            def focus_in(ev):
                if not var.get():
                    entry.delete(0,'end'); entry.configure(foreground='#212529')
            def focus_out(ev):
                val=entry.get().strip()
                if val:
                    var.set(val); entry.configure(foreground='#212529')
                elif var.get():
                    entry.delete(0,'end'); entry.insert(0,var.get()); entry.configure(foreground='#212529')
                else:
                    entry.delete(0,'end'); entry.insert(0,ph); entry.configure(foreground='#adb5bd')
                refresh_badge()
            def var_trace(*_):
                if var.get():
                    entry.delete(0,'end'); entry.insert(0,var.get()); entry.configure(foreground='#212529')
                elif self.root.focus_get() is not entry:
                    entry.delete(0,'end'); entry.insert(0,ph); entry.configure(foreground='#adb5bd')
                refresh_badge()
            entry.bind('<FocusIn>', focus_in)
            entry.bind('<FocusOut>', focus_out)
            var.trace_add('write', var_trace)
            entry.insert(0, ph); entry.configure(foreground='#adb5bd')
            self._set_badge(badge, False)
        _bind(self.old_entry, self.old_path, self.old_status, '请选择参考文件...')
        _bind(self.new_entry, self.new_path, self.new_status, '请选择待检文件...')
    def _on_status_click(self, event):
        """点击状态徽章：已导入→打开文件，未导入→提示"""
        widget = event.widget
        if widget == self.old_status:
            path = self.old_path.get()
            if not path:
                messagebox.showinfo("提示", "请先导入参考报告文件")
            else:
                try: os.startfile(path)
                except Exception as e: messagebox.showerror("错误", f"无法打开文件: {e}")
        elif widget == self.new_status:
            path = self.new_path.get()
            if not path:
                messagebox.showinfo("提示", "请先导入待检报告文件")
            else:
                try: os.startfile(path)
                except Exception as e: messagebox.showerror("错误", f"无法打开文件: {e}")
    def _raise_modal(self):
        w=getattr(self,'_active_modal',None)
        try:
            if w is not None and w.winfo_exists(): w.lift(); w.focus_force(); return True
        except Exception: pass
        return False
    def _modal_call(self,fn):
        if self._modal_busy:
            self._raise_modal(); return
        self._modal_busy=True
        try:
            fn()
        finally:
            self._active_modal=None; self._modal_busy=False
    def browse(self,var):
        p=filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")])
        if p: var.set(p)
    def toggle_topmost(self): self.root.attributes('-topmost',self.topmost.get())
    def open_check_options(self):
        if self.check_project: messagebox.showinfo("提示","当前为规则检查模式，无需设置常规差异检测"); return
        dlg=CheckOptionsDialog(self.root,self.check_options,self.com_verify.get())
        if dlg.result is not None:
            self.check_options=dlg.result; self.com_verify.set(dlg.com_on)
            en=sum(1 for v in self.check_options.values() if v); self.log(f"检测设置已更新：{en}/{len(self.check_options)} 项已启用，高级审核{'开启' if dlg.com_on else '关闭'}")
    def open_project_dialog(self):
        if self._modal_busy: self._raise_modal(); return
        if not self.old_path.get(): messagebox.showwarning("提示","请先选择参考报告文件路径"); return
        self._modal_busy=True
        try:
            dlg=CheckProjectDialog(self.root,self.old_path.get(),self.new_path.get(),self.check_project); self._active_modal=dlg
            if dlg.result is not None:
                self.check_project=dlg.result; self._set_start_btn_text(); self.settings_btn.configure(state='disabled'); self.project_btn.configure(text="高级审核\n规则配置"); self.log(f"检查项目已加载：{self.check_project.project_name}，包含 {len(self.check_project.rules)} 条规则")
        finally:
            self._active_modal=None; self._modal_busy=False
    def clear_check_project(self):
        self.check_project=None; self.start_btn.configure(text="常规差异\n逐行对比",bootstyle=INFO,command=self.start_compare,width=10); self.settings_btn.configure(state='normal'); self.config_btn.configure(text="高级\n审核",bootstyle="outline-primary",width=7,command=self.load_check_project); self.project_btn.configure(text="高级审核\n规则配置",command=self.open_project_dialog); self.tree.delete(*self.tree.get_children()); self.detail.delete('1.0','end'); self.diff_items=[]; self.result_data=None; self.log("已退出规则检查模式，恢复常规差异对比")
    def load_check_project(self):
        if self._modal_busy: self._raise_modal(); return
        self._modal_busy=True
        try:
            dlg=ProjectListDialog(self.root,PROGRAM_DIR); self._active_modal=dlg
            if dlg.result:
                with open(dlg.result,'r',encoding='utf-8') as f: data=json.load(f)
                self.check_project=CheckProject.from_dict(data); self._set_start_btn_text(); self.settings_btn.configure(state='disabled'); self.config_btn.configure(text="退出配置\n高级审核",bootstyle="warning-outline",width=8,command=self.clear_check_project); self.log(f"检查项目已加载：{self.check_project.project_name}，包含 {len(self.check_project.rules)} 条规则")
        except Exception as e: messagebox.showerror("错误",f"打开配置列表失败：{str(e)}")
        finally:
            self._active_modal=None; self._modal_busy=False
    def _log_tag(self,msg):
        """根据日志内容返回着色 tag"""
        if '✗' in msg or '✕' in msg: return 'log_bad'
        if any(k in msg for k in ('异常','失败','错误','无法','跳过复核')): return 'log_bad'
        if '✓' in msg or '✔' in msg: return 'log_ok'
        if any(k in msg for k in ('高级审核完成','解析完成','检查总计耗时','对比阶段耗时')): return 'log_bold'
        return None

    def _insert_summary_line(self,msg):
        """检查完成统计行分段着色：需人工复核=红粗，已豁免=蓝粗"""
        import re as _re
        m=_re.match(r'^(检查完成(?:\(异常中断\))?[：:]\s*)',msg)
        prefix=m.group(1) if m else ''
        if prefix: self.log_text.insert('end',prefix,'log_bold')
        rest=msg[len(prefix):]
        # 按关键统计项边界切分，保留中间的逗号分隔符
        segs=_re.split(r'(?=需人工复核|已豁免|插件)',rest)
        for s in segs:
            if not s: continue
            if s.startswith('需人工复核'): self.log_text.insert('end',s,'log_red_bold')
            elif s.startswith('已豁免'): self.log_text.insert('end',s,'log_blue_bold')
            else: self.log_text.insert('end',s,'log_bold')
        self.log_text.insert('end','\n')

    def log(self,msg):
        def _log():
            ls=self.log_text.index('end-2l'); last=self.log_text.get(ls,'end-1c')
            is_hb=('正在加载' in msg or '已耗时' in msg)
            if ('正在加载' in last or '已耗时' in last) and is_hb: self.log_text.delete(ls,'end-1c')
            if is_hb:
                self.log_text.insert('end',msg+'\n','log_hb')  # 心跳刷新行：灰色，不加时间戳
            elif msg.startswith('检查完成'):
                self.log_text.insert('end',time.strftime('%H:%M:%S')+' ','ts')
                self._insert_summary_line(msg)
            else:
                self.log_text.insert('end',time.strftime('%H:%M:%S')+' ','ts')
                tag=self._log_tag(msg)
                self.log_text.insert('end',msg+'\n',tag) if tag else self.log_text.insert('end',msg+'\n')
            self.log_text.see('end')
        self.root.after(0,_log)
    def update_progress(self,val,stat=""): self.root.after(0,lambda:self.progress.configure(value=val))
    def set_progress_mode(self,mode):
        def _s():
            self.progress.configure(mode=mode)
            if mode=='indeterminate': self.progress.start(50)
            else: self.progress.stop()
        self.root.after(0,_s)
    def _excel_open_paths(self):
        try:
            app=win32com.client.GetActiveObject("Excel.Application")
            return {normalize_path(w.FullName) for w in app.Workbooks}
        except Exception: return None

    @staticmethod
    def _is_file_locked(fp):
        """检测文件是否被其他进程锁定（如 Excel）"""
        try:
            f=open(fp,'ab'); f.close(); return False
        except (IOError,PermissionError): return True

    def _prep_advanced_audit(self,old,new):
        # 第一步：确保两份报告已在 Excel 打开（os.startfile）；第二步：COM 通道确认窗。False=中止
        miss=[]
        paths=self._excel_open_paths()
        if paths is None:
            miss=[old,new]
        else:
            if normalize_path(old) not in paths: miss.append(old)
            if normalize_path(new) not in paths: miss.append(new)
        # 文件锁兜底：COM 没找到（可能返回了其他 Excel 实例），但文件已被锁定说明确实打开了
        if miss and self._is_file_locked(old) and self._is_file_locked(new):
            self.log("检测到文件已在 Excel 中打开，跳过开启确认")
            miss=[]
        if miss:
            choice=messagebox.askyesnocancel("高级审核","高级审核需通过 Excel COM 调用显示层，报告文件尚未打开。\n是＝立即打开报告并继续；否＝仅做常规对比（不审核）；取消＝中止本次检查。")
            if choice is None: return False
            if choice:
                for fp in miss:
                    try: os.startfile(fp)
                    except Exception as e: messagebox.showerror("错误",f"无法打开文件：{e}"); return False
                self.log("正在打开报告文件，请等待 Excel 加载...")
                ok_all=False; deadline=time.time()+40
                while time.time()<deadline:
                    time.sleep(1); ps=self._excel_open_paths()
                    if ps is not None and normalize_path(old) in ps and normalize_path(new) in ps: ok_all=True; break
                    # 文件锁兜底：os.startfile 可能在另一个实例中打开
                    if self._is_file_locked(old) and self._is_file_locked(new):
                        self.log("文件已在 Excel 中打开（其他实例），跳过等待")
                        ok_all=True; break
                if not ok_all and not messagebox.askyesno("高级审核","报告打开超时，是否仍继续高级审核？"): return False
            else:
                self.com_verify.set(False); self.log("已跳过高级审核，仅执行常规对比"); return True
        dlg=ComCheckDialog(self.root)
        if dlg.result is not True: self.log("已取消高级审核确认，本次检查中止"); return False
        return True
    def start_compare(self):
        old=self.old_path.get(); new=self.new_path.get()
        if not old or not new: messagebox.showerror("错误","请选择两个文件"); return
        if not os.path.isfile(old) or not os.path.isfile(new): messagebox.showerror("错误","文件不存在"); return
        if not old.lower().endswith('.xlsx') or not new.lower().endswith('.xlsx'): messagebox.showerror("错误","仅支持 .xlsx 格式文件"); return
        if self.com_verify.get() and not self._prep_advanced_audit(old,new): return
        self.stop_event.clear(); self.start_btn.configure(text="停止检查",bootstyle="danger",command=self.stop_compare)
        for b in (self.project_btn,self.config_btn,self.settings_btn): b.configure(state='disabled')
        self.tree.delete(*self.tree.get_children()); self.detail.delete('1.0','end'); self.diff_items=[]; self.log_text.insert('end',"="*10+" 开始新的检查 "+"="*10+'\n','log_blue_bold'); self.log_text.see('end')
        current_opts=dict(self.check_options); pm=self.plugin_manager; cp=self.check_project; tol=self.color_tolerance.get(); do_com=self.com_verify.get()
        def worker():
            comparer=None; compare_t0=time.time()
            try:
                comparer=OpenpyxlComparer(old,new,self.log,self.update_progress,check_options=current_opts,plugin_manager=pm,progress_mode_fn=self.set_progress_mode,check_project=cp,stop_event=self.stop_event,mode='diff',color_tolerance=tol)
                comparer.run()
                if do_com:
                    self.log("启动 Excel COM 显示层数据采集...")
                    try:
                        verifier=ExcelCOMVerifier(old,new,self.log,progress_fn=self.update_progress,progress_mode_fn=self.set_progress_mode)
                        # 传递字体等价表（从 comparer 的 cache 里拿）
                        old_fe = getattr(getattr(comparer, 'old_cache', None), 'font_equiv', None)
                        new_fe = getattr(getattr(comparer, 'new_cache', None), 'font_equiv', None)
                        verifier.collect_style_data(comparer.diffs, old_font_equiv=old_fe, new_font_equiv=new_fe)
                    except Exception as e:
                        self.log(f"COM数据采集异常: {e}，退回 openpyxl 数据")
                # COM 采集完成后，再执行规则引擎过滤（样式类优先用 COM 数据判定）
                if comparer.check_project:
                    self.update_progress(92, "执行进阶规则过滤...")
                    comparer._apply_rule_filter(comparer.diffs, comparer.old_wb_ref, comparer.new_wb_ref)
                # debug 导出全量 diff 数据
                try:
                    import json as _json
                    _dp = os.path.join(os.path.dirname(new), f"debug_dump_{time.strftime('%Y%m%d_%H%M%S')}.json")
                    _owb = getattr(comparer, 'old_wb_ref', None)
                    _nwb = getattr(comparer, 'new_wb_ref', None)
                    _dd = []
                    for d in comparer.diffs:
                        e = {k: v for k, v in d.items() if k != 'com_style'}
                        if d.get('com_style'):
                            e['com_old'] = d['com_style'].get('old')
                            e['com_new'] = d['com_style'].get('new')
                        # 读单元格实际值（openpyxl原始值）；shift 配对场景 old 侧读配对旧格，与比对逻辑一致
                        try:
                            if _owb and _nwb and d.get('sheet') in _owb.sheetnames:
                                from openpyxl.utils import column_index_from_string as _cifs
                                _m = getattr(comparer, '_last_shift_old_map', {}) if d.get('type')=='单元格删除' else getattr(comparer, '_last_shift_new_map', {})
                                _hit = _m.get((d.get('sheet'), d.get('address')))
                                _osh, _oad = (_hit[0][1], _hit[0][2]) if _hit else (d.get('sheet'), d.get('address'))
                                for _tag, _wb, _sh, _ad in (('old', _owb, _osh, _oad), ('new', _nwb, d.get('sheet'), d.get('address'))):
                                    _cs = ''.join(ch for ch in _ad if ch.isalpha())
                                    _rs = ''.join(ch for ch in _ad if ch.isdigit())
                                    if _cs and _rs and _sh in _wb.sheetnames:
                                        _r, _c = int(_rs), _cifs(_cs)
                                        e[f'{_tag}_value'] = str(_wb[_sh].cell(_r, _c).value)
                                        e[f'{_tag}_numfmt'] = _wb[_sh].cell(_r, _c).number_format
                                if _hit: e['shift_paired_old'] = f"{_osh}!{_oad}"
                        except Exception: pass
                        _dd.append(e)
                    with open(_dp, 'w', encoding='utf-8') as _f:
                        _json.dump(_dd, _f, ensure_ascii=False, indent=1, default=str)
                    self.log(f"调试数据已导出: {_dp}")
                except Exception as _e:
                    self.log(f"调试导出失败: {_e}")
                self.log(f"检查总计耗时 {_fmt_duration(time.time()-compare_t0)}")
                self.old_sheet_order=comparer.sheet_order; self.result_data=(comparer.diffs,comparer.sheet_diffs,comparer.stats); self.root.after(0,self.populate_tree)
            except KeyboardInterrupt:
                if comparer: self.old_sheet_order=comparer.sheet_order; self.result_data=(comparer.diffs,comparer.sheet_diffs,comparer.stats)
                else: self.result_data=([],[],{'total_cells':0,'diff_cells':0,'added_sheets':[],'removed_sheets':[],'images_diff':0})
                self.root.after(0,self.populate_tree)
            except Exception as e:
                import traceback; self.root.after(0,lambda:messagebox.showerror("对比失败",f"{e}\n\n{traceback.format_exc()}"))
            finally: self.root.after(0,self.on_comparison_finished)
        threading.Thread(target=worker,daemon=True).start()
    def _set_start_btn_text(self):
        if self.check_project:
            name=self.check_project.project_name or ''
            dw=sum(2 if ord(ch)>0x2e7f else 1 for ch in name)
            w=max(10,min(22,dw+6))
            self.start_btn.configure(text=f"[{name}]",width=w)
        else: self.start_btn.configure(text="常规差异\n逐行对比",width=10)
    def on_comparison_finished(self):
        self._set_start_btn_text(); self.start_btn.configure(bootstyle=INFO,command=self.start_compare)
        self.start_btn.configure(state='normal'); self.project_btn.configure(state='normal'); self.config_btn.configure(state='normal')
        if not self.check_project: self.settings_btn.configure(state='normal')
        if self.stop_event.is_set(): self.log("检查已停止")
        self.progress['value']=100
    def _set_twisty(self,iid,is_open):
        vals=list(self.tree.item(iid,'values') or ('','',''))
        while len(vals)<3: vals.append('')
        vals[0]='[-]' if is_open else '[+]'
        self.tree.item(iid,values=vals)
    def _toggle_all_nodes(self):
        kids=self.tree.get_children('')
        if any(self.tree.item(k,'open') for k in kids):
            for k in kids:
                self.tree.item(k,open=False); self._set_twisty(k,False)
            self.tree.heading('action',text='展开')
        else:
            for k in kids:
                self.tree.item(k,open=True); self._set_twisty(k,True)
            self.tree.heading('action',text='收起')
    def populate_tree(self):
        self._pt_err=None
        try:
            self._populate_tree_inner()
        except Exception as e:
            self._pt_err=e
            import traceback; tb=traceback.format_exc()
            self.log(f"[populate_tree 异常] {e}")
            self.log(tb)
        finally:
            if self._pt_err:
                try:
                    diffs,_,_=self.result_data
                    nr=sum(1 for d in diffs if not d.get('rule_pass') and d.get('sheet')!='🔍 数据检查')
                    rp=sum(1 for d in diffs if d.get('rule_pass'))
                    pl=sum(1 for d in diffs if d.get('sheet')=='🔍 数据检查')
                    style_rp=sum(1 for d in diffs if d.get('rule_pass') and d.get('rule_name')=='样式显示一致')
                    rule_rp=rp-style_rp
                    rp_tag = f"（规则{rule_rp} + 样式{style_rp}）" if rp else ''
                    self.log(f"检查完成(异常中断): 需人工复核 {nr} 项，已豁免 {rp} 项{rp_tag}，插件 {pl} 项")
                except Exception: pass
    def _populate_tree_inner(self):
        self.tree.heading('action',text='收起')
        diffs,sheet_diffs,stats=self.result_data; normal=[]; rule_pass=[]; plugin=[]
        for d in diffs:
            if d['sheet']=='🔍 数据检查': plugin.append(d)
            elif d.get('rule_pass'): rule_pass.append(d)
            else: normal.append(d)
        if sheet_diffs:
            sn=self.tree.insert('','end',text=f'📋 Sheet 结构差异（{len(sheet_diffs)}）',open=True,values=('[-]','',''),tags=('sheet','twisty'))
            for sd in sheet_diffs: node=self.tree.insert(sn,'end',text=sd['desc'],values=('[-]',sd['name'],sd['type'])); self.diff_items.append((node,{'type':'sheet_struct','data':sd}))
        if plugin:
            pn=self.tree.insert('','end',text=f'🔍 数据检查结果（{len(plugin)}）',open=True,values=('[-]','',''),tags=('sheet','twisty'))
            for d in plugin: node=self.tree.insert(pn,'end',text=d['desc'][:80],values=('[-]',d['address'],d['type'])); self.diff_items.append((node,{'type':'cell','data':d}))
        if normal:
            pn=self.tree.insert('','end',text=f'⚠️ 需人工复核（未豁免）（{len(normal)}）',open=True,values=('[-]','',''),tags=('warning_sheet',)); dmap={}
            for d in normal: dmap.setdefault(d['sheet'],[]).append(d)
            for sname in self.old_sheet_order:
                if sname in dmap:
                    merged=self._merge_cell_ranges(dmap[sname]); n_disp=sum(d.get("_merged_count",1) for d in merged)
                    sn=self.tree.insert(pn,'end',text=f"📄 {sname}（{n_disp}）",open=True,values=('[-]','',''),tags=('sheet','twisty'))
                    for d in merged:
                        atag=('advanced_check',) if d.get('advanced_check') else (('rule_fail',) if d.get('rule_name') and not d.get('rule_pass') else ())
                        node=self.tree.insert(sn,'end',text=(d.get('short_desc') or d['desc'])[:120],values=('[-]',d['address'],d['type']),tags=atag); self.diff_items.append((node,{'type':'cell','data':d}))
            for sname in dmap:
                if sname not in self.old_sheet_order:
                    merged=self._merge_cell_ranges(dmap[sname]); n_disp=sum(d.get("_merged_count",1) for d in merged)
                    sn=self.tree.insert(pn,'end',text=f"📄 {sname}（{n_disp}）",open=True,values=('[-]','',''),tags=('sheet','twisty'))
                    for d in merged:
                        atag=('advanced_check',) if d.get('advanced_check') else (('rule_fail',) if d.get('rule_name') and not d.get('rule_pass') else ())
                        node=self.tree.insert(sn,'end',text=(d.get('short_desc') or d['desc'])[:120],values=('[-]',d['address'],d['type']),tags=atag); self.diff_items.append((node,{'type':'cell','data':d}))
        if rule_pass:
            total_pass=len(rule_pass)
            pn=self.tree.insert('','end',text=f'✅ 已豁免（可展开）（{total_pass}）',open=False,values=('[+]','',''),tags=('sheet','twisty'))
            style_pass=[d for d in rule_pass if d.get('rule_name')=='样式显示一致']
            rule_match_pass=[d for d in rule_pass if d.get('rule_name')!='样式显示一致']
            for group_name, group_data in [('规则匹配通过', rule_match_pass), ('样式显示一致', style_pass)]:
                if not group_data: continue
                gn=self.tree.insert(pn,'end',text=f'{group_name}（{len(group_data)}）',open=False,values=('[+]','',''),tags=('sheet',))
                dmap={}
                for d in group_data: dmap.setdefault(d['sheet'],[]).append(d)
                for sname in self.old_sheet_order:
                    if sname in dmap:
                        merged=self._merge_cell_ranges(dmap[sname]); n_disp=sum(d.get('_merged_count',1) for d in merged)
                        sn=self.tree.insert(gn,'end',text=f"📄 {sname}（{n_disp}）",open=True,values=('[-]','',''),tags=('sheet','twisty'))
                        for d in merged: node=self.tree.insert(sn,'end',text=(d.get('short_desc') or d['desc'])[:100],values=('[-]',d['address'],d['type'])); self.diff_items.append((node,{'type':'cell','data':d}))
                for sname in dmap:
                    if sname not in self.old_sheet_order:
                        merged=self._merge_cell_ranges(dmap[sname]); n_disp=sum(d.get('_merged_count',1) for d in merged)
                        sn=self.tree.insert(gn,'end',text=f"📄 {sname}（{n_disp}）",open=True,values=('[-]','',''),tags=('sheet','twisty'))
                        for d in merged: node=self.tree.insert(sn,'end',text=(d.get('short_desc') or d['desc'])[:100],values=('[-]',d['address'],d['type'])); self.diff_items.append((node,{'type':'cell','data':d}))

        style_rp=sum(1 for d in rule_pass if d.get('rule_name')=='样式显示一致')
        rule_rp=len(rule_pass)-style_rp
        rp_tag = f"（规则{rule_rp} + 样式{style_rp}）" if rule_pass else ''
        self.log(f"检查完成：需人工复核 {len(normal)} 项，已豁免 {len(rule_pass)} 项{rp_tag}，插件 {len(plugin)} 项")
    @staticmethod
    def _merge_cell_ranges(items):
        """将同行同规则同类型的相邻单元格合并为区域表达，如 D29,E29,F29,G29 → D29:G29。
        返回排序后的合并项列表，count 为原始条目数。"""
        groups={}  # (rule_name,type,row) → {cols, items, sheet, desc, ...}
        flat=[]
        for item in items:
            addr=item.get('address','')
            if ':' in addr: flat.append(item); continue
            col_s=''.join(ch for ch in addr if ch.isalpha()); row_s=''.join(ch for ch in addr if ch.isdigit())
            if not col_s or not row_s: flat.append(item); continue
            col=column_index_from_string(col_s); row=int(row_s)
            key=(item.get('rule_name',''),item.get('type',''),row)
            if key not in groups: groups[key]={'cols':[],'items':[],'sheet':item['sheet'],'desc':item['desc'],'rule_expect':item.get('rule_expect',''),'rule_pass':item.get('rule_pass',False),'com_confirmed':False,'count':0}
            groups[key]['cols'].append(col); groups[key]['items'].append(item); groups[key]['count']+=1
            if item.get('com_confirmed'): groups[key]['com_confirmed']=True
        result=list(flat)
        for key,data in groups.items():
            cols=sorted(set(data['cols'])); ranges=[]; cur=[cols[0]]
            for c in cols[1:]:
                if c==cur[-1]+1: cur.append(c)
                else: ranges.append(cur); cur=[c]
            ranges.append(cur)
            for cr in ranges:
                if len(cr)==1:
                    for it in data['items']:
                        if ''.join(ch for ch in it.get('address','') if ch.isalpha())==get_column_letter(cr[0]): result.append(it); break
                else:
                    fc=get_column_letter(cr[0]); lc=get_column_letter(cr[-1]); row=key[2]
                    merged=dict(data['items'][0]); merged['address']=f"{fc}{row}:{lc}{row}"; merged['desc']=f"{data['desc']} ({len(cr)}项)"; merged['_merged_count']=len(cr)
                    result.append(merged)
        def _sk(it):
            a=it.get('address','')
            if ':' in a: a=a.split(':')[0]
            cs=''.join(ch for ch in a if ch.isalpha()); rs=''.join(ch for ch in a if ch.isdigit())
            return (99999,0) if not cs or not rs else (int(rs),column_index_from_string(cs))
        result.sort(key=_sk)
        return result

    def on_tree_select(self,event):
        sel=self.tree.selection()
        if not sel: return
        target=next((d for n,d in self.diff_items if n==sel[0]),None)
        if not target: return
        if target['type'] in ('cell','sheet_struct'):
            d=target['data']
            lines=[f"描述: {d['desc']}"]
            if d.get('rule_name'):
                lines.append(f"规则: {d['rule_name']}")
                diff_desc=d.get('rule_diff_desc')
                if diff_desc:
                    lines.append("结果:")
                    for line in diff_desc.split('\n'):
                        lines.append(f"  {line}")
                else:
                    lines.append("结果: 无差异描述")
            self.detail.delete('1.0','end'); self.detail.insert('1.0','\n'.join(lines))
    def on_tree_click(self,event):
        if self.tree.identify_region(event.x,event.y)!='cell': return
        if self.tree.identify_column(event.x)!='#1': return
        iid=self.tree.identify_row(event.y)
        if not iid: return
        if self.tree.get_children(iid):
            # 父级行：[-]/[+] 单击切换展开/收起
            is_open=not self.tree.item(iid,'open')
            self.tree.item(iid,open=is_open); self._set_twisty(iid,is_open); self.tree.see(iid)
        else:
            # 叶子行：保持原行为（收起所属父级），同步父级标记为 [+]
            parent=self.tree.parent(iid)
            if parent:
                self.tree.item(parent,open=False); self._set_twisty(parent,False); self.tree.see(parent)
    def jump_to_excel(self,file_path,sheet_name,cell_addr):
        # 返回 'opened'(已跳转)/False；不再静默启动 Excel（由上层询问后 os.startfile 打开）
        try:
            try: excel=win32com.client.GetActiveObject("Excel.Application")
            except Exception: return False,"Excel 未运行"
            wb=None
            for w in excel.Workbooks:
                if normalize_path(w.FullName)==normalize_path(file_path): wb=w; break
            # 文件锁兜底：GetActiveObject 返回了其他 Excel 实例，用 GetObject 直接获取 workbook
            if wb is None and self._is_file_locked(file_path):
                try: wb=win32com.client.GetObject(file_path)
                except Exception: pass
            if wb is None: return False,f"文件未在 Excel 中打开：{file_path}"
            wb.Activate()
            app=wb.Application
            try: ws=wb.Worksheets(sheet_name)
            except Exception as e: return False,f"工作表不存在：{sheet_name}（{e}）"
            ws.Activate()
            # 提取第一个单元格用于滚动定位（兼容范围地址如 D29:G29）
            first_addr = cell_addr.split(':')[0] if ':' in cell_addr else cell_addr
            col=''.join(ch for ch in first_addr if ch.isalpha())
            row=''.join(ch for ch in first_addr if ch.isdigit())
            if not col or not row: return False,f"无效单元格地址：{cell_addr}"
            app.ActiveWindow.ScrollRow=int(row); app.ActiveWindow.ScrollColumn=column_index_from_string(col); ws.Range(cell_addr).Select(); return 'opened',None
        except Exception as e: return False,str(e)
    def on_tree_double_click(self,event):
        sel=self.tree.selection()
        if not sel: return
        info=next((d for n,d in self.diff_items if n==sel[0]),None)
        if not info or info['type']!='cell': return
        d=info['data']; sheet=d.get('sheet'); addr=d.get('address')
        if not sheet or not addr: return
        if not re.match(r'^[A-Za-z]+[0-9]+(:[A-Za-z]+[0-9]+)?$',addr): self.log(f"跳过跳转：无效地址 {addr}"); return
        def _ensure(path,label):
            ok,err=self.jump_to_excel(path,sheet,addr)
            if ok=='opened': return True
            if messagebox.askyesno("跳转",f"{label}未打开：{err}\n是否立即打开该报告以便跳转？"):
                try:
                    os.startfile(path)
                    for _ in range(30):
                        time.sleep(1)
                        if self.jump_to_excel(path,sheet,addr)[0]=='opened':
                            self.log(f"已跳转到 {sheet}!{addr}（{label}）"); return True
                    messagebox.showwarning("跳转失败",f"{label}打开超时，请确认 Excel 已加载后重试")
                except Exception as e: messagebox.showerror("跳转失败",f"无法打开文件：{e}")
            return False
        _ensure(self.old_path.get(),"参考报告"); _ensure(self.new_path.get(),"待检报告")
    def stop_compare(self):
        self.stop_event.set(); self.log("正在停止检查..."); self.start_btn.configure(state='disabled')

if __name__ == "__main__":
    app=tb.Window(themename="flatly")
    DiffViewer(app)
    app.mainloop()
